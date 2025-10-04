####################
# Core application #
####################

import streamlit as st
import pandas as pd
import datetime
import io
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import pandas.api.types as ptypes
import uuid
import copy
from typing import Optional, List, Dict, Any, Tuple

from app_const import *
from app_utils import *
from carnet_addr import *
import tracer
import sql_api as sql 
import gsheet_api as gs
import dropbox_api as dp
import sync_worker as wk
import undo

####################
# Contexte Manager #
####################

def charger_contexte_depuis_sql():

    if "df" not in st.session_state:

        # Récupération df, meta, ca à partir de la base SQLite
        df, meta, ca = sql.charger_contexte()

        # Mise à jour wb, fn, fp
        try:
            wb = None
            fn  = meta["fn"]
            fp  = meta["fp"]
            if not (fp is None or str(fp).strip() == ""):
                wb = dp.download_excel_from_dropbox(fp)
        except Exception as e:
            print(f"Erreur au chargement du modèle Excel depuis DropBox : {e}")

        # Mise à jour paramètres
        try:
            st.session_state.MARGE = to_timedelta(meta["MARGE"], default=MARGE)
            if meta["MARGE"] is None:
                sql.sauvegarder_param("MARGE")
            st.session_state.DUREE_REPAS = to_timedelta(meta["DUREE_REPAS"], default=DUREE_REPAS)
            if meta["DUREE_REPAS"] is None:
                sql.sauvegarder_param("DUREE_REPAS")
            st.session_state.DUREE_CAFE = to_timedelta(meta["DUREE_CAFE"], default=DUREE_CAFE)
            if meta["DUREE_CAFE"] is None:
                sql.sauvegarder_param("DUREE_CAFE")

            st.session_state.itineraire_app = meta["itineraire_app"]
            st.session_state.city_default = meta["city_default"]
            st.session_state.traiter_pauses = str(meta["traiter_pauses"]).strip().lower() == "true"
        except Exception as e:
            print(f"Erreur au chargement des paramètres depuis SQLite : {e}")

        # Mise à jour période de programmation
        try:
            val = meta["periode_a_programmer_debut"]
            if val is not None and str(val).strip() != "":
                st.session_state.periode_a_programmer_debut = datetime.date.fromisoformat(val.split(" ")[0])
            val = meta["periode_a_programmer_fin"]
            if val is not None and str(val).strip() != "":
                st.session_state.periode_a_programmer_fin = datetime.date.fromisoformat(val.split(" ")[0])
        except Exception as e:
            print(f"Erreur au chargement de la période de programmation depuis SQLite : {e}")
        
        # Si période programmation absente des meta rattrapage via init standard à partir des activités programmées du contexte que l'on vient de charger  
        if "periode_a_programmer_debut" not in st.session_state or "periode_a_programmer_fin" not in st.session_state:
            initialiser_periode_programmation(df) 
            sql.sauvegarder_param("periode_a_programmer_debut")
            sql.sauvegarder_param("periode_a_programmer_fin")
    
        st.session_state["push_periode_programmation_modele_values"] = True 

        df = _nettoyer_donnees(df, fn)
        initialiser_etat_contexte(df, wb, fn, fp, ca)
        undo.init(verify=False)
        maj_contexte(maj_donnees_calculees=True, maj_options_date=True) 

        st.session_state.activites_programmees = st.session_state.activites_programmees.drop(columns="__options_date", errors="ignore")
        st.session_state.activites_non_programmees = st.session_state.activites_non_programmees.drop(columns="__options_date", errors="ignore")
        
        selection = st.session_state.activites_non_programmees.index[0] if len(st.session_state.activites_non_programmees) > 0 else None
        demander_selection("activites_non_programmees", selection, deselect="activites_programmees")
        st.session_state.menu_activites = {
            "menu": "menu_activites_non_programmees",
            "index_df": selection
        }

        wk.enqueue_save_full(df, meta, ca)

def charger_contexte_depuis_gsheet():

    if "gsheets" in st.session_state:
        
        curseur_attente()

        try:

            # Récupération df, meta, ca à partir de la base SQLite
            df, meta, ca = gs.charger_contexte()
        
            # Mise à jour wb, fn, fp
            try:
                wb = None
                fn  = meta["fn"]
                fp  = meta["fp"]
                if not (fp is None or str(fp).strip() == ""):
                    wb = dp.download_excel_from_dropbox(fp)
            except Exception as e:
                print(f"Erreur au chargement du modèle Excel depuis DropBox : {e}")

            # Mise à jour paramètres
            try:
                st.session_state.MARGE = to_timedelta(meta["MARGE"], default=MARGE)
                if meta["MARGE"] is None:
                    gs.sauvegarder_param("MARGE")
                st.session_state.DUREE_REPAS = to_timedelta(meta["DUREE_REPAS"], default=DUREE_REPAS)
                if meta["DUREE_REPAS"] is None:
                    gs.sauvegarder_param("DUREE_REPAS")
                st.session_state.DUREE_CAFE = to_timedelta(meta["DUREE_CAFE"], default=DUREE_CAFE)
                if meta["DUREE_CAFE"] is None:
                    gs.sauvegarder_param("DUREE_CAFE")

                st.session_state.itineraire_app = meta["itineraire_app"]
                st.session_state.city_default = meta["city_default"]
                st.session_state.traiter_pauses = str(meta["traiter_pauses"]).strip().lower() == "true"
            except Exception as e:
                print(f"Erreur au chargement des paramètres depuis SQLite : {e}")

            # Mise à jour période de programmation
            try:
                val = meta["periode_a_programmer_debut"]
                if val is not None and str(val).strip() != "":
                    st.session_state.periode_a_programmer_debut = datetime.date.fromisoformat(val.split(" ")[0])
                val = meta["periode_a_programmer_fin"]
                if val is not None and str(val).strip() != "":
                    st.session_state.periode_a_programmer_fin = datetime.date.fromisoformat(val.split(" ")[0])
            except Exception as e:
                print(f"Erreur au chargement de la période de programmation depuis SQLite : {e}")
    
            # Si période programmation absente des meta rattrapage via init standard à partir des activités programmées du contexte que l'on vient de charger  
            if "periode_a_programmer_debut" not in st.session_state or "periode_a_programmer_fin" not in st.session_state:
                initialiser_periode_programmation(df) 
                gs.sauvegarder_param("periode_a_programmer_debut")
                gs.sauvegarder_param("periode_a_programmer_fin")
    
            st.session_state["push_periode_programmation_modele_values"] = True 

            initialiser_dtypes(df)
            df = _nettoyer_donnees(df, fn)
            df = add_persistent_uuid(df)
            df = add_hyperliens(df)
            ca = nettoyer_ca(ca)
            initialiser_etat_contexte(df, wb, fn, fp, ca)
            undo.init(verify=False)
            maj_contexte(maj_donnees_calculees=True, maj_options_date=True) 
            selection = st.session_state.activites_non_programmees.index[0] if len(st.session_state.activites_non_programmees) > 0 else None
            demander_selection("activites_non_programmees", selection, deselect="activites_programmees")
            st.session_state.menu_activites = {
                "menu": "menu_activites_non_programmees",
                "index_df": selection
            }
            curseur_normal()
        
        except Exception as e:
            print(f"Erreur au chargement des données depuis la Google Sheets : {e}")
            curseur_normal

##########################
# Fonctions applicatives #
##########################

# Indique si une row est une activité programmée
def est_activite_programmee(row):
    if isinstance(row, pd.DataFrame):
        row=row.iloc[0] # sinon and plante car pd.isna et pd.notna renvoient des series
    return (est_float_valide(row["Date"]) and 
             pd.notna(row["Debut"]) and 
             pd.notna(row["Duree"]) and 
             pd.notna(row["Activite"]))

# Renvoie le dataframe des activités programmées
def get_activites_programmees(df):
    return df[
        df["Date"].apply(est_float_valide) & 
        df["Debut"].notna() & 
        df["Duree"].notna() &
        df["Activite"].notna()
    ].sort_values(by=["Date", "Debut_dt"], ascending=[True, True])

# Indique si une row est une activité non programmée
def est_activite_non_programmee(row):
    if isinstance(row, pd.DataFrame):
        row=row.iloc[0] # sinon and plante car pd.isna et pd.notna renvoient des series
    return (pd.isna(row["Date"]) and 
             pd.notna(row["Debut"]) and 
             pd.notna(row["Duree"]) and 
             pd.notna(row["Activite"]))

# Renvoie le dataframe des activités non programmées
def get_activites_non_programmees(df):
    return df[df["Date"].isna() & 
              df["Debut"].notna() & 
              df["Duree"].notna() &
              df["Activite"].notna()
    ].sort_values(by=["Date", "Debut_dt"], ascending=[True, True])

# Indique si une activité donnée par son descripteur dans le df est réservée
def est_activite_reserve(ligne_df):
    return str(ligne_df["Reserve"]).strip().lower() == "oui"

# Ajout d'une nouvelle activité à la bd contexte
# @chrono
def ajouter_activite(idx=None, nom=None, jour=None, debut=None, duree=None, lieu=None, relache=None, hyperlien=None):
    def get_nom_nouvelle_activite(df):
        noms_existants = df["Activite"].dropna().astype(str).str.strip().tolist()
        while True:
            st.session_state.compteur_activite += 1
            nom_candidat = f"Activité {st.session_state.compteur_activite}"
            if nom_candidat not in noms_existants:
                return nom_candidat
            
    def get_next_free_index(df):
        existing = set(df.index)
        i = 0
        while i in existing:
            i += 1
        return i
    
    df = st.session_state.get("df", None)
    if df is None:
        return
    
    idx = get_next_free_index(df) if idx is None else idx
    nom = get_nom_nouvelle_activite(df) if nom is None else nom
    jour = pd.NA if jour is None else jour
    debut = "09h00" if debut is None else debut
    duree = "1h00" if duree is None else duree
    lieu = pd.NA if lieu is None else lieu
    relache = pd.NA if relache is None else relache
    hyperlien = f"https://www.festivaloffavignon.com/resultats-recherche?recherche={nom.replace(' ', '+')}" if hyperlien is None else hyperlien

    df.at[idx, "Date"] = jour
    df.at[idx, "Debut"] = debut
    df.at[idx, "Duree"] = duree
    df.at[idx, "Activite"] = nom
    df.at[idx, "Lieu"] = lieu
    df.at[idx, "Relache"] = relache
    df.at[idx, "Hyperlien"] = hyperlien
    add_persistent_uuid(df, idx)
    _maj_donnees_calculees_row(idx, full=False)

    row = df.loc[[idx]]

    if est_activite_programmee(row):
        st.session_state.activites_programmees = pd.concat([st.session_state.activites_programmees, row]).sort_values(by=["Date", "Debut"], ascending=[True, True])

        row = _creer_df_display_activites_programmees(row)
        st.session_state.activites_programmees_df_display = pd.concat([st.session_state.activites_programmees_df_display, row]).sort_values(by=["Date", "Début"], ascending=[True, True])
        st.session_state.activites_programmees_df_display_copy = st.session_state.activites_programmees_df_display.copy()

        maj_creneaux_disponibles()

    elif est_activite_non_programmee(row):
        st.session_state.activites_non_programmees = pd.concat([st.session_state.activites_non_programmees, row]).sort_values(by=["Date", "Debut"], ascending=[True, True])

        row = _creer_df_display_activites_non_programmees(row)
        st.session_state.activites_non_programmees_df_display = pd.concat([st.session_state.activites_non_programmees_df_display, row]).sort_values(by=["Date", "Début"], ascending=[True, True])
        st.session_state.activites_non_programmees_df_display_copy = st.session_state.activites_non_programmees_df_display.copy()

        maj_creneaux_disponibles()
    
    return idx

def _creer_df_display_activites_non_programmees(activites_non_programmees, maj_options_date=True):
    df_display = activites_non_programmees.copy()
    df_display["__index"] = df_display.index
    if "__options_date" not in df_display or maj_options_date:
        df_display["__options_date"] = calculer_options_date_activites_non_programmees(df_display) 
        df_display["__options_date"] = df_display["__options_date"].map(safe_json_dump)
    df_display["Date"] = df_display["Date"].apply(lambda x: str(int(x)) if pd.notna(x) and float(x).is_integer() else "")
    df_display["__desel_ver"] = st.session_state.activites_programmees_sel_request["desel"]["ver"] if "activites_programmees_sel_request" in st.session_state else 0
    df_display["__desel_id"] =  get_uuid(df_display, st.session_state.activites_programmees_sel_request["desel"]["id"]) if "activites_programmees_sel_request" in st.session_state else None
    df_display["__sel_ver"] = st.session_state.activites_programmees_sel_request["sel"]["ver"] if "activites_programmees_sel_request" in st.session_state else 0
    df_display["__sel_id"] =  get_uuid(df_display, st.session_state.activites_programmees_sel_request["sel"]["id"]) if "activites_programmees_sel_request" in st.session_state else None
    df_display["__sel_source"] = "api"
    df_display["__df_push_ver"] = 0
    df_display = _ensure_addr_cols(df_display)
    df_display.drop(columns=["Debut_dt", "Duree_dt"], inplace=True)
    df_display.rename(columns=RENOMMAGE_COLONNES, inplace=True)
    df_display = df_display.where(df_display.notna(), None)
    return df_display

def _creer_df_display_activites_programmees(activites_programmees, maj_options_date=True):
    df_display = activites_programmees.copy()
    df_display["__jour"] = df_display["Date"].apply(dateint_to_jour)
    df_display["__index"] = df_display.index
    if "__options_date" not in df_display or maj_options_date:
        df_display["__options_date"] = _calculer_options_date_activites_programmees(df_display) 
        df_display["__options_date"] = df_display["__options_date"].map(safe_json_dump)
    df_display["__non_reserve"] = df_display["Reserve"].astype(str).str.strip().str.lower() != "oui"
    df_display["Date"] = df_display["Date"].apply(lambda x: str(int(x)) if pd.notna(x) and float(x).is_integer() else "")
    df_display["__desel_ver"] = st.session_state.activites_programmees_sel_request["desel"]["ver"] if "activites_programmees_sel_request" in st.session_state else 0
    df_display["__desel_id"] =  get_uuid(df_display, st.session_state.activites_programmees_sel_request["desel"]["id"]) if "activites_programmees_sel_request" in st.session_state else None
    df_display["__sel_ver"] = st.session_state.activites_programmees_sel_request["sel"]["ver"] if "activites_programmees_sel_request" in st.session_state else 0
    df_display["__sel_id"] =  get_uuid(df_display, st.session_state.activites_programmees_sel_request["sel"]["id"]) if "activites_programmees_sel_request" in st.session_state else None
    df_display["__sel_source"] = "api"
    df_display["__df_push_ver"] = 0
    df_display = _ensure_addr_cols(df_display)
    df_display.drop(columns=["Debut_dt", "Duree_dt"], inplace=True)
    df_display.rename(columns=RENOMMAGE_COLONNES, inplace=True)
    df_display = df_display.where(df_display.notna(), None)
    return df_display

# Met à jour le contexte pour une activité dont la date de programmation passe de jour à None
# Si le param jour est à None on prend comme date de programmation antérieure la valeur présente dans l'activité 
# (ce qui suppose que cette valeur n'a pas été modifiée préalablement par un bd_modifier_cell). 
# @chrono
def _deprogrammer(idx, jour=None):
    
    if "df" not in st.session_state:
        return
    
    if idx not in st.session_state.df.index:
        return
    
    if "activites_programmees" not in st.session_state:
        return
    
    if "activites_non_programmees" not in st.session_state:
        return

    if idx in st.session_state.activites_programmees.index:
        row = st.session_state.activites_programmees.loc[[idx]]
        if jour is None:
            jour = row.loc[idx]["Date"]
        row.at[idx, "Date"] = None
        st.session_state.activites_programmees.drop(index=idx, inplace=True)
        st.session_state.activites_non_programmees = pd.concat([st.session_state.activites_non_programmees, row]).sort_values(by=["Debut_dt"], ascending=[True])

        row = st.session_state.activites_programmees_df_display.loc[[idx]]
        row.at[idx, "Date"] = ""
        row.drop(columns=["__jour", "__non_reserve"], inplace=True)
        st.session_state.activites_programmees_df_display.drop(index=idx, inplace=True)
        st.session_state.activites_non_programmees_df_display = pd.concat([st.session_state.activites_non_programmees_df_display, row]).sort_values(by=["Début"], ascending=[True])

        _maj_options_date(st.session_state.df, st.session_state.activites_programmees, st.session_state.activites_programmees_df_display, jour)
        _maj_options_date(st.session_state.df, st.session_state.activites_programmees, st.session_state.activites_non_programmees_df_display, jour)

        st.session_state.activites_programmees_df_display_copy = st.session_state.activites_programmees_df_display.copy()
        st.session_state.activites_non_programmees_df_display_copy = st.session_state.activites_non_programmees_df_display.copy()

        maj_creneaux_disponibles()

# Déprogrammation d'une activité programmée (si pause suppression, si activité ordinaire date à None)
def deprogrammer_activite_programmee(idx):
    df = st.session_state.df
    if est_pause(df.loc[idx]):
        supprimer_activite(idx)
    else:
        if idx not in st.session_state.df.index:
            return
        jour = st.session_state.df.loc[idx]["Date"]
        modifier_df_cell(st.session_state.df, idx, "Date", None)
        _deprogrammer(idx, jour)

# Met à jour les variables d'état relatives aux activités programmées
# @chrono
def maj_activites_programmees(maj_options_date=True):
    if st.session_state.get("df", None) is None:
        return  
    activites_programmees = get_activites_programmees(st.session_state.df)
    st.session_state.activites_programmees = activites_programmees
    df_display = _creer_df_display_activites_programmees(activites_programmees, maj_options_date)
    st.session_state.activites_programmees_df_display = df_display
    st.session_state.activites_programmees_df_display_copy = df_display.copy()

# Met à jour les variables d'état relatives aux activités non programmées
# @chrono
def maj_activites_non_programmees(maj_options_date=True):
    if st.session_state.get("df", None) is None:
        return
    activites_non_programmees = get_activites_non_programmees(st.session_state.df)
    st.session_state.activites_non_programmees = activites_non_programmees
    df_display = _creer_df_display_activites_non_programmees(activites_non_programmees, maj_options_date)
    st.session_state.activites_non_programmees_df_display = df_display
    st.session_state.activites_non_programmees_df_display_copy = df_display.copy()

def maj_ca_display():
    if st.session_state.get("ca", None) is None:
        return
    st.session_state.ca_display = st.session_state.ca.copy()

# Met à jour le contexte complet (activités programmées, non programmées et créneaux disponibles)
def maj_contexte(maj_donnees_calculees=True, maj_options_date=True):
    st.session_state.setdefault("bd_maj_contexte_cmd", {"maj_donnees_calculees": maj_donnees_calculees, "maj_options_date": maj_options_date})
    tracer.log(f"Debut", types=["gen"])
    if maj_donnees_calculees:
        maj_donnees_calculees_df()
    maj_activites_programmees(maj_options_date) # pour mise à jour menus options date
    maj_activites_non_programmees(maj_options_date) # pour mise à jour menus options date
    maj_creneaux_disponibles()
    maj_ca_display()
    tracer.log(f"Fin", types=["gen"])
    del st.session_state["bd_maj_contexte_cmd"]

# Met à jour la variable d'état qui donne la liste des créneaux disponibles
# @chrono
def maj_creneaux_disponibles():
    df = st.session_state.get("df")
    if df is None:
        return
    
    activites_programmees = st.session_state.get("activites_programmees")
    if activites_programmees is None:
        return
    
    traiter_pauses = st.session_state.get("traiter_pauses", False)
    st.session_state.creneaux_disponibles = get_creneaux(df, activites_programmees, traiter_pauses) 
    # if st.session_state.creneaux_disponibles is not None and len(st.session_state.creneaux_disponibles) > 0:
    #     demander_selection("creneaux_disponibles", st.session_state.creneaux_disponibles.index[0])

# Met à jour les données calculées d'une ligne
def _maj_donnees_calculees_row(idx, full=True):
    df = st.session_state.get("df", None)
    if df is None:
        return
    if idx not in df.index:
        return
    try:
        if len(df) > 0:
            debut = heure_parse(df.loc[idx, "Debut"])
            duree = duree_parse(df.loc[idx, "Duree"])
            
            df.at[idx, "Debut_dt"] = debut
            df.at[idx, "Duree_dt"] = duree

            fin = calculer_fin_row(df.loc[idx])
            df.at[idx, "Fin"] = fin

            if full:
                df = st.session_state.get("activites_programmees", None)
                if df is not None and idx in df.index:
                    df.at[idx, "Debut_dt"] = debut
                    df.at[idx, "Duree_dt"] = duree
                    df.at[idx, "Fin"] = fin

                df = st.session_state.get("activites_programmees_df_display", None)
                if df is not None and idx in df.index:
                    df.at[idx, "Fin"] = fin
                    st.session_state.activites_programmees_df_display_copy = st.session_state.activites_programmees_df_display.copy()
                
                df = st.session_state.get("activites_non_programmees", None)
                if df is not None and idx in df.index:
                    df.at[idx, "Debut_dt"] = debut
                    df.at[idx, "Duree_dt"] = duree
                    df.at[idx, "Fin"] = fin
                df = st.session_state.get("activites_non_programmees_df_display", None)
                if df is not None and idx in df.index:
                    df.at[idx, "Fin"] = fin
                    st.session_state.activites_programmees_df_display_copy = st.session_state.activites_programmees_df_display.copy()

    except:
        pass        

# Met à jour les données calculées sur st.session_state
# A utiliser conjointement avec maj_activites_programmees, maj_activites_non_programmees et , maj_creneaux_disponibles
# pour reconstituer le contexte apres chargement de nouvelles données via fichier ou google sheet
def maj_donnees_calculees_df():
    df = st.session_state.get("df", None)
    if df is None:
        return
    try:
        if len(df) > 0:
            df["Debut_dt"] = df["Debut"].apply(heure_parse)
            df["Duree_dt"] = df["Duree"].apply(duree_parse)
            df["Fin"] = df.apply(calculer_fin_row, axis=1)    
            df["Hyperlien"] = get_liens_activites()   
    except:
        pass        

def modifier_cellule(idx, col, val, section_critique=False):

    if section_critique:
        st.session_state.setdefault("bd_modifier_cellule_cmd", 
            {
                "idx": idx,
                "col": col,
                "val": val
            }
        )

    tracer.log(f"Debut {idx} {col} {val}", types=["gen"])

    df = st.session_state.df
    oldval = df.loc[idx, col]
    modifier_df_cell(df, idx, col, val)
    if col == "Date":
        jour = safe_int(val)

        # Programmation d'une activité non programmée
        if (pd.isna(oldval) or oldval == "") and not (pd.isna(val) or val == ""):
            if jour is not None:
                _programmer(idx, jour)
        
        # Déprogrammation d'une activité programmée
        elif not (pd.isna(oldval) or oldval == "") and (pd.isna(val) or val == ""):
            jour = safe_int(oldval)
            if jour is not None:
                _deprogrammer(idx, jour)            

        # Reprogrammation d'une activité programmée    
        elif est_activite_programmee(df.loc[idx]):
            modifier_df_cell(st.session_state.activites_programmees, idx, col, val)
            modifier_df_display_cell(st.session_state.activites_programmees_df_display, idx, df_display_col_nom(col), str(val))
            modifier_df_display_cell(st.session_state.activites_programmees_df_display, idx, "__jour", dateint_to_jour(val) if safe_int(val) is not None else None)
            st.session_state.activites_programmees = st.session_state.activites_programmees.sort_values(by=["Date", "Debut"], ascending=[True, True])
            st.session_state.activites_programmees_df_display = st.session_state.activites_programmees_df_display.sort_values(by=["Date", "Début"], ascending=[True, True])

            _maj_options_date(df, st.session_state.activites_programmees, st.session_state.activites_programmees_df_display, safe_int(oldval))
            _maj_options_date(df, st.session_state.activites_programmees, st.session_state.activites_non_programmees_df_display, safe_int(oldval))
            _maj_options_date(df, st.session_state.activites_programmees, st.session_state.activites_programmees_df_display, jour)
            _maj_options_date(df, st.session_state.activites_programmees, st.session_state.activites_non_programmees_df_display, jour)

            st.session_state.activites_programmees_df_display_copy = st.session_state.activites_programmees_df_display.copy()
            maj_creneaux_disponibles()
    
    else:
        # Affectation de df_activite et df_display (ainsi que de l'id correspondant) selon que l'on traite une activité programmée ou non programmée
        activite_programmee = est_activite_programmee(df.loc[idx])
        df_activite = st.session_state.activites_programmees if activite_programmee else st.session_state.activites_non_programmees
        df_display = st.session_state.activites_programmees_df_display if activite_programmee else st.session_state.activites_non_programmees_df_display
        uuid = df.at[idx, "__uuid"]
        idd = get_index_from_uuid(df_display, uuid)
        if idd is None:
            print(f"Erreur de recherche d'uuid dans modifier_cellule: {uuid}")
            return
        
        # Maj de la colonne modifiée sur df_activite et df_display
        modifier_df_cell(df_activite, idx, col, val)
        modifier_df_cell(df_display, idd, df_display_col_nom(col), val)

        # Maj des données calculées sur df, df_activite et df_display
        if col in ["Debut", "Duree"]: 
            _maj_donnees_calculees_row(idx)

        # Maj des colonnes hyperlien et adresse
        if col == "Activite": 
            lnk = df.loc[idx, "Hyperlien"]
            if isinstance(lnk, str) and isinstance(oldval, str) and isinstance(val, str):
                ancien_nom_dans_lnk = oldval.replace(' ', '+')
                if ancien_nom_dans_lnk in lnk:
                    lnk = lnk.replace(ancien_nom_dans_lnk, val.replace(' ', '+'))
                    modifier_df_cell(df, idx, "Hyperlien", lnk)
                    modifier_df_cell(df_activite, idx, "Hyperlien", lnk)
                    modifier_df_cell(df_display, idd, "Hyperlien", lnk)
        elif col == "Lieu":
            _set_addr_cols(df_display, idd, val)
    
        # Maj spécifiques aux activités programmées
        if activite_programmee:

            if col == "Debut": # Retri du df_activité et du df_display
                st.session_state.activites_programmees = df_activite.sort_values(by=["Date", "Debut"], ascending=[True, True])
                st.session_state.activites_programmees_df_display = df_display.sort_values(by=["Date", "Début"], ascending=[True, True])

            elif col == "Reserve": # Maj des colonnes "__non_reserve" et "__options_date" sur le df_display
                # Maj "__non_reserve"
                non_reserve = str(df_display.loc[idd][df_display_col_nom("Reserve")].strip().lower()) != "oui"
                modifier_df_cell(df_display, idd, "__non_reserve", non_reserve)
                # Maj "__options_date"
                row = df_display.loc[idd]
                options_date = safe_json_dump(get_jours_possibles_from_activite_programmee(row))
                modifier_df_cell(df_display, idd, "__options_date", options_date)

            if col in ["Debut", "Duree", "Relache"]: # Maj des "__options_date" sur les deux df_display pour le jour considéré
                jour = df.loc[idx]["Date"]
                _maj_options_date(df, st.session_state.activites_programmees, st.session_state.activites_programmees_df_display, jour)
                _maj_options_date(df, st.session_state.activites_programmees, st.session_state.activites_non_programmees_df_display, jour)

            # Maj de la copie du df_display
            st.session_state.activites_programmees_df_display_copy = st.session_state.activites_programmees_df_display.copy()

        # Maj spécifiques aux activités non programmées
        else:

            if col == "Debut": # Retri du df_activité et du df_display
                st.session_state.activites_non_programmees = df_activite.sort_values(by=["Date", "Debut"], ascending=[True, True])
                st.session_state.activites_non_programmees_df_display = df_display.sort_values(by=["Date", "Début"], ascending=[True, True])
        
            if col in ["Debut", "Duree", "Relache"]: # Maj "__options_date" 
                row = df_display.loc[idd]
                options_date = safe_json_dump(get_jours_possibles_from_activite_non_programmee(row))
                modifier_df_cell(df_display, idd, "__options_date", options_date)

            # Maj de la copie du df_display
            st.session_state.activites_non_programmees_df_display_copy = st.session_state.activites_non_programmees_df_display.copy()

        maj_creneaux_disponibles()

    tracer.log(f"Fin {idx} {col} {val}", types=["gen"])
    
    if section_critique:
        del st.session_state["bd_modifier_cellule_cmd"]

# Met à jour le contexte pour une activité dont la date de programmation passe de None à jour
# @chrono
def _programmer(idx, jour=None):
    
    if "df" not in st.session_state:
        return
    
    if idx not in st.session_state.df.index:
        return
    
    if "activites_programmees" not in st.session_state:
        return
    
    if "activites_non_programmees" not in st.session_state:
        return

    if jour is None:
        return

    if idx in st.session_state.activites_non_programmees.index:

        row = st.session_state.activites_non_programmees.loc[[idx]]
        row.at[idx, "Date"] = jour
        st.session_state.activites_non_programmees.drop(index=idx, inplace=True)
        st.session_state.activites_programmees = pd.concat([st.session_state.activites_programmees, row]).sort_values(by=["Date", "Debut_dt"], ascending=[True, True])

        row = st.session_state.activites_non_programmees_df_display.loc[[idx]]
        row.at[idx, "Date"] = str(jour)
        row["__jour"] = row["Date"].apply(dateint_to_jour)
        row["__non_reserve"] = row["Réservé"].astype(str).str.strip().str.lower() != "oui"
        st.session_state.activites_non_programmees_df_display.drop(index=idx, inplace=True)
        st.session_state.activites_programmees_df_display = pd.concat([st.session_state.activites_programmees_df_display, row]).sort_values(by=["Date", "Début"], ascending=[True, True])

        _maj_options_date(st.session_state.df, st.session_state.activites_programmees, st.session_state.activites_programmees_df_display, jour)
        _maj_options_date(st.session_state.df, st.session_state.activites_programmees, st.session_state.activites_non_programmees_df_display, jour)

        st.session_state.activites_programmees_df_display_copy = st.session_state.activites_programmees_df_display.copy()
        st.session_state.activites_non_programmees_df_display_copy = st.session_state.activites_non_programmees_df_display.copy()
        
        maj_creneaux_disponibles()

# Suppression d'une activité d'un df
def supprimer_activite(idx):
    if idx not in st.session_state.df.index:
        return
    jour = st.session_state.df.loc[idx]["Date"]
    uuid = st.session_state.df.loc[idx]["__uuid"]
    st.session_state.df.loc[idx] = pd.NA
    st.session_state.df.at[idx, "__uuid"] = uuid
    st.session_state.activites_programmees = supprimer_row_df(st.session_state.activites_programmees, idx)
    st.session_state.activites_non_programmees = supprimer_row_df(st.session_state.activites_non_programmees, idx)
    st.session_state.activites_programmees_df_display = supprimer_row_df_display(st.session_state.activites_programmees_df_display, idx)
    st.session_state.activites_programmees_df_display_copy = supprimer_row_df_display(st.session_state.activites_programmees_df_display_copy, idx)
    st.session_state.activites_non_programmees_df_display = supprimer_row_df_display(st.session_state.activites_non_programmees_df_display, idx)
    st.session_state.activites_non_programmees_df_display_copy = supprimer_row_df_display(st.session_state.activites_non_programmees_df_display_copy, idx)
    _maj_options_date(st.session_state.df, st.session_state.activites_programmees, st.session_state.activites_programmees_df_display, jour)
    _maj_options_date(st.session_state.df, st.session_state.activites_programmees, st.session_state.activites_non_programmees_df_display, jour)
    maj_creneaux_disponibles()

def _ensure_addr_cols(df):
    if "__addr_enc"   not in df.columns: df["__addr_enc"]   = None
    carnet = st.session_state.get("ca")
    city_default = st.session_state.get("city_default", "")
    mask = df["Lieu"].notna()
    for i in df.index[mask]:
        if pd.isna(df.at[i, "__addr_enc"]) or not str(df.at[i, "__addr_enc"]).strip():
            addr_h, addr_enc = resolve_address_fast(df.at[i, "Lieu"], carnet, city_default=city_default)
            df.at[i, "__addr_enc"] = addr_enc
    return df

def _set_addr_cols(df, idx, lieu):
    carnet = st.session_state.get("ca")
    city_default = st.session_state.get("city_default", "")
    addr_h, addr_enc = resolve_address_fast(lieu, carnet, city_default=city_default)
    df.at[idx, "__addr_enc"] = addr_enc

def _calculer_options_date_activites_programmees(df_display):
    # Hash non pertinent en l'état car cette fonction n'est appelée par maj_activites_non_programmees que si les données d'entrée on changé
    # hash_val  = hash_df(
    #     df_display, 
    #     colonnes_a_garder=["Date", "Debut", "Duree"], 
    #     params=[
    #         st.session_state.periode_a_programmer_debut.isoformat(), 
    #         st.session_state.periode_a_programmer_fin.isoformat(),
    #         str(st.session_state.MARGE.total_seconds()),
    #         str(st.session_state.DUREE_REPAS.total_seconds()),
    #         str(st.session_state.DUREE_CAFE.total_seconds())])
    # hash_key = "options_date_activites_programmees__hash"
    # key = "options_date_activites_programmees"
    # if st.session_state.get(hash_key) != hash_val:
    #     st.session_state[key] = df_display.apply(lambda row: get_jours_possibles_from_activite_programmee(row), axis=1)
    #     st.session_state[hash_key] = hash_val
    # return st.session_state[key]
    return df_display.apply(lambda row: get_jours_possibles_from_activite_programmee(row), axis=1)

# Calcule les options des dates pour les activiés non programmées
def calculer_options_date_activites_non_programmees(df_display):
    # Hash non pertinent en l'état car cette fonction n'est appelée par maj_activites_non_programmees que si les données d'entrée on changé
    # hash_val  = hash_df(
    #     df_display, 
    #     colonnes_a_garder=["Date", "Debut", "Duree"], 
    #     params=[
    #         st.session_state.periode_a_programmer_debut.isoformat(), 
    #         st.session_state.periode_a_programmer_fin.isoformat(),
    #         str(st.session_state.MARGE.total_seconds()),
    #         str(st.session_state.DUREE_REPAS.total_seconds()),
    #         str(st.session_state.DUREE_CAFE.total_seconds())])
    # hash_key = "options_date_activites_non_programmees__hash"
    # key = "options_date_activites_non_programmees"
    # if st.session_state.get(hash_key) != hash_val:
    #     st.session_state[key] = df_display.apply(lambda row: get_jours_possibles_from_activite_non_programmee(row), axis=1)
    #     st.session_state[hash_key] = hash_val
    # return st.session_state[key]
    return df_display.apply(lambda row: get_jours_possibles_from_activite_non_programmee(row), axis=1)

# Met à jour la colonne __options_date d'un df_display donné pour un jour donné
def _maj_options_date(df, activites_programmees, df_display, jour):
    """
    - jour: int (jour modifié)
    Met à jour uniquement les lignes dont __options_date contient `jour`.
    Retourne la liste des index modifiés.
    """
    if jour is None or pd.isna(jour):
        return

    jour_str = str(jour)

    changed_idx = []

    # Pré-filtrage simple : on parcourt uniquement les lignes où la chaîne n'est pas vide.
    # (on pourrait accélérer avec .dropna() / .astype(str), mais restons sûrs)
    for i, s in df_display["__options_date"].items():
        if not s:
            continue
    
        # parse -> set[str]
        opts = parse_options_date(s)
        
        # Activité courante
        row = df_display.loc[i]

        # S'il s'agit d'une activité programmée au jour dit...
        if row["Date"] == jour_str:
            # S'il s'agit d'une activité réservée on vérifie que le menu est vide. Si ce n'est pas le cas on le vide.
            if est_activite_reserve(df.loc[i]):
                if opts != set():
                    df_display.at[i, "__options_date"] = dump_options_date(set())
                    changed_idx.append(i)
            # Sinon on vérifie que le menu n'est pas vide (cas d'une activité qui serait passée de réservée à non réservée).
            # Dans ce cas on reconstruit le menu.
            else:
                if opts == set():
                    df_display.at[i, "__options_date"] = dump_options_date(get_jours_possibles_from_activite_programmee(row))
                    changed_idx.append(i)
            # Sinon rien d'autre à faire
            # car s'il s'agit d'une activité reprogrammée au jour dit ce jour était déjà dans le menu avant reprogrammation et doit y rester
            # et sinon ce jour est déjà dans le menu et doit y rester aussi pour que le déploiement dudit menu n'oblige pas à changer de jour.
        
        else:

            # si le jour n'était pas présent ET que la règle ne le concerne pas, on peut sauter
            # (mais on doit tout de même appeler la règle si tu veux ajouter quand c'est possible)
            allowed = est_jour_possible(df, activites_programmees, i, jour)

            # remove si plus possible
            if not allowed and jour_str in opts:
                opts.remove(jour_str)
                if len(opts) == 1 and '' in opts:
                    opts = set() # un menu ne doit pas avoir un seul élément vide
                df_display.at[i, "__options_date"] = dump_options_date(opts)
                changed_idx.append(i)

            # add si maintenant possible
            elif allowed and jour_str not in opts:
                opts.add(jour_str)
                if len(opts) == 1:
                    opts.add('') # il faut un item vide dans un menu avec des jours valides pour permettre la déprogrammation
                df_display.at[i, "__options_date"] = dump_options_date(opts)
                changed_idx.append(i)

    return changed_idx

# Affichage de la période à programmer
def initialiser_periode_programmation(df):

    if "nouveau_fichier" not in st.session_state:
        st.session_state.nouveau_fichier = True
    
    # Initialisation de la periode si nouveau fichier
    if st.session_state.nouveau_fichier == True:
        # Reset du flag déclenché par uploader
        st.session_state.nouveau_fichier = False

        # Initialisation des variables de début et de fin de période à programmer
        periode_a_programmer_debut = None 
        periode_a_programmer_fin = None

        dates_valides = get_dates_from_df(df)

        if not dates_valides.empty:
            # Conversion en datetime
            dates_datetime = dates_valides.apply(dateint_to_date)
            if not dates_datetime.empty:
                periode_a_programmer_debut = dates_datetime.min()
                periode_a_programmer_fin = dates_datetime.max()

        if periode_a_programmer_debut is None or periode_a_programmer_fin is None:
            dates_festival = get_dates_festival()
            periode_a_programmer_debut = dates_festival["debut"]
            periode_a_programmer_fin = dates_festival["fin"]
        
        st.session_state.periode_a_programmer_debut = periode_a_programmer_debut
        st.session_state.periode_a_programmer_fin = periode_a_programmer_fin
    
    if "periode_a_programmer_debut" not in st.session_state or "periode_a_programmer_fin" not in st.session_state:
        dates_festival = get_dates_festival()
        st.session_state.periode_a_programmer_debut = dates_festival["debut"]
        st.session_state.periode_a_programmer_fin = dates_festival["fin"]
    
# Nettoie les données du tableau Excel importé
def _nettoyer_donnees(df, fn):
    try:
        # Nettoyage noms de colonnes : suppression espaces et accents
        df.columns = df.columns.str.strip().str.replace("\u202f", " ").str.normalize("NFKD").str.encode("ascii", errors="ignore").str.decode("utf-8")

        if not all(col in df.columns for col in COLONNES_ATTENDUES):
            st.session_state.contexte_invalide_message = f"Le fichier {fn} n'est pas au format Excel ou ne contient pas toutes les colonnes attendues: " + ", ".join(COLONNES_ATTENDUES_ACCENTUEES) + "."
        else:
            initialiser_dtypes(df)

            if (len(df) > 0):
                # Suppression des lignes presque vides i.e. ne contenant que des NaN ou des ""
                df = df[~df.apply(lambda row: all(pd.isna(x) or str(x).strip() == "" for x in row), axis=1)].reset_index(drop=True)

                # Nettoyage Heure (transforme les datetime, time et None en str mais ne garantit pas le format HHhMM, voir est_heure_valide pour cela)
                df["Debut"] = df["Debut"].apply(heure_str).astype("string")

                # Nettoyage Duree (transforme les timedelta et None en str mais ne garantit pas le format HhMM, voir est_duree_valide pour cela)
                df["Duree"] = df["Duree"].apply(duree_str).astype("string")

                # Colonne Relache castée en object avec NaN remplacés par "" et le reste en str
                df["Relache"] = df["Relache"].astype("object").fillna("").astype(str)

            # Valide le contexte si pas d'exception dans le traitement précédent
            if "contexte_invalide" in st.session_state:
                del st.session_state["contexte_invalide"]

        return df
            
    except Exception as e:
        st.error(f"Erreur lors du décodage du fichier : {e}")
        df = pd.DataFrame(columns=COLONNES_ATTENDUES)
        initialiser_dtypes(df)


# Renvoie les hyperliens de la colonne Activité 
def get_liens_activites(wb):
    liens_activites = {}
    try:
        ws = wb.worksheets[0]
        for cell in ws[1]:
            if cell.value and str(cell.value).strip().lower() in ["activité"]:
                col_activite_index = cell.column
        for row in ws.iter_rows(min_row=2, min_col=col_activite_index, max_col=col_activite_index):
            cell = row[0]
            if cell.hyperlink:
                liens_activites[cell.value] = cell.hyperlink.target
            else:
                # Construire l'URL de recherche par défaut
                if cell.value is not None:
                    url = f"https://www.festivaloffavignon.com/resultats-recherche?recherche={cell.value.replace(' ', '+')}"
                    liens_activites[cell.value] = url  # L'enregistrer dans la session
        return liens_activites
    except:
        return liens_activites

# Vérifie la cohérence des informations du dataframe et affiche le résultat dans un expander
def verifier_coherence(df):
    
    # @st.cache_data
    def get_log_verifier_coherence(df):
        # try:
        erreurs = []

        def est_entier(x):
            try:
                return not pd.isna(x) and str(x).strip() != "" and int(float(x)) == float(x)
            except Exception:
                return False
        
        if len(df) <= 0:
            return
        
        # 1. 🔁 Doublons
        df_valid = df[df["Activite"].notna() & (df["Activite"].astype(str).str.strip() != "")]
        df_valid = df_valid.copy()  # pour éviter SettingWithCopyWarning
        df_valid["_activite_clean"] = df_valid["Activite"].astype(str).str.strip().str.lower()
        doublons = df_valid[df_valid.duplicated(subset=["_activite_clean"], keep=False)]

        if not doublons.empty:
            bloc = []
            for _, row in doublons.iterrows():
                if not est_pause(row):
                    try:
                        date_str = str(int(float(row["Date"]))) if pd.notna(row["Date"]) else "Vide"
                    except (ValueError, TypeError):
                        date_str = "Vide"
                    heure_str = str(row["Debut"]).strip() if pd.notna(row["Debut"]) else "Vide"
                    duree_str = str(row["Duree"]).strip() if pd.notna(row["Duree"]) else "Vide"
                    
                    if not bloc:
                        bloc = ["🟠 Doublons d'activités :"]

                    bloc.append(f"{date_str} - {heure_str} - {row['Activite']} ({duree_str})")
            erreurs.append("\n".join(bloc))
            
        # 2. ⛔ Chevauchements
        chevauchements = []
        df_sorted = df.sort_values(by=["Date", "Debut_dt"])
        for i in range(1, len(df_sorted)):
            r1 = df_sorted.iloc[i - 1]
            r2 = df_sorted.iloc[i]
            if r1.isna().all() or r2.isna().all():
                continue
            if pd.notna(r1["Date"]) and pd.notna(r2["Date"]) and r1["Date"] == r2["Date"]:
                fin1 = r1["Debut_dt"] + r1["Duree_dt"]
                debut2 = r2["Debut_dt"]
                if debut2 < fin1:
                    chevauchements.append((r1, r2))
        if chevauchements:
            bloc = ["🔴 Chevauchements:"]
            for r1, r2 in chevauchements:
                bloc.append(
                    f"{r1['Activite']} ({r1['Debut']} / {r1['Duree']}) chevauche {r2['Activite']} ({r2['Debut']} / {r2['Duree']}) le {r1['Date']}"
                )
            erreurs.append("\n".join(bloc))

        # 3. 🕒 Erreurs de format
        bloc_format = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            # Date : uniquement si non NaN
            if pd.notna(row["Date"]) and not est_entier(row["Date"]):
                bloc_format.append(f"Date invalide à la ligne {idx + 2} : {row['Date']}")

            # Ne tester Heure/Duree que si Activite ou Autres est renseigné
            if str(row["Activite"]).strip() != "":
                if not re.match(r"^\d{1,2}h\d{2}$", str(row["Debut"]).strip()):
                    bloc_format.append(f"Heure invalide à la ligne {idx + 2} : {row['Debut']}")
                if not re.match(r"^\d{1,2}h\d{2}$", str(row["Duree"]).strip()):
                    bloc_format.append(f"Durée invalide à la ligne {idx + 2} : {row['Duree']}")
            
            # Test de la colonne Relache
            if not est_relache_valide(row["Relache"]):
                bloc_format.append(f"Relache invalide à la ligne {idx + 2} : {row['Relache']}")

        # 4. 📆 Spectacles un jour de relâche (Date == Relache)
        bloc_relache = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            if (
                est_entier(row["Date"]) and
                est_entier(row["Relache"]) and
                int(float(row["Date"])) == int(float(row["Relache"])) and
                str(row["Activite"]).strip() != ""
            ):
                bloc_relache.append(
                    f"{row['Activite']} prévu le jour de relâche ({int(row['Date'])}) à la ligne {idx + 2}"
                )
        if bloc_relache:
            erreurs.append("🛑 Spectacles programmés un jour de relâche:\n" + "\n".join(bloc_relache))

        # 5. 🕳️ Heures non renseignées
        bloc_heure_vide = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            if str(row["Activite"]).strip() != "":
                if pd.isna(row["Debut"]) or str(row["Debut"]).strip() == "":
                    bloc_heure_vide.append(f"Heure vide à la ligne {idx + 2}")
        if bloc_heure_vide:
            erreurs.append("⚠️ Heures non renseignées:\n" + "\n".join(bloc_heure_vide))

        # 6. 🕓 Heures au format invalide
        bloc_heure_invalide = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            if str(row["Activite"]).strip() != "":
                h = row["Debut"]
                if pd.notna(h) and str(h).strip() != "":
                    h_str = str(h).strip().lower()
                    is_time_like = isinstance(h, (datetime.datetime, datetime.time))
                    valid_format = bool(re.match(r"^\d{1,2}h\d{2}$", h_str) or re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", h_str))
                    if not is_time_like and not valid_format:
                        bloc_heure_invalide.append(f"Heure invalide à la ligne {idx + 2} : {h}")
        if bloc_heure_invalide:
            erreurs.append("⛔ Heures mal formatées:\n" + "\n".join(bloc_heure_invalide))

        # 7. 🕳️ Durées non renseignées ou nulles
        bloc_duree_nulle = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            if isinstance(row["Duree_dt"], pd.Timedelta) and row["Duree_dt"] == pd.Timedelta(0):
                if pd.isna(row["Duree"]) or str(row["Duree"]).strip() == "":
                    msg = f"Durée vide à la ligne {idx + 2}"
                else:
                    msg = f"Durée égale à 0 à la ligne {idx + 2} : {row['Duree']}"
                bloc_duree_nulle.append(msg)
        if bloc_duree_nulle:
            erreurs.append("⚠️ Durées nulles ou vides:\n" + "\n".join(bloc_duree_nulle))

        # 8. ⏱️ Durées au format invalide
        bloc_duree_invalide = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            if str(row["Activite"]).strip() != "":
                d = row["Duree"]
                if pd.notna(d) and str(d).strip() != "":
                    d_str = str(d).strip().lower()
                    is_timedelta = isinstance(d, pd.Timedelta)
                    valid_format = bool(re.match(r"^\d{1,2}h\d{2}$", d_str) or re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", d_str))
                    if not is_timedelta and not valid_format:
                        bloc_duree_invalide.append(f"Durée invalide à la ligne {idx + 2} : {d}")
        if bloc_duree_invalide:
            erreurs.append("⛔ Durées mal formatées:\n" + "\n".join(bloc_duree_invalide))

        contenu = "<div style='font-size: 14px;'>"
        for bloc in erreurs:
            lignes = bloc.split("\n")
            if lignes[0].startswith(("🟠", "🔴", "⚠️", "🛑", "⛔")):
                contenu += f"<p><strong>{lignes[0]}</strong></p><ul>"
                for ligne in lignes[1:]:
                    contenu += f"<li>{ligne}</li>"
                contenu += "</ul>"
            else:
                contenu += f"<p>{bloc}</p>"
        contenu += "</div>"
        return contenu
        
    with st.expander("Cohérence des données"):
        st.markdown(get_log_verifier_coherence(df), unsafe_allow_html=True)

def valider_valeur(df, colonne, nouvelle_valeur):           
    erreur = None
    if colonne == "Debut" and not est_heure_valide(nouvelle_valeur):
        erreur = "⛔ Format attendu : HHhMM (ex : 10h00)"
    elif colonne == "Duree" and not est_duree_valide(nouvelle_valeur):
        erreur = "⛔ Format attendu : HhMM (ex : 1h00 ou 0h30)"
    elif colonne == "Relache" and not est_relache_valide(nouvelle_valeur):
        erreur = "⛔ Format attendu : voir A propos / Format des données"
    elif colonne == "Reserve" and not est_reserve_valide(nouvelle_valeur):
        erreur = "⛔ Format attendu : Oui, Non"
    elif ptypes.is_numeric_dtype(df[colonne]) and not ptypes.is_numeric_dtype(nouvelle_valeur):
        try:
            if "." not in nouvelle_valeur and "," not in nouvelle_valeur and "e" not in nouvelle_valeur.lower():
                nouvelle_valeur = int(nouvelle_valeur)
            else:
                nouvelle_valeur = float(nouvelle_valeur)
        except:
            erreur = "⛔ Format numérique attendu"
    return erreur

# Affecte une nouvelle valeur à une cellule du df de base donnée par son index et sa colonne
def affecter_valeur_df(index, colonne, nouvelle_valeur, section_critique=None):
    
    def set_section_critique_step(section_critique, step):
        if section_critique is not None:
            section_critique["step"] = step

    df = st.session_state.df
    valeur_courante = df.at[index, colonne]
    step = section_critique["step"] if section_critique is not None else 0
    tracer.log(f"step {step}")
    erreur = None

    if step == 0:
        erreur = valider_valeur(df, colonne, nouvelle_valeur)
        if not erreur:
            set_section_critique_step(section_critique, 1)
            if colonne == "Debut" :
                heures, minutes = nouvelle_valeur.split("h")
                nouvelle_valeur = f"{int(heures):02d}h{int(minutes):02d}"
            if (pd.isna(valeur_courante) and pd.notna(nouvelle_valeur)) or nouvelle_valeur != valeur_courante:
                try:
                    df.at[index, colonne] = nouvelle_valeur
                except Exception as e:
                    erreur = f"⛔ {e}"
                else:
                    set_section_critique_step(section_critique, 2)
                    df.at[index, colonne] = valeur_courante
                    undo.save()
                    modifier_cellule(index, colonne, nouvelle_valeur)
                    sql.sauvegarder_row(index)
    elif step == 1:
        if colonne == "Debut" :
            heures, minutes = nouvelle_valeur.split("h")
            nouvelle_valeur = f"{int(heures):02d}h{int(minutes):02d}"
        if (pd.isna(valeur_courante) and pd.notna(nouvelle_valeur)) or nouvelle_valeur != valeur_courante:
            try:
                df.at[index, colonne] = nouvelle_valeur
            except Exception as e:
                erreur = f"⛔ {e}"
            else:
                set_section_critique_step(section_critique, 2)
                df.at[index, colonne] = valeur_courante
                undo.save()
                modifier_cellule(index, colonne, nouvelle_valeur, section_critique=section_critique)
                sql.sauvegarder_row(index)
    elif step == 2:
        df.at[index, colonne] = valeur_courante
        undo.save()
        modifier_cellule(index, colonne, nouvelle_valeur, section_critique=section_critique)
        sql.sauvegarder_row(index)
        
    return erreur

# Affecte une nouvelle valeur à une cellule d'une row d'un df 
def affecter_valeur_row(row, colonne, nouvelle_valeur):
    valeur_courante = row[colonne]
    erreur = valider_valeur(row, colonne, nouvelle_valeur)
    if not erreur:
        if colonne == "Debut" :
            heures, minutes = nouvelle_valeur.split("h")
            nouvelle_valeur = f"{int(heures):02d}h{int(minutes):02d}"
        if (pd.isna(valeur_courante) and pd.notna(nouvelle_valeur)) or nouvelle_valeur != valeur_courante:
            try:
                row[colonne] = nouvelle_valeur
            except Exception as e:
                erreur = f"⛔ {e}"

    return erreur

# Vérifie qu'une valeur est bien Oui Non
def est_reserve_valide(val):
    return str(val).strip().lower() in ["oui", "non", ""]

# token = liste de jours (1 ou plusieurs séparés par -),
# éventuellement suivis de /mois et éventuellement /année (2 ou 4 chiffres)
_TOKEN = r"(?:\d{1,2}(?:\s*-\s*\d{1,2})*)(?:/\d{1,2}(?:/\d{2,4})?)?"
_MOTIF_RELACHE = re.compile(
    rf"^\s*(?:{_TOKEN}|pair|impair)(\s*,\s*(?:{_TOKEN}|pair|impair))*\s*$",
    re.IGNORECASE
)

# ---------- parsing ----------
def _valid_date(y: int, m: int, d: int) -> bool:
    try:
        datetime.date(y, m, d)
        return True
    except ValueError:
        return False

def _y2k(yy: int) -> int:
    return 2000 + yy if yy < 100 else yy

def _mk_dateint(y: int, m: int, d: int) -> Optional[int]:
    try:
        datetime.date(y, m, d)
        return y*10000 + m*100 + d
    except ValueError:
        return None

def _tokenize_specs(s: str) -> List[str]:
    """
    Découpe par virgules de haut niveau (on ne coupe pas à l'intérieur des parenthèses).
    """
    if not isinstance(s, str):
        return []
    out, cur, depth = [], [], 0
    for ch in s:
        if ch == '(':
            depth += 1
            cur.append(ch)
        elif ch == ')':
            depth = max(0, depth-1)
            cur.append(ch)
        elif ch == ',' and depth == 0:
            tok = ''.join(cur).strip()
            if tok:
                out.append(tok)
            cur = []
        else:
            cur.append(ch)
    tok = ''.join(cur).strip()
    if tok:
        out.append(tok)
    return out

def _parse_one_token(tok: str, default_year: int, default_month: int) -> dict:
    """
    Retourne un dict décrivant la règle, ou {} si le token est invalide.
    Types possibles:
      - interval  : bornes d1-d2, avec délimiteurs
                    [d1-d2]  (intervalle fermé = relâche)
                    <d1-d2>  (fenêtre de jeu = hors relâche)
                    option: /mm(/yyyy) traînant -> mois/année implicites pour d1 et d2
      - list      : (a,b,c)/mm(/yyyy) -> jours de relâche explicites
      - days      : 8,25(/mm(/yyyy))  -> jours de relâche explicites
      - parity    : "jours pairs" | "jours impairs"
    """
    t = tok.strip().lower()

    # --- 1) Intervalle avec crochets (relâche) OU chevrons (fenêtre de jeu) ---
    # [d1-d2]  ou  <d1-d2>  + option /mm(/yyyy)
    m = re.fullmatch(
        r"(?P<L>\[|<)\s*(?P<d1>\d{1,2})\s*-\s*(?P<d2>\d{1,2})\s*(?P<R>\]|>)\s*(?:/\s*(?P<mm>\d{1,2})(?:/\s*(?P<yyyy>\d{2,4}))?)?",
        t
    )
    if m:
        d1 = int(m.group("d1")); d2 = int(m.group("d2"))
        mm = int(m.group("mm")) if m.group("mm") else default_month
        yy = _y2k(int(m.group("yyyy"))) if m.group("yyyy") else default_year

        if not (_valid_date(yy, mm, d1) and _valid_date(yy, mm, d2)):
            return {}

        is_window = (m.group("L") == "<" and m.group("R") == ">")   # <d1-d2>  = fenêtre de jeu
        return {
            "type": "interval",
            "y": yy, "m": mm, "d1": d1, "d2": d2,
            "mode": "window" if is_window else "closed"  # "closed" = relâche
        }

    # --- 2) Regroupement par mois: (a,b,c)/mm(/yyyy) -> relâches ciblées ---
    m = re.fullmatch(r"\(\s*([\d\s,]+)\s*\)\s*/\s*(\d{1,2})(?:/\s*(\d{2,4}))?", t)
    if m:
        days_part, mm, yyyy = m.groups()
        mm = int(mm)
        yy = _y2k(int(yyyy)) if yyyy else default_year
        days = [int(x) for x in re.findall(r"\d{1,2}", days_part)]
        if not days: 
            return {}
        # valide toutes les dates
        for d in days:
            if not _valid_date(yy, mm, d):
                return {}
        return {"type": "list", "y": yy, "m": mm, "days": days}

    # --- 3) Parité (relâche) ---
    if re.fullmatch(r"(jours?\s+)?pairs?", t):
        return {"type": "parity", "parity": "pairs"}
    if re.fullmatch(r"(jours?\s+)?impairs?", t):
        return {"type": "parity", "parity": "impairs"}


    # --- 4) Jours simples (relâche) : "8,25", ou "8,25/07", ou "8,25/07/2025"
    m = re.fullmatch(r"([\d\s,]+)(?:/\s*(\d{1,2})(?:/\s*(\d{2,4}))?)?", t)
    if m:
        days_part, mm, yyyy = m.groups()
        days = [int(x) for x in re.findall(r"\d{1,2}", days_part)]
        if not days:
            return {}
        mm = int(mm) if mm else default_month
        yy = _y2k(int(yyyy)) if yyyy else default_year
        for d in days:
            if not _valid_date(yy, mm, d):
                return {}
        return {"type": "days", "y": yy, "m": mm, "days": days}

    # rien de reconnu
    return {}

def _parse_relache(relache_val: str, *, default_year: Optional[int], default_month: Optional[int]) -> List[Dict[str, Any]]:
    if not relache_val or str(relache_val).strip() == "":
        return []  # aucune relâche => hors relâche partout
    tokens = _tokenize_specs(relache_val)
    # Défauts si manquants (utilise mois/année “courants” du contexte)
    if default_year is None or default_month is None:
        today = datetime.date.today()
        default_year  = today.year
        default_month = today.month

    rules: List[Dict[str, Any]] = []
    for tok in tokens:
        rule = _parse_one_token(tok, default_year, default_month)
        if rule:
            rules.append(rule)

    # Attache “jours pairs/impairs” au dernier intervalle sans parité si le token
    # de parité est juste après.
    attached: List[Dict[str, Any]] = []
    i = 0
    while i < len(rules):
        r = rules[i]
        if r["type"] == "interval" and i+1 < len(rules) and rules[i+1]["type"] == "parity" and (r.get("parity") is None):
            r = {**r, "parity": rules[i+1]["parity"]}
            attached.append(r)
            i += 2
        else:
            attached.append(r)
            i += 1
    return attached

def _parse_day_maybe_dmY(token: str, def_y: int, def_m: int) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    """
    Parse 'd' | 'd/m' | 'd/m/yy' | 'd/m/yyyy' -> (Y,M,D), avec défauts (def_y, def_m).
    """
    t = token.strip()
    m = re.fullmatch(r"(\d{1,2})(?:/(\d{1,2})(?:/(\d{2,4}))?)?", t)
    if not m:
        return (None, None, None)
    d = int(m.group(1))
    mm = int(m.group(2)) if m.group(2) else def_m
    if m.group(3):
        yy = int(m.group(3))
        yy = _y2k(yy) if yy < 100 else yy
    else:
        yy = def_y
    if _mk_dateint(yy, mm, d) is None:
        return (None, None, None)
    return (yy, mm, d)

# ---------- API demandée ----------
def est_relache_valide(val: object, *, default_year: int | None = None, default_month: int | None = None) -> bool:
    """
    Valide la chaîne :
      - vide/NaN => True
      - sinon, TOUS les tokens doivent être valides (pas seulement un).
    """
    s = "" if val is None else str(val).strip()
    if s == "":
        return True

    tokens = _tokenize_specs(s)  # ta fonction existante de découpage par virgules de haut niveau
    if not tokens:
        return False

    if default_year is None or default_month is None:
        today = datetime.date.today()
        default_year, default_month = today.year, today.month

    # ❗️On exige que chaque token passe _parse_one_token
    for tok in tokens:
        rule = _parse_one_token(tok, default_year, default_month)
        if not rule:
            return False
    return True

def est_hors_relache(relache_val: Optional[str], date_val: Optional[int], today: Optional[datetime.date] = None) -> bool:
    """
    True  = jour jouable (hors relâche)
    False = jour de relâche

    relache_val peut contenir, séparés par des virgules :
      - Intervalles de relâche             : [5-26], [05-10]/07, [25/07-01/08]
      - Fenêtres de jeu                    : <5-26>, <05-10>/07, <25/07-01/08>
      - Regroupements de jours de relâche  : (8,25), (9,16,23)/07
      - Relâche jours pairs ou impairs     : "jours pairs" / "jours impairs"

    Règles de décision :
      1) Si la date matche un bloc de RELÂCHE (intervalle / regroupement / parité) -> False
      2) Sinon, s'il existe des fenêtres de JEU (<...>) :
            - si la date est DANS AU MOINS l'une d'entre elles (bornes incluses) -> True
            - sinon -> False
      3) Sinon (aucune fenêtre et pas de relâche) -> True
    """
    if relache_val is None or str(relache_val).strip() == "" or date_val is None:
        return True

    try:
        dv = int(date_val)
    except Exception:
        return True

    dy = dv // 10000
    dm = (dv // 100) % 100
    dd = dv % 100

    base = today or datetime.date.today()
    def_y, def_m = base.year, base.month

    txt = str(relache_val).strip().lower()

    # --- Parité (relâche) ---
    parite_relache = None  # "pair" | "impair" | None
    if re.search(r"\brel[aâ]che\s+jours?\s+pairs?\b", txt) or re.search(r"\bjours?\s+pairs?\b", txt):
        parite_relache = "pair"
    if re.search(r"\brel[aâ]che\s+jours?\s+impairs?\b", txt) or re.search(r"\bjours?\s+impairs?\b", txt):
        parite_relache = "impair"

    closed_intervals: List[Tuple[int,int]] = []  # relâche
    open_intervals:   List[Tuple[int,int]] = []  # JEU
    regroup_days:     List[int]            = []  # relâche

    # 1) Intervalles fermés [A-B] avec /mm(/yyyy) traînant
    for m in re.finditer(r"\[\s*([0-9/]+)\s*-\s*([0-9/]+)\s*\]\s*(?:/(\d{1,2})(?:/(\d{2,4}))?)?", txt):
        a_txt, b_txt, mm_txt, yy_txt = m.groups()
        mm_def = int(mm_txt) if mm_txt else def_m
        if yy_txt:
            yy_def = int(yy_txt); yy_def = _y2k(yy_def) if yy_def < 100 else yy_def
        else:
            yy_def = def_y

        Ay, Am, Ad = _parse_day_maybe_dmY(a_txt, yy_def, mm_def)
        By, Bm, Bd = _parse_day_maybe_dmY(b_txt, yy_def, mm_def)
        if None not in (Ay, Am, Ad, By, Bm, Bd):
            a_di = _mk_dateint(Ay, Am, Ad)
            b_di = _mk_dateint(By, Bm, Bd)
            if a_di is not None and b_di is not None:
                lo, hi = (a_di, b_di) if a_di <= b_di else (b_di, a_di)
                closed_intervals.append((lo, hi))

    # 2) Fenêtres de jeu <A-B> avec /mm(/yyyy) traînant   ⟵ (remplace l'ancien motif ]A-B[)
    for m in re.finditer(r"<\s*([0-9/]+)\s*-\s*([0-9/]+)\s*>\s*(?:/(\d{1,2})(?:/(\d{2,4}))?)?", txt):
        a_txt, b_txt, mm_txt, yy_txt = m.groups()
        mm_def = int(mm_txt) if mm_txt else def_m
        if yy_txt:
            yy_def = int(yy_txt); yy_def = _y2k(yy_def) if yy_def < 100 else yy_def
        else:
            yy_def = def_y

        Ay, Am, Ad = _parse_day_maybe_dmY(a_txt, yy_def, mm_def)
        By, Bm, Bd = _parse_day_maybe_dmY(b_txt, yy_def, mm_def)
        if None not in (Ay, Am, Ad, By, Bm, Bd):
            a_di = _mk_dateint(Ay, Am, Ad)
            b_di = _mk_dateint(By, Bm, Bd)
            if a_di is not None and b_di is not None:
                lo, hi = (a_di, b_di) if a_di <= b_di else (b_di, a_di)
                open_intervals.append((lo, hi))

    # 3) Regroupements (a,b,c)/mm(/yyyy)  → relâche
    for m in re.finditer(r"\(\s*([\d\s,]+)\s*\)\s*(?:/(\d{1,2})(?:/(\d{2,4}))?)?", txt):
        jours_txt, mm_txt, yy_txt = m.groups()
        mm_def = int(mm_txt) if mm_txt else def_m
        if yy_txt:
            yy_def = int(yy_txt); yy_def = _y2k(yy_def) if yy_def < 100 else yy_def
        else:
            yy_def = def_y
        try:
            jours = [int(x.strip()) for x in jours_txt.split(",") if x.strip()]
        except Exception:
            jours = []
        for jd in jours:
            di = _mk_dateint(yy_def, mm_def, jd)
            if di is not None:
                regroup_days.append(di)

    # 4) Jours isolés de relâche "22/10" (hors parenthèses), séparés par virgules
    for part in [p.strip() for p in txt.split(",")]:
        if not part or "jour" in part:  # ignore parité
            continue
        if re.search(r"^\[.*\]$|^<.*>$|^\(.*\)$", part):  # saute si déjà capturé
            continue
        mday = re.fullmatch(r"(\d{1,2})(?:/(\d{1,2})(?:/(\d{2,4}))?)?", part)
        if not mday:
            continue
        d = int(mday.group(1))
        mm = int(mday.group(2)) if mday.group(2) else def_m
        if mday.group(3):
            yy = int(mday.group(3)); yy = _y2k(yy) if yy < 100 else yy
        else:
            yy = def_y
        di = _mk_dateint(yy, mm, d)
        if di is not None:
            regroup_days.append(di)

    # ----- Décision -----

    # (1) relâche explicite : fermé / regroupements / parité
    for lo, hi in closed_intervals:
        if lo <= dv <= hi:
            return False
    if dv in set(regroup_days):
        return False
    if parite_relache in ("pair", "impair"):
        is_even = (dd % 2 == 0)
        if (parite_relache == "pair" and is_even) or (parite_relache == "impair" and not is_even):
            return False

    # (2) fenêtres de jeu présentes ? -> on ne joue QUE dedans
    if open_intervals:
        for lo, hi in open_intervals:
            if lo <= dv <= hi:
                return True
        return False

    # (3) par défaut : joué
    return True

# Création de la liste des créneaux avant/après pour chaque activité programmée 
# le df des activités programmées est supposé etre trié par jour ("Date") et par heure de début ("Debut")
def get_creneaux(df, activites_programmees, traiter_pauses):

    def creer_creneau(row, borne_min, borne_max, avant, apres, type_creneau):
        titre = row["Activite"] if not pd.isna(row["Activite"]) else ""
        date_str = str(int(row["Date"])) if pd.notnull(row["Date"]) else ""
        return {
            "Date": date_str, # str pour ne pas avoir d'icone de filtre sur la colonne
            "Debut": borne_min.strftime('%Hh%M'),
            "Fin": borne_max.strftime('%Hh%M'),
            "Activité avant": avant,
            "Activité après": apres,
            "__type_creneau": type_creneau,
            "__index": row.name,
            "__uuid": str(uuid.uuid4())
        }
    
    params_to_hash = [
        traiter_pauses, 
        st.session_state.get("MARGE", MARGE).total_seconds(), 
        st.session_state.get("DUREE_REPAS", DUREE_REPAS).total_seconds(), 
        st.session_state.get("DUREE_CAFE", DUREE_CAFE).total_seconds(),
        st.session_state.get("periode_a_programmer_debut", BASE_DATE).isoformat(),
        st.session_state.get("periode_a_programmer_fin", BASE_DATE).isoformat(),
    ]

    hash_val  = hash_df(df, colonnes_a_garder=[col for col in df.columns if col not in ["Debut_dt", "Duree_dt", "__uuid"]], params=params_to_hash)
    hash_key = "creneaux__hash"
    key = "creneaux"
    
    if st.session_state.get(hash_key) != hash_val:
        
        creneaux = []
        bornes = []

        # Traitement des jours libres 
        jours_libres = []
        for jour in range(date_to_dateint(st.session_state.periode_a_programmer_debut), date_to_dateint(st.session_state.periode_a_programmer_fin) + 1):
            if jour not in activites_programmees["Date"].values:
                jours_libres.append(jour)
        for jour in jours_libres:
            if exist_activites_programmables(jour, traiter_pauses):
                row = pd.Series({col: None for col in df.columns})
                row["Date"] = jour
                borne_min = datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))
                borne_max = datetime.datetime.combine(BASE_DATE, datetime.time(23, 59))
                creneaux.append(creer_creneau(row, borne_min, borne_max, "", "", "Journée"))

        if len(activites_programmees) > 0:
            # Initialisation de jour_courant au premier jour des activités programmées
            jour_courant = activites_programmees.iloc[0]["Date"]

            for _, row in activites_programmees.iterrows():

                # Heure de début d'activité
                heure_debut = row["Debut_dt"]
                # Heure de fin d'activité
                heure_fin = heure_debut + row["Duree_dt"] if pd.notnull(heure_debut) and pd.notnull(row["Duree_dt"]) else None
                # initialisation du tableau enregistrant pour chaque jour les bornes des creneaux rencontrés pour eviter les doublons
                if row ["Date"] != jour_courant:
                    bornes = []
                    jour_courant = row ["Date"]

                # Ajout des creneaux avant l'activité considérée s'ils existent
                if pd.notnull(heure_debut):
                    if get_activites_programmables_avant(df, activites_programmees, row, traiter_pauses):
                        borne_min, borne_max, pred = get_creneau_bounds_avant(activites_programmees, row)
                        if (borne_min, borne_max) not in bornes:
                            bornes.append((borne_min, borne_max))
                            creneaux.append(creer_creneau(row, borne_min, borne_max, pred["Activite"] if pred is not None else "", row["Activite"], "Avant"))

                # Ajout des creneaux après l'activité considérée s'ils existent
                if pd.notnull(heure_fin):
                    if get_activites_programmables_apres(df, activites_programmees, row, traiter_pauses):
                        borne_min, borne_max, next = get_creneau_bounds_apres(activites_programmees, row)
                        borne_max = borne_max if borne_max is not None else datetime.datetime.combine(BASE_DATE, datetime.time(23, 59))
                        if (borne_min, borne_max) not in bornes:
                            bornes.append((borne_min, borne_max))
                            creneaux.append(creer_creneau(row, borne_min, borne_max, row["Activite"], next["Activite"] if next is not None else "", "Après"))
        creneaux = sorted(creneaux, key=lambda x: int(x["Date"]))
        creneaux = pd.DataFrame(creneaux)
        st.session_state[key] = creneaux
        st.session_state[hash_key] = hash_val
    return st.session_state[key]

# Renvoie les bornes du créneau existant avant une activité donnée par son descripteur ligne_ref
def get_creneau_bounds_avant(activites_programmees, ligne_ref):
    date_ref = ligne_ref["Date"]
    debut_ref = ligne_ref["Debut_dt"] if pd.notnull(ligne_ref["Debut_dt"]) else datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))

    # Chercher l'activité programmée précédente sur le même jour
    programmes_jour_ref = activites_programmees[activites_programmees["Date"] == date_ref]
    programmes_jour_ref = programmes_jour_ref.sort_values(by="Debut_dt")
    prev = programmes_jour_ref[programmes_jour_ref["Debut_dt"] < debut_ref].tail(1)

    # Calculer l'heure de début minimum du créneau
    if not prev.empty:
        prev_fin = datetime.datetime.combine(BASE_DATE, prev["Debut_dt"].iloc[0].time()) + prev["Duree_dt"].iloc[0]
        debut_min = prev_fin
    else:
        debut_min = datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))

    # Calculer l'heure de fin max du créneau
    fin_max = datetime.datetime.combine(BASE_DATE, debut_ref.time())

    return debut_min, fin_max, prev.iloc[0] if not prev.empty else None

# Renvoie les bornes du créneau existant après une activité donnée par son descripteur ligne_ref
# S'il n'y a pas d'activité suivante pour le même jour renvoie None pour fin_max
def get_creneau_bounds_apres(activites_programmees, ligne_ref):
    date_ref = ligne_ref["Date"]
    debut_ref = ligne_ref["Debut_dt"] if pd.notnull(ligne_ref["Debut_dt"]) else datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))
    duree_ref = ligne_ref["Duree_dt"] if pd.notnull(ligne_ref["Duree_dt"]) else datetime.timedelta(0)
    fin_ref = debut_ref + duree_ref if pd.notnull(debut_ref) and pd.notnull(duree_ref) else debut_ref    

    # Ajuster la date_ref au jour de fin si le jour de fin n'est pas le jour de début
    if fin_ref.day != debut_ref.day:
        date_reelle = dateint_to_date(ligne_ref["Date"])
        date_debut_reelle = debut_ref.replace(year=date_reelle.year, month=date_reelle.month, day=date_reelle.day)
        date_fin_reelle = date_debut_reelle + duree_ref 
        date_ref = date_to_int(date_fin_reelle.year, date_fin_reelle.month, date_fin_reelle.day)

    # Chercher l'activité programmée suivante sur le même jour de référence
    programmes_jour_ref = activites_programmees[activites_programmees["Date"] == date_ref]
    programmes_jour_ref = programmes_jour_ref.sort_values(by="Debut_dt")
    next = programmes_jour_ref[programmes_jour_ref["Debut_dt"] + programmes_jour_ref["Duree_dt"] > fin_ref].head(1)

    # Calculer l'heure de fin max du créneau
    if not next.empty:
        fin_max = datetime.datetime.combine(BASE_DATE, next["Debut_dt"].iloc[0].time())
    else:
        fin_max = None # datetime.datetime.combine(BASE_DATE, datetime.time(23, 59))

    # Calculer l'heure de début minimum du créneau
    debut_min = datetime.datetime.combine(BASE_DATE, fin_ref.time())

    return debut_min, fin_max, next.iloc[0] if not next.empty else None

# Renvoie la liste des activités programmables avant une activité donnée par son descripteur ligne_ref
def get_activites_programmables_avant(df, activites_programmees, ligne_ref, traiter_pauses=True):
    date_ref = ligne_ref["Date"]

    proposables = [] 

    debut_min, fin_max, _ = get_creneau_bounds_avant(activites_programmees, ligne_ref)
    if debut_min >= fin_max:
        return proposables  # Pas d'activités programmables avant si le créneau est invalide

    for _, row in df[df["Date"].isna()].iterrows():
        if pd.isna(row["Debut_dt"]) or pd.isna(row["Duree_dt"]):
            continue
        h_debut = datetime.datetime.combine(BASE_DATE, row["Debut_dt"].time())
        h_fin = h_debut + row["Duree_dt"]
        # L'activité doit commencer après debut_min et finir avant fin_max
        if h_debut >= debut_min + st.session_state.MARGE and h_fin <= fin_max - st.session_state.MARGE and est_hors_relache(row["Relache"], date_ref):
            nouvelle_ligne = row.drop(labels=["Debut_dt", "Duree_dt"]).to_dict()
            nouvelle_ligne["__type_activite"] = "ActiviteExistante"
            nouvelle_ligne["__index"] = row.name
            proposables.append(nouvelle_ligne)
    if traiter_pauses:
        ajouter_pauses(proposables, activites_programmees, ligne_ref, "Avant")
    return proposables

# Renvoie la liste des activités programmables après une activité donnée par son descripteur ligne_ref
def get_activites_programmables_apres(df, activites_programmees, ligne_ref, traiter_pauses=True):
    date_ref = ligne_ref["Date"]
    debut_ref = ligne_ref["Debut_dt"] if pd.notnull(ligne_ref["Debut_dt"]) else datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))
    duree_ref = ligne_ref["Duree_dt"] if pd.notnull(ligne_ref["Duree_dt"]) else datetime.timedelta(0)
    fin_ref = debut_ref + duree_ref if pd.notnull(debut_ref) and pd.notnull(duree_ref) else None   

    proposables = []

    debut_min, fin_max, _ = get_creneau_bounds_apres(activites_programmees, ligne_ref) # Attention fin_max est None si créneau se termine apres 23h59
    if fin_max is not None and debut_min >= fin_max:
        return proposables  # Pas d'activités programmables avant si le créneau est invalide

    if fin_ref.day != debut_ref.day:
        return proposables  # Pas d'activités programmables après si le jour a changé

    for _, row in df[df["Date"].isna()].iterrows():
        if pd.isna(row["Debut_dt"]) or pd.isna(row["Duree_dt"]):
            continue
        h_debut = datetime.datetime.combine(BASE_DATE, row["Debut_dt"].time())
        h_fin = h_debut + row["Duree_dt"]
        # L'activité doit commencer après debut_min et finir avant fin_max en tenant compte des marges et des relaches
        if h_debut >= debut_min + st.session_state.MARGE and (fin_max is None or h_fin <= fin_max - st.session_state.MARGE) and est_hors_relache(row["Relache"], date_ref):
            nouvelle_ligne = row.drop(labels=["Debut_dt", "Duree_dt"]).to_dict()
            nouvelle_ligne["__type_activite"] = "ActiviteExistante"
            nouvelle_ligne["__index"] = row.name
            proposables.append(nouvelle_ligne)
    if traiter_pauses:
        ajouter_pauses(proposables, activites_programmees, ligne_ref, "Après")
    return proposables

# Renvoie les activités programmables sur une journée entière donc les activités qui ne sont pas relache ce jour
def get_activites_programmables_sur_journee_entiere(date_ref, traiter_pauses=True):
    proposables = []

    for _, row in st.session_state.activites_non_programmees.iterrows():
        if est_hors_relache(row["Relache"], date_ref):
            nouvelle_ligne = row.drop(labels=["Debut_dt", "Duree_dt"]).to_dict()
            nouvelle_ligne["__type_activite"] = "ActiviteExistante"
            nouvelle_ligne["__index"] = row.name
            proposables.append(nouvelle_ligne)
    
    if traiter_pauses:
        h_dej = datetime.datetime.combine(BASE_DATE, datetime.time(12, 0))
        type_repas = "déjeuner"
        proposables.append(
            completer_ligne({
                "Debut": datetime.datetime.combine(BASE_DATE, datetime.time(12, 0)).strftime('%Hh%M'),
                "Fin": (h_dej + st.session_state.DUREE_REPAS).strftime('%Hh%M'),
                "Duree": duree_str(st.session_state.DUREE_REPAS),
                "Activite": f"Pause {type_repas}",
                "__type_activite": type_repas,
                "__uuid": str(uuid.uuid4()),
            }))
        
        h_dej = datetime.datetime.combine(BASE_DATE, datetime.time(12, 0))
        type_repas = "dîner"
        proposables.append(
            completer_ligne({
                "Debut": datetime.datetime.combine(BASE_DATE, datetime.time(20, 0)).strftime('%Hh%M'),
                "Fin": (h_dej + st.session_state.DUREE_REPAS).strftime('%Hh%M'),
                "Duree": duree_str(st.session_state.DUREE_REPAS),
                "Activite": f"Pause {type_repas}",
                "__type_activite": type_repas,
                "__uuid": str(uuid.uuid4()),
            }))
    
    return proposables

# Renvoie True s'il existe des activités programmables sur une journée entière donc des activités qui ne sont pas relache ce jour
def exist_activites_programmables(date_ref, traiter_pauses=False):
    if traiter_pauses:
        return True
    for _, row in st.session_state.activites_non_programmees.iterrows():
        if est_hors_relache(row["Relache"], date_ref):
            return True
    return False

# Vérifie si une pause d'un type donné est déjà présente pour un jour donné dans le dataframe des activités planiées
def pause_deja_existante(activites_programmees, jour, type_pause):
    activites_programmes_du_jour = activites_programmees[activites_programmees["Date"] == jour]
    return activites_programmes_du_jour["Activite"].astype(str).str.contains(type_pause, case=False, na=False).any() 

# Ajoute les pauses possibles (déjeuner, dîner, café) à une liste d'activités programmables pour une activité donnée par son descripteur ligne_ref
def ajouter_pauses(proposables, activites_programmees, ligne_ref, type_creneau):

    # Pause repas
    def ajouter_pause_repas(proposables, date_ref, debut_min, fin_max, pause_debut_min, pause_debut_max, type_repas):
        if not pause_deja_existante(activites_programmees, date_ref, type_repas):
            if type_creneau == "Avant":
                h_dej = min(max(fin_max - st.session_state.DUREE_REPAS - st.session_state.MARGE, 
                    datetime.datetime.combine(BASE_DATE, pause_debut_min)), 
                    datetime.datetime.combine(BASE_DATE, pause_debut_max))
                if h_dej - st.session_state.MARGE >= debut_min and h_dej + st.session_state.MARGE <= fin_max:
                    nouvelle_ligne = completer_ligne({
                        "Debut": h_dej.strftime('%Hh%M'),
                        "Fin": (h_dej + st.session_state.DUREE_REPAS).strftime('%Hh%M'),
                        "Duree": duree_str(st.session_state.DUREE_REPAS),
                        "Activite": f"Pause {type_repas}",
                        "__type_activite": type_repas,
                        "__uuid": str(uuid.uuid4()),
                    })
                    proposables.append(nouvelle_ligne)
            elif type_creneau == "Après": # Attention : dans ce cas fin_max est None si le créneau se termine apres 23h59
                h_dej = min(max(debut_min + st.session_state.MARGE, 
                    datetime.datetime.combine(BASE_DATE, pause_debut_min)), 
                    datetime.datetime.combine(BASE_DATE, pause_debut_max))
                if h_dej - st.session_state.MARGE >= debut_min and (fin_max is None or h_dej + st.session_state.MARGE <= fin_max):
                    nouvelle_ligne = completer_ligne({
                        "Debut": h_dej.strftime('%Hh%M'),
                        "Fin": (h_dej + st.session_state.DUREE_REPAS).strftime('%Hh%M'),
                        "Duree": duree_str(st.session_state.DUREE_REPAS),
                        "Activite": f"Pause {type_repas}",
                        "__type_activite": type_repas,
                        "__uuid": str(uuid.uuid4()),
                    })
                    proposables.append(nouvelle_ligne)
    
    def ajouter_pause_cafe(proposables, debut_min, fin_max):
        if not est_pause(ligne_ref):
            Lieu_ref = ligne_ref["Lieu"]
            if type_creneau == "Avant":
                i = activites_programmees.index.get_loc(ligne_ref.name)  
                Lieu_ref_prev = activites_programmees.iloc[i - 1]["Lieu"] if i > 0 else None
                h_cafe = fin_max - st.session_state.DUREE_CAFE
                if not pd.isna(Lieu_ref) and not pd.isna(Lieu_ref_prev) and Lieu_ref == Lieu_ref_prev: 
                    # Dans ce cas pas la peine de tenir compte de la marge avec l'activité précédente
                    if h_cafe >= debut_min: 
                        nouvelle_ligne = completer_ligne({
                            "Debut": h_cafe.strftime('%Hh%M'),
                            "Fin": (h_cafe + st.session_state.DUREE_CAFE).strftime('%Hh%M'),
                            "Duree": duree_str(st.session_state.DUREE_CAFE),
                            "Activite": f"Pause café",
                            "__type_activite": "café",
                            "__uuid": str(uuid.uuid4()),
                        })
                        proposables.append(nouvelle_ligne)
                else: 
                    # Dans ce cas on tient compte de la marge avec l'activité précédente sauf si debut_min = 0h00
                    marge_cafe = st.session_state.MARGE if debut_min != datetime.datetime.combine(BASE_DATE, datetime.time(0, 0)) else datetime.timedelta(minutes=0) 
                    if h_cafe >= debut_min + marge_cafe:
                        nouvelle_ligne = completer_ligne({
                            "Debut": h_cafe.strftime('%Hh%M'),
                            "Fin": (h_cafe + st.session_state.DUREE_CAFE).strftime('%Hh%M'),
                            "Duree": duree_str(st.session_state.DUREE_CAFE),
                            "Activite": "Pause café",
                            "__type_activite": "café",
                            "__uuid": str(uuid.uuid4()),
                        })
                        proposables.append(nouvelle_ligne)
            elif type_creneau == "Après": # Attention : dans ce cas fin_max est None si le créneau se termine apres 23h59
                i = activites_programmees.index.get_loc(ligne_ref.name)  
                Lieu_ref_suiv = activites_programmees.iloc[i + 1]["Lieu"] if i < len(activites_programmees) - 1 else None
                h_cafe = debut_min
                if not pd.isna(Lieu_ref) and not pd.isna(Lieu_ref_suiv) and Lieu_ref == Lieu_ref_suiv: 
                    # Dans ce cas pas la peine de tenir compte de la marge avec l'activité suivante 
                    if fin_max is None or h_cafe + st.session_state.DUREE_CAFE <= fin_max: 
                        nouvelle_ligne = completer_ligne({
                            "Debut": h_cafe.strftime('%Hh%M'),
                            "Fin": (h_cafe + st.session_state.DUREE_CAFE).strftime('%Hh%M'),
                            "Duree": duree_str(st.session_state.DUREE_CAFE),
                            "Activite": "Pause café",
                            "__type_activite": "café",
                            "__uuid": str(uuid.uuid4()),
                        })
                        proposables.append(nouvelle_ligne)
                else: 
                    # Dans ce cas on tient compte de la marge avec l'activité suivante sauf si fin_max is None (créneau se termine après 23h59)
                    marge_cafe = st.session_state.MARGE if fin_max is not None else datetime.timedelta(minutes=0)
                    if fin_max is None or h_cafe + st.session_state.DUREE_CAFE <= fin_max - marge_cafe:
                        nouvelle_ligne = completer_ligne({
                            "Debut": h_cafe.strftime('%Hh%M'),
                            "Fin": (h_cafe + st.session_state.DUREE_CAFE).strftime('%Hh%M'),
                            "Duree": duree_str(st.session_state.DUREE_CAFE),
                            "Activite": "Pause café",
                            "__type_activite": "café",
                            "__uuid": str(uuid.uuid4()),
                        })
                        proposables.append(nouvelle_ligne)

    date_ref = ligne_ref["Date"]

    # Récupération des bornes du créneau
    if type_creneau == "Avant":
        debut_min, fin_max, _ = get_creneau_bounds_avant(activites_programmees, ligne_ref)
    elif type_creneau == "Après":
        debut_min, fin_max, _ = get_creneau_bounds_apres(activites_programmees, ligne_ref)
    else:
        raise ValueError("type_creneau doit être 'Avant' ou 'Après'")

    # Pause déjeuner
    ajouter_pause_repas(proposables, date_ref, debut_min, fin_max, PAUSE_DEJ_DEBUT_MIN, PAUSE_DEJ_DEBUT_MAX, "déjeuner")

    # Pause dîner
    ajouter_pause_repas(proposables, date_ref, debut_min, fin_max, PAUSE_DIN_DEBUT_MIN, PAUSE_DIN_DEBUT_MAX, "dîner")

    # Pause café
    ajouter_pause_cafe(proposables, debut_min, fin_max)

def serialiser_contexte(df, ca=None):
    """
    Écrit df_sorted dans la feuille openpyxl ws ou à défaut dans un buffer, en alignant par nom de 
    colonne (désaccentué/minuscule). Crée les colonnes manquantes à la fin si create_missing=True.
    La colonne 'Hyperlien' n’est pas écrite mais sert à poser un lien sur la colonne Activité.
    """
    def _norm(s: str) -> str:
        if s is None:
            return ""
        s = str(s).strip()
        s = unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("ascii")
        s = re.sub(r"\s+", " ", s).lower()
        return s

    def _to_excel_value(v):
        if pd.isna(v):
            return None
        try:
            iv = int(v)
            if str(iv) == str(v).strip():
                return iv
        except Exception:
            pass
        return v

    # Réindexer proprement pour éviter les trous
    df_sorted = df.copy()
    df_sorted = df_sorted.sort_values(by=["Date", "Debut_dt"])
    df_sorted = df_sorted.reset_index(drop=True)
    hyperliens = df_sorted["Hyperlien"] if "Hyperlien" in df_sorted.columns else None
    df_sorted = df_sorted.drop(columns=["Hyperlien", "Debut_dt", "Duree_dt", "__options_date", "__uuid"], errors='ignore')
    df_sorted["Date"] = df_sorted["Date"].apply(lambda v: dateint_to_str(v) if pd.notna(v) else "")

    # Récupération de la worksheet à traiter
    wb = st.session_state.get("wb")

    if wb is not None:
        try:
            ws = wb.worksheets[0]

            # Effacer le contenu de la feuille Excel existante
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

            # Réinjecter les données du df dans la feuille Excel
            col_activite = None
            for cell in ws[1]:
                if cell.value and str(cell.value).strip().lower() in ["activité"]:
                    col_activite = cell.column
            source_font = ws.cell(row=1, column=1).font

            # Normalise les entêtes existants dans la feuille
            ws_headers = [cell.value for cell in ws[1]]
            header_map = { _norm(h): j for j, h in enumerate(ws_headers, start=1) if h }

            # Liste des colonnes du DF à écrire (on exclut Hyperlien)
            write_cols = [c for c in df_sorted.columns if _norm(c) != "hyperlien"]

            # Associer chaque colonne du DF à un index de colonne Excel (création si nécessaire)
            col_to_ws_idx = {}
            next_col_idx = len(ws_headers) + 1
            for col in write_cols:
                key = _norm(col)
                col_idx = header_map.get(key)
                if col_idx is None:
                    col_idx = next_col_idx
                    ws.cell(row=1, column=col_idx).value = col
                    header_map[key] = col_idx
                    next_col_idx += 1
                col_to_ws_idx[col] = col_idx  # peut rester None si on ne crée pas

            # Colonne Activité dans la feuille
            ACTIVITE_KEYS = {"activite"}
            activite_ws_col = None
            for k, j in header_map.items():
                if k in ACTIVITE_KEYS:
                    activite_ws_col = j
                    break
            if activite_ws_col is None:
                for col in write_cols:
                    if _norm(col) in ACTIVITE_KEYS and col_to_ws_idx.get(col):
                        activite_ws_col = col_to_ws_idx[col]
                        break

            # Écriture des lignes
            for i, (_, row) in enumerate(df_sorted.iterrows(), start=2):
                # 1) valeurs normales
                for col in write_cols:
                    col_idx = col_to_ws_idx.get(col)
                    if not col_idx:
                        continue
                    cell = ws.cell(row=i, column=col_idx)
                    cell.value = _to_excel_value(row[col])

                # 2) hyperlien si présent
                if activite_ws_col and "Hyperlien" in df_sorted.columns:
                    url = row.get("Hyperlien")
                    c = ws.cell(row=i, column=activite_ws_col)
                    if isinstance(url, str) and url.strip():
                        c.hyperlink = url
                        c.font = Font(color="0000EE", underline="single")
                    else:
                        c.hyperlink = None
                        c.font = copy(source_font)  # réinitialisation du style

            # -------- Sheet 1 : adrs (nouveau) --------
            # crée/attrape la feuille 1
            if len(wb.worksheets) > 1:
                ws_adrs = wb.worksheets[1]
                ws_adrs.title = "adrs"
            else:
                ws_adrs = wb.create_sheet(title="adrs", index=1)

            # Clear la feuille adrs
            if ws_adrs.max_row > 0:
                ws_adrs.delete_rows(1, ws_adrs.max_row)

            # Écrire le DataFrame ca s'il existe, sinon laisser vide avec entête vide
            if isinstance(ca, pd.DataFrame) and not ca.empty:
                # écrire entêtes + lignes
                ca_clean = ca.drop(columns=["__uuid"], errors='ignore')
                for r in dataframe_to_rows(ca_clean, index=False, header=True):
                    ws_adrs.append(r)
            else:
                # on laisse la feuille vide (ou ajoute juste l'en-tête si tu préfères)
                pass

            # Sauvegarde dans buffer mémoire
            buffer = io.BytesIO()
            wb.save(buffer)

            # Revenir au début du buffer pour le téléchargement
            buffer.seek(0)
            return buffer
        
        except Exception as e:
            print(f"Erreur dans serialiser_contexte via Dropbox: {e}")

    # Sauvegarde simple dans buffer mémoire en cas d'échec avec workbook Dropbox
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_sorted.to_excel(writer, sheet_name="data", index=False)
        if isinstance(ca, pd.DataFrame) and not ca.empty:
            ca.to_excel(writer, sheet_name="adrs", index=False)
        else:
            # crée quand même la feuille "adrs" vide
            pd.DataFrame().to_excel(writer, sheet_name="adrs", index=False)

    # Revenir au début du buffer pour le téléchargement
    buffer.seek(0)
    return buffer

def sauvegarder_contexte(df_hash=None):


    # Version modale
    @st.dialog("Sauvegarder données")
    def show_dialog_sauvegarder_contexte(nom_fichier, df_hash=None, ca_hash=None):
        st.markdown("Voulez-vous sauvegarder les données ?")
        col1, col2 = st.columns([1, 1])
        with col1:
            df = st.session_state.get("df")
            ca = st.session_state.get("ca")
            if df_hash is None:
                df_hash = hash_df(df, colonnes_a_enlever=["Debut_dt", "Duree_dt", "__options_date", "__uuid"])
            ca_hash = hash_df(ca)

            combo_hash = (df_hash, ca_hash)
            combo_hash_prev = st.session_state.get("__contexte_combo_hash")
            buffer = st.session_state.get("__contexte_buffer")

            if (combo_hash != combo_hash_prev or buffer is None) and est_contexte_valide():
                buffer = serialiser_contexte(df, ca)
                st.session_state["__contexte_combo_hash"] = combo_hash
                st.session_state["__contexte_buffer"] = buffer

            # Bouton de téléchargement
            if st.download_button(
                label="Valider",
                data=st.session_state["__contexte_buffer"],
                file_name=nom_fichier,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=CENTRER_BOUTONS
            ):
                st.rerun()
        with col2:
            if st.button(LABEL_BOUTON_ANNULER, use_container_width=CENTRER_BOUTONS):
                st.rerun()

    nom_fichier = st.session_state.get("fn", "planning_avignon.xlsx")
    if st.button(
        label=LABEL_BOUTON_SAUVEGARDER,
        use_container_width=CENTRER_BOUTONS,
        disabled=not est_contexte_valide()):
        show_dialog_sauvegarder_contexte(nom_fichier)

    # Version Non Modale
    # nom_fichier = st.session_state.get("fn", "planning_avignon.xlsx")
    # df = st.session_state.get("df")
    # ca = st.session_state.get("ca")
    
    # if df_hash is None:
    #     df_hash = hash_df(df, colonnes_a_enlever=["Debut_dt", "Duree_dt", "__options_date", "__uuid"])
    # ca_hash = hash_df(ca)

    # combo_hash = (df_hash, ca_hash)
    # combo_hash_prev = st.session_state.get("__contexte_combo_hash")
    # buffer = st.session_state.get("__contexte_buffer")

    # if (combo_hash != combo_hash_prev or buffer is None) and est_contexte_valide():
    #     buffer = serialiser_contexte(df, ca)
    #     st.session_state["__contexte_combo_hash"] = combo_hash
    #     st.session_state["__contexte_buffer"] = buffer

    # # Bouton de téléchargement
    # st.download_button(
    #     label=LABEL_BOUTON_SAUVEGARDER,
    #     data=st.session_state.get("__contexte_buffer", ""),
    #     file_name=nom_fichier,
    #     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    #     use_container_width=CENTRER_BOUTONS,
    #     disabled=not est_contexte_valide()
    # )

# Programme une activité non programmée à une date donnée
def programmer_activite_non_programmee(date_ref, activite):

    df = st.session_state.df
    type_activite = activite["__type_activite"]
    undo.save()
    if type_activite == "ActiviteExistante":
        # Pour les spectacles, on programme la date et l'heure
        index = activite["__index"]
        modifier_cellule(index, "Date", date_ref)
    elif type_activite == "déjeuner":
        # Pour les pauses, on ne programme pas d'heure spécifique
        index = len(df)
        ajouter_activite(
            idx=index, 
            nom="Pause déjeuner",
            jour=date_ref, 
            debut=activite["Debut"],
            duree=formatter_timedelta(st.session_state.DUREE_REPAS),
            )
    elif type_activite == "dîner":
        # Pour les pauses, on ne programme pas d'heure spécifique
        index = len(df)
        ajouter_activite(
            idx=index, 
            nom="Pause dîner",
            jour=date_ref, 
            debut=activite["Debut"],
            duree=formatter_timedelta(st.session_state.DUREE_REPAS),
            )
    elif type_activite == "café":
        # Pour les pauses, on ne programme pas d'heure spécifique
        index = len(df)
        ajouter_activite(
            idx=index, 
            nom="Pause café",
            jour=date_ref, 
            debut=activite["Debut"],
            duree=formatter_timedelta(st.session_state.DUREE_CAFE),
            )
    else:
        return

    demander_selection("activites_programmees", index, deselect="activites_non_programmees")
    demander_selection("creneaux_disponibles", get_creneau_proche(st.session_state.creneaux_disponibles, activite)[0])
    st.session_state["activites_programmables_selected_row"] = None
    # forcer_reaffichage_df("creneaux_disponibles")
    sql.sauvegarder_row(index)
    st.rerun()

# Renvoie les jours possibles pour programmer une activité donnée par son idx
def get_jours_possibles(df, activites_programmees, idx_activite):
    try:
        jours_possibles = []

        # Retour si index non valide
        if idx_activite not in df.index:
            return jours_possibles

        # Récupérer la durée de l'activité à considérer
        ligne_a_considerer = df.loc[idx_activite]
        debut = ligne_a_considerer["Debut_dt"]
        fin = ligne_a_considerer["Debut_dt"] + ligne_a_considerer["Duree_dt"]

        if activites_programmees is not None:
            for jour in range(date_to_dateint(st.session_state.periode_a_programmer_debut), date_to_dateint(st.session_state.periode_a_programmer_fin) + 1):
                
                if not est_hors_relache(ligne_a_considerer["Relache"], jour):
                    continue

                activites_programmes_du_jour = activites_programmees[activites_programmees["Date"] == jour].sort_values("Debut_dt")

                if not activites_programmes_du_jour.empty:
                    # Créneau entre minuit et première activité du jour
                    premiere_activite_du_jour = activites_programmes_du_jour.iloc[0]
                    borne_inf = datetime.datetime.combine(BASE_DATE, datetime.time.min)  # 00h00
                    borne_sup = premiere_activite_du_jour["Debut_dt"]
                    if debut >= borne_inf and fin <= borne_sup - st.session_state.MARGE:
                        jours_possibles.append(jour)
                        continue  # on prend le premier créneau dispo du jour

                    # Ensuite, créneaux entre chaque activité programmée
                    for _, ligne in activites_programmes_du_jour.iterrows():
                        borne_inf, borne_sup, _ = get_creneau_bounds_apres(activites_programmes_du_jour, ligne)
                        if debut >= borne_inf + st.session_state.MARGE and (borne_sup is None or fin <= borne_sup - st.session_state.MARGE):
                            jours_possibles.append(jour)
                            break  # jour validé, on passe au suivant
                else: # jour libre
                    jours_possibles.append(jour)
    except Exception as e:
        print(f"Erreur in get_jours_possibles : {e}")
    return jours_possibles

# Indique si un jour donné est possible pour programmer une activité donnée par son idx
def est_jour_possible(df, activites_programmees, idx_activite, jour):
    try:
        # Retour si index non valide
        if idx_activite not in df.index:
            return False

        # Récupérer la durée de l'activité à considérer
        ligne_a_considerer = df.loc[idx_activite]
        debut = ligne_a_considerer["Debut_dt"]
        fin = ligne_a_considerer["Debut_dt"] + ligne_a_considerer["Duree_dt"]

        if activites_programmees is not None:
                
            if not est_hors_relache(ligne_a_considerer["Relache"], jour):
                return False

            activites_programmes_du_jour = activites_programmees[activites_programmees["Date"] == jour].sort_values("Debut_dt")

            if not activites_programmes_du_jour.empty:
                # Créneau entre minuit et première activité du jour
                premiere_activite_du_jour = activites_programmes_du_jour.iloc[0]
                borne_inf = datetime.datetime.combine(BASE_DATE, datetime.time.min)  # 00h00
                borne_sup = premiere_activite_du_jour["Debut_dt"]
                if debut >= borne_inf + st.session_state.MARGE and fin <= borne_sup - st.session_state.MARGE:
                    return True

                # Ensuite, créneaux entre chaque activité programmée
                for _, ligne in activites_programmes_du_jour.iterrows():
                    borne_inf, borne_sup, _ = get_creneau_bounds_apres(activites_programmes_du_jour, ligne)
                    if debut >= borne_inf + st.session_state.MARGE and (borne_sup is None or fin <= borne_sup - st.session_state.MARGE):
                        return True
            else: # jour libre
                return True
    except Exception as e:
        print(f"Erreur in est_jour_possible : {e}")
    return False

# idem get_jours_possibles avec en paramètre une row d'activité programmée contenant en colonne __index l'index du df de base
# Les paramètres df et activites_programmees sont supposés etre stockés dans st.session_state
def get_jours_possibles_from_activite_programmee(row: pd.Series):
    jours = get_jours_possibles(st.session_state.df, st.session_state.activites_programmees, row["__index"])
    jour_courant = int(row["Date"]) if pd.notna(row["Date"]) and row["Date"] is not None else row["Date"]
    if pd.notna(row["__index"]) and row["__index"] in st.session_state.df.index:
        if not est_activite_reserve(st.session_state.df.loc[row["__index"]]):
            jours = [jour_courant] + jours + [""] if jours != [] else [jour_courant] + [""]
        else: 
            jours = []
    return sorted([str(j) for j in jours]) if isinstance(jours, list) else []

# idem get_jours_possibles avec en paramètre une row d'activité non programmée contenant en colonne __index l'index du df de base
# Les paramètres df et activites_programmees sont supposés etre stockés dans st.session_state
def get_jours_possibles_from_activite_non_programmee(row: pd.Series):
    jours = get_jours_possibles(st.session_state.df, st.session_state.activites_programmees, row["__index"])
    jours = [""] + jours if jours != [] else []
    return [str(j) for j in jours] if isinstance(jours, list) else []

# Calcule les options des dates pour les activiés programmées
# Programme une activité choisie en fonction des jours possibles
def programmer_activite_par_choix_activite():

    df = st.session_state.get("df")
    if df is None or len(df) <= 0:
        return

    st.markdown("##### Programmation d'une nouvelle activité")

    # Filtrer les activités non programmées
    activites_programmees = st.session_state.get("activites_programmees")
    activites_non_programmees = st.session_state.get("activites_non_programmees")

    # Liste d'options formatées
    options_activites = []
    for idx, row in activites_non_programmees.iterrows():
        if get_jours_possibles(df, activites_programmees, idx):
            label = f"[{row["Debut"]} - {row["Fin"]}] - {str(row["Activite"]).strip()}"
            options_activites.append((label, idx))

    # Afficher la selectbox des activités
    activite_selectionee = st.selectbox("Choix de l'activité à programmer :", options_activites, format_func=lambda x: x[0])
    if activite_selectionee:
        idx_choisi = activite_selectionee[1]

        # Déterminer les jours disponibles 
        jours_possibles = get_jours_possibles(df, activites_programmees, idx_choisi)
        jours_label = [dateint_to_str(x) for x in jours_possibles]

        jour_selection = st.selectbox("Choix du jour :", jours_label)

        # Bouton pour confirmer
        if jour_selection:
            if st.button(LABEL_BOUTON_PROGRAMMER, key="AjouterAuPlanningParChoixActivite"):
                jour_choisi = date_to_dateint(jour_selection)

                # On peut maintenant modifier le df
                df.at[idx_choisi, "Date"] = jour_choisi
                st.rerun()

def get_creneau_proche(creneaux: pd.DataFrame, activite):
    """
    A partir d'une liste de créneaux, renvoie le créneau le plus proche d'une activité donnée selon les critères suivants:
      1 : Date activité manquante → premier créneau qui contient l’activité
      2 : même jour, créneau dont Début >= Fin activité
      3 : même jour, créneau qui contient totalement l’activité
      4 : même jour, créneau dont Fin <= Début activité
      5 : jour futur le plus proche
      6 : jour passé le plus proche
      fallback : aucune correspondance, première ligne
    
    S’il n’y a aucun candidat (i.e. vide ou Date non utilisables) -> (index de la 1ère ligne, ligne, 'fallback').
    
    Paramètres:
    - creneaux: liste de creneaux fournie sous la forme d'un DataFrame tel que renvoyé par get_creneaux
    - activite: activité
    
    retours:
    - index du créneau sélectionné
    - créneau sélectionné
    - critère de choix
    """
    def _interval_distance_to_window(d, f, win_start, win_end):
        """
        Distance entre l'intervalle [d,f] et la fenêtre [win_start, win_end].
        0 si recouvrement; sinon la distance minimale entre bords.
        Robuste si d ou f manquent (on réduit à un point).
        """
        if pd.isna(d) and pd.isna(f):
            return 10**9
        if pd.isna(d): d = f
        if pd.isna(f): f = d
        if d > f:
            d, f = f, d
        # Overlap ?
        if not (f < win_start or d > win_end):
            return 0
        # Sinon distance au plus proche bord
        if f < win_start:
            return win_start - f
        else:  # d > win_end
            return d - win_end

    if creneaux is None or creneaux.empty or activite is None:
        return None, None, None

    hdeb = activite.get("Debut")
    hfin = activite.get("Fin")
    win_start = hhmm_to_min(hdeb)
    win_end   = hhmm_to_min(hfin)

    jour = safe_int(activite.get("Date"))

    work = creneaux.copy()
    work["Date"]   = pd.to_numeric(work["Date"], errors="coerce")
    work["_debut"] = work["Debut"].map(hhmm_to_min)
    work["_fin"]   = work["Fin"].map(hhmm_to_min)

    # corrige inversions Debut/Fin
    mask_swap = work["_debut"].notna() & work["_fin"].notna() & (work["_fin"] < work["_debut"])
    if mask_swap.any():
        tmp = work.loc[mask_swap, "_debut"].copy()
        work.loc[mask_swap, "_debut"] = work.loc[mask_swap, "_fin"]
        work.loc[mask_swap, "_fin"]   = tmp

    # ---------- 1 : Date activité manquante → créneau qui contient l’activité
    if jour is None or pd.isna(jour):
        for idx, creneau in work.iterrows():
            p = get_proposables(creneau, traiter_pauses=False)
            if activite["__uuid"] in p["__uuid"].values:
                return idx, creneau, "1"
        idx0 = creneaux.index[0]
        return idx0, creneaux.loc[idx0], "fallback"

    # ---------- 2 : même jour, Debut >= Fin activité
    j = int(jour)
    if win_end is not None:
        r2 = work[(work["Date"] == j) & (work["_debut"].notna()) & (work["_debut"] >= win_end)]
        if not r2.empty:
            idx = (r2["_debut"] - win_end).idxmin()
            return idx, creneaux.loc[idx], "2"

    # ---------- 3 : même jour, créneau contenant l’activité
    if win_start is not None and win_end is not None:
        r3 = work[(work["Date"] == j)
                 & work["_debut"].notna() & work["_fin"].notna()
                 & (work["_debut"] <= win_start) & (work["_fin"] >= win_end)]
        if not r3.empty:
            slack_left  = win_start - r3["_debut"]
            slack_right = r3["_fin"] - win_end
            r3 = r3.assign(_slack_total = slack_left + slack_right,
                           _slack_left = slack_left,
                           _slack_right = slack_right)
            cand = r3.sort_values(
                by=["_slack_total", "_slack_left", "_slack_right", "_debut", "_fin"],
                ascending=[True, True, True, True, True]
            ).iloc[0]
            return cand.name, creneaux.loc[cand.name], "3"

    # ---------- 4 : même jour, Fin <= Début activité
    if win_start is not None:
        r4 = work[(work["Date"] == j) & (work["_fin"].notna()) & (work["_fin"] <= win_start)]
        if not r4.empty:
            idx = (win_start - r4["_fin"]).idxmin()
            return idx, creneaux.loc[idx], "4"

    # distance au créneau pour futur/passé
    work["_win_dist"] = work.apply(
        lambda r: _interval_distance_to_window(r["_debut"], r["_fin"], win_start, win_end), axis=1
    )

    # ---------- 5 : jour futur le plus proche
    r5 = work[(work["Date"].notna()) & (work["Date"] >= j)]
    if not r5.empty:
        r5 = r5.assign(_day_dist=(r5["Date"] - j).astype("int64"))
        cand = r5.sort_values(by=["_day_dist", "_win_dist", "_debut", "_fin"],
                              ascending=[True, True, True, True]).iloc[0]
        return cand.name, creneaux.loc[cand.name], "5"

    # ---------- 6 : jour passé le plus proche
    r6 = work[(work["Date"].notna()) & (work["Date"] <= j)]
    if not r6.empty:
        r6 = r6.assign(_day_dist=(j - r6["Date"]).astype("int64"))
        cand = r6.sort_values(by=["_day_dist", "_win_dist", "_debut", "_fin"],
                              ascending=[True, True, True, True]).iloc[0]
        return cand.name, creneaux.loc[cand.name], "6"

    # ---------- fallback
    idx0 = creneaux.index[0]
    return idx0, creneaux.loc[idx0], "fallback"

def get_proposables(creneau, traiter_pauses=False):

    proposables = []

    df = st.session_state.get("df")
    if df is None or len(df) <= 0:
        return proposables

    type_creneau = creneau["__type_creneau"]
    idx = creneau["__index"]
    date_ref = int(creneau["Date"]) # date_ref doit être en int !

    if type_creneau == "Avant":
        activites_programmees = st.session_state.get("activites_programmees")
        if activites_programmees is None:
            return proposables
        try:
            ligne_ref = activites_programmees.loc[idx]
        except Exception as e:
            print(f"Erreur afficher_creneaux_disponibles : {e}")
            return proposables
        proposables = get_activites_programmables_avant(df, activites_programmees, ligne_ref, traiter_pauses)

    elif type_creneau == "Après":
        activites_programmees = st.session_state.get("activites_programmees")
        if activites_programmees is None:
            return proposables
        try:
            ligne_ref = activites_programmees.loc[idx]
        except Exception as e:
            print(f"Erreur afficher_creneaux_disponibles : {e}")
            return proposables
        proposables = get_activites_programmables_apres(df, activites_programmees, ligne_ref, traiter_pauses)

    elif type_creneau == "Journée":
        proposables = get_activites_programmables_sur_journee_entiere(date_ref, traiter_pauses)

    proposables = pd.DataFrame(proposables).sort_values(by=["Debut"], ascending=[True]) if proposables else pd.DataFrame(proposables)
    proposables["Date"] = creneau["Date"] # ou str(date_ref) car col Date au format string dans les df_display !

    return proposables

# Initialisation des variables d'état du contexte après chargement des données du contexte
def initialiser_etat_contexte(df, wb, fn, fp, ca):
    st.session_state.df = df
    st.session_state.wb = wb
    st.session_state.fn = fn
    st.session_state.fp = fp
    st.session_state.ca = ca
    st.session_state.nouveau_fichier = True
    st.session_state.compteur_activite = 0
    st.session_state.menu_activites = {"menu": "menu_activites_non_programmees", "index_df": None}
    st.session_state.menu_activites_programmees = None
    st.session_state.menu_activites_non_programmees = None
    st.session_state.forcer_menu_activites_programmees = False
    st.session_state.forcer_menu_activites_non_programmees = False
    st.session_state.forcer_maj_menu_activites_programmees = False
    st.session_state.forcer_maj_menu_activites_non_programmees = copy.deepcopy(SEL_REQUEST_DEFAUT)
    st.session_state.activites_programmees_sel_request = copy.deepcopy(SEL_REQUEST_DEFAUT)
    st.session_state.activites_non_programmees_sel_request = copy.deepcopy(SEL_REQUEST_DEFAUT)
    st.session_state.creneaux_disponibles_sel_request = copy.deepcopy(SEL_REQUEST_DEFAUT)
    st.session_state.activites_programmables_sel_request =copy.deepcopy(SEL_REQUEST_DEFAUT)

    # forcer_reaffichage_df("creneaux_disponibles")

# Charge le fichier Excel contenant les activités à programmer
def charger_contexte_depuis_fichier():

    @st.dialog("Charger fichier")
    def dialog_charger_fichier():
        # Chargement du fichier Excel contenant les activités à programmer
        fd = st.file_uploader(
            "Choix du fichier Excel contenant les activités à programmer", 
            type=["xlsx"], 
            label_visibility="collapsed",
            key="file_uploader",
        )

        if st.button("Charger", use_container_width=CENTRER_BOUTONS, disabled=fd is None):

            try:
                st.session_state.contexte_invalide = True
                curseur_attente()
                df = pd.read_excel(fd)
                df["Date"] = df["Date"].apply(lambda v: date_to_dateint(v, datetime.date.today().year, datetime.date.today().month))
                wb = load_workbook(fd)
                lnk = get_liens_activites(wb)
                sheetnames = wb.sheetnames
                ca = pd.read_excel(fd, sheet_name=sheetnames[1]) if len(sheetnames) > 1 else None
                df = _nettoyer_donnees(df, fd.name) # si ok RAZ du contexte_invalide

                if "contexte_invalide" not in st.session_state:
                    df = add_persistent_uuid(df)
                    df = add_hyperliens(df, lnk)
                    ca = nettoyer_ca(ca)
                    fn = fd.name if fd is not None else ""
                    fp = dp.upload_excel_to_dropbox(fd.getvalue(), fd.name) if fd is not None else ""
                    undo.save()
                    initialiser_etat_contexte(df, wb, fn, fp, ca)
                    initialiser_periode_programmation(df)
                    st.session_state["push_periode_programmation_modele_values"] = True 
                    # undo.init(verify=False)
                    maj_contexte(maj_donnees_calculees=True)
                    sql.sauvegarder_contexte()
                    selection = st.session_state.activites_non_programmees.index[0] if len(st.session_state.activites_non_programmees) > 0 else None
                    demander_selection("activites_non_programmees", selection, deselect="activites_programmees")
                    st.session_state.menu_activites = {
                        "menu": "menu_activites_non_programmees",
                        "index_df": selection
                    }
                    st.session_state.forcer_maj_menu_activites_non_programmees = True
                    st.rerun()

            except Exception as e:
                st.sidebar.error(f"Erreur de chargement du fichier : {e}")

    if st.button("Charger", use_container_width=CENTRER_BOUTONS):
        dialog_charger_fichier()

# Initialisation des types d'un df vide
def initialiser_dtypes(df):
    for col in df.columns:
        if col in COLONNES_TYPE_INT:
            df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
        elif col in COLONNES_TYPE_STRING:
            df[col] = df[col].astype("string")
        elif col in COLONNES_TYPE_OBJECT:
            df[col] = df[col].astype("object")
    if "Debut_dt" not in df.columns:
        df["Debut_dt"] = pd.Series(dtype="datetime64[ns]")
    else:
        df["Debut_dt"] = df["Debut_dt"].astype("datetime64[ns]")
    if "Duree_dt" not in df.columns:
        df["Duree_dt"] = pd.Series(dtype="timedelta64[ns]")
    else:
        df["Duree_dt"] = df["Duree_dt"].astype("timedelta64[ns]")

# Initialisation d'un nouveau contexte
def initialiser_nouveau_contexte():

    if "contexte_invalide" in st.session_state:
        del st.session_state["contexte_invalide"]

    df = pd.DataFrame(columns=COLONNES_ATTENDUES)
    df = add_persistent_uuid(df)
    df = add_hyperliens(df)
    initialiser_dtypes(df)
    wb = None
    fn = "planning_avignon.xlsx"
    fp = ""
    ca = pd.DataFrame(columns=COLONNES_ATTENDUES_CARNET_ADRESSES)
    ca = add_persistent_uuid(ca)
    
    initialiser_etat_contexte(df, wb, fn, fp, ca)
    initialiser_periode_programmation(df)
    sql.sauvegarder_contexte()

# Création d'un nouveau contexte
def creer_nouveau_contexte():
    if st.button(LABEL_BOUTON_NOUVEAU, use_container_width=CENTRER_BOUTONS, key="creer_nouveau_contexte"):
        curseur_attente()
        undo.save()
        initialiser_nouveau_contexte()
        maj_contexte(maj_donnees_calculees=True)
        st.rerun()

# Indique si le contexte est vlide pour traitement
def est_contexte_valide():
    return "df" in st.session_state and isinstance(st.session_state.df, pd.DataFrame) and "contexte_invalide" not in st.session_state

# Section critique pour la déprogrammation d'une activité programmée.
# Section critique car la modification de cellule depuis la grille est validée par un click row 
# qui peut entraîner une interruption du script python et donc une incohérence de contexte.
# Le mécanisme de section critique permet une relance automatique du traitement jusqu'à complétion 
# en cas d'interruption par un rerun Streamlit : une commande est enregistrée dans st.session_state 
# et est automatiquement relancée en début de rerun tant qu'elle n'est pas terminée.
def activites_programmees_deprogrammer(idx):
    
    st.session_state.setdefault("activites_programmees_deprogrammer_cmd", 
        {
            "idx": idx,
            "step": 0,
        }
    )

    tracer.log(f"Début {idx}")

    st.session_state.forcer_menu_activites_non_programmees = True
    deprogrammer_activite_programmee(idx)

    # Workaround pour forcer le réaffichage de la grille.
    # Sinon figeage grille après modification de cellule.
    forcer_reaffichage_activites_programmees() 

    # forcer_reaffichage_df("creneaux_disponibles")
    sql.sauvegarder_row(idx)

    tracer.log(f"Fin {idx}")
    del st.session_state["activites_programmees_deprogrammer_cmd"]

# Section critique pour la reprogrammation d'une activité programmée.
# Section critique car la modification de cellule depuis la grille est validée par un click row 
# qui peut entraîner une interruption du script python et donc une incohérence de contexte.
# Le mécanisme de section critique permet une relance automatique du traitement jusqu'à complétion 
# en cas d'interruption par un rerun Streamlit : une commande est enregistrée dans st.session_state 
# et est automatiquement relancée en début de rerun tant qu'elle n'est pas terminée.
def activites_programmees_reprogrammer(idx, jour):
    
    st.session_state.setdefault("activites_programmees_reprogrammer_cmd", 
        {
            "idx": idx,
            "jour": jour,
            "step": 0,
        }
    )

    tracer.log(f"Début {idx} {jour}")

    modifier_cellule(idx, "Date", jour)

    # Workaround pour forcer le réaffichage de la grille.
    # Sinon figeage grille après modification de cellule.
    forcer_reaffichage_activites_programmees() 

    sql.sauvegarder_row(idx)

    tracer.log(f"Fin {idx} {jour}")
    del st.session_state["activites_programmees_reprogrammer_cmd"]

# Section critique pour la modification de cellules d'une activité programmée.
# Section critique car la modification de cellule depuis la grille est validée par un click row 
# qui peut entraîner une interruption du script python et donc une incohérence de contexte.
# Le mécanisme de section critique permet une relance automatique du traitement jusqu'à complétion 
# en cas d'interruption par un rerun Streamlit : une commande est enregistrée dans st.session_state 
# et est automatiquement relancée en début de rerun tant qu'elle n'est pas terminée.
def activites_programmees_modifier_cellule(idx, col, val):
    
    st.session_state.setdefault("activites_programmees_modifier_cellule_cmd", 
        {
            "idx": idx,
            "col": col,
            "val": val,
            "step": 0,
        }
    )

    tracer.log(f"Début {idx} {col} {val}")

    erreur = affecter_valeur_df(idx, col, val, section_critique=st.session_state.activites_programmees_modifier_cellule_cmd)

    # Workaround pour forcer le réaffichage de la grille.
    # Sinon figeage grille après modification de cellule.
    forcer_reaffichage_activites_programmees() 

    if not erreur:
        if col in ["Debut", "Duree", "Activité"]:
            # forcer_reaffichage_df("creneaux_disponibles")
            pass
    else:
        st.session_state.aggrid_activites_programmees_erreur = erreur

    tracer.log(f"Fin {idx} {col} {val}")
    del st.session_state["activites_programmees_modifier_cellule_cmd"]

# Section critique pour la programmation d'une activité non programmée.
# Section critique car la modification de cellule depuis la grille est validée par un click row 
# qui peut entraîner une interruption du script python et donc une incohérence de contexte.
# Le mécanisme de section critique permet une relance automatique du traitement jusqu'à complétion 
# en cas d'interruption par un rerun Streamlit : une commande est enregistrée dans st.session_state 
# et est automatiquement relancée en début de rerun tant qu'elle n'est pas terminée.
def activites_non_programmees_programmer(idx, jour):
    
    st.session_state.setdefault("activites_non_programmees_programmer_cmd", 
        {
            "idx": idx,
            "jour": jour,
            "step": 0,
        }
    )

    tracer.log(f"Début {idx} {jour}")

    st.session_state.forcer_menu_activites_programmees = True
    modifier_cellule(idx, "Date", int(jour))

    # Workaround pour forcer le réaffichage de la grille.
    # Sinon figeage grille après modification de cellule.
    forcer_reaffichage_activites_non_programmees() 

    # forcer_reaffichage_df("creneaux_disponibles")
    sql.sauvegarder_row(idx)

    tracer.log(f"Fin {idx} {jour}")
    del st.session_state["activites_non_programmees_programmer_cmd"]

# Section critique pour la modification de cellules d'une activité non programmée.
# Section critique car la modification de cellule depuis la grille est validée par un click row 
# qui peut entraîner une interruption du script python et donc une incohérence de contexte.
# Le mécanisme de section critique permet une relance automatique du traitement jusqu'à complétion 
# en cas d'interruption par un rerun Streamlit : une commande est enregistrée dans st.session_state 
# et est automatiquement relancée en début de rerun tant qu'elle n'est pas terminée.
def activites_non_programmees_modifier_cellule(idx, col, val):
    
    st.session_state.setdefault("activites_non_programmees_modifier_cellule_cmd", 
        {
            "idx": idx,
            "col": col,
            "val": val,
            "step": 0,
        }
    )

    tracer.log(f"Début {idx} {col} {val}")

    erreur = affecter_valeur_df(idx, col, val, section_critique=st.session_state.activites_non_programmees_modifier_cellule_cmd)

    # Workaround pour forcer le réaffichage de la grille.
    # Sinon figeage grille après modification de cellule.
    forcer_reaffichage_activites_non_programmees() 
    
    if not erreur:
        # forcer_reaffichage_df("activites_programmables")
        pass
    else:
        st.session_state.aggrid_activites_non_programmees_erreur = erreur

    tracer.log(f"Fin {idx} {col} {val}")
    del st.session_state["activites_non_programmees_modifier_cellule_cmd"]

