####################
# Core application #
####################

import streamlit as st
import pandas as pd
import datetime
import io
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import requests
from bs4 import BeautifulSoup
import pandas.api.types as ptypes
from st_aggrid import AgGrid, DataReturnMode, GridOptionsBuilder, JsCode, GridUpdateMode
from io import BytesIO
import uuid
import json
from streamlit_javascript import st_javascript
from urllib.parse import quote_plus
import copy
# import pkg_resources

from app_const import *
from app_utils import *
import tracer
import sql_api as sql 
import gsheet_api as gs
import sync_worker as wk
import undo

###########
# JsCodes #
###########

# JsCode chargé de gérer la sélection/déselection programmée de lignes dans les AgGrid, 
# le flip-flop entre grilles "activites_programmees" et "activites_non_programmees" via __sel_source
# et le renvoie correct des modifications de cellules prises en compte par le DOM via response["data"].
# Il exploite les colonnes de travail suivantes:
# __sel_id = id de la ligne à sélectionner (None si pas de contrainte de sélection).
# __sel_ver = version de la demande de sélection (doit être incrémentée à chaque demande).
# __desel_id = id de la ligne devant rester visible lors de la déselection (None si aucune contrainte de visibilité lors de la désélection).
# __desel_ver = version de la demande de désélection (doit être incrémentée à chaque demande).
# __sel_source = information renvoyée par le DOM (event.source exposé par onRowSelected) indiquant si la source de selection est "user" ou "api" selon que la demande de sélection provient d'un click utilisateur ou d'une requête python via JsCode.
# __df_push_ver = permet au JsCode de déclencher un selectionChanged "api" lorsqu'il détecte une incrémentation de la première ligne sur cette colonne, ce qui permet à Streamlit de renvoyer la data mise à jour dans response["data"] sans attendre un clic utilisateur. 
# Ces colonnes sont configurées par les fonctions utilisateur demander_selection(), demander_deselection() et signaler_df_push()
# L'information de retour __sel_source est exploitée par le mecanisme de flip flop entre grilles "activites_programmees" et "activites_non_programmees" via le response["data"] de l'aggrid,
# ceci afin de ne déclencher le changement d'activité sélectionnée que sur clic user (cf. fonction afficher_activites_programmees() et afficher_activites_non_programmees()).
# Ce JsCode doit être branché sur le onGridReady (voir les grid_options configurées avec les fonctions init_grid_options_xxx).
JS_SELECT_DESELECT_ONCE = JsCode(r"""
function(p){
  var api=p&&p.api; if(!api) return;

  // --- identifiant d'instance par IFRAME, pour éviter tout chevauchement entre grilles ---
  var fe = window.frameElement || null;
  var instId = (function(){
    if (!fe) return 'grid-' + Math.random().toString(36).slice(2);
    var v = fe.getAttribute('data-ag-inst');
    if (!v) { v = 'grid-' + Date.now().toString(36).slice(2) + '-' + Math.random().toString(36).slice(2);
              fe.setAttribute('data-ag-inst', v); }
    return v;
  })();

  // --- caches par instance (désélection / sélection) ---
  window.__agSelCache   = window.__agSelCache   || {};
  window.__agDeselCache = window.__agDeselCache || {};
  var selCache   = window.__agSelCache;
  var deselCache = window.__agDeselCache;

  // ======================= Helpers DataFrame/meta =======================
  function firstRow(){
    var n=api.getDisplayedRowCount?api.getDisplayedRowCount():0;
    return (n>0) ? api.getDisplayedRowAtIndex(0) : null;
  }

  function updateMetaIfChanged(key, val){
    try{
      var r0=firstRow(); if(!r0||!r0.data) return;
      if(r0.data[key] === val) return;
      var row=Object.assign({}, r0.data);
      row[key]=val;
      api.applyTransaction && api.applyTransaction({ update:[row] });
    }catch(e){}
  }

  function scan(col){
    var n=api.getDisplayedRowCount?api.getDisplayedRowCount():0;
    for(var i=0;i<n;i++){
      var r=api.getDisplayedRowAtIndex(i);
      if(r&&r.data&&r.data[col]!=null) return String(r.data[col]);
    }
    return null;
  }

  function readMeta(){
    return {
      deselVer: scan("__desel_ver"),
      deselId:  scan("__desel_id"),
      selId:    scan("__sel_id"),
      selVer:   scan("__sel_ver")
    };
  }

  // ======================= DF push nudge =======================
  // On déclenche un "nudge" (évènement selectionChanged API) quand __df_push_ver change,
  // pour forcer st_aggrid à renvoyer response["data"] à Python.
  var lastDfVer = null;
  function dfPushVer(){ return scan("__df_push_ver"); }

  function nudgeReturn(){
    try{
      // Marque comme API pour ne pas perturber ta logique flip-flop
      updateMetaIfChanged("__sel_source","api");

      var selected = api.getSelectedNodes ? api.getSelectedNodes() : null;
      if (selected && selected.length){
        var n = selected[0];
        n.setSelected(false, true, true);
        n.setSelected(true,  true, true);
      } else {
        var first = api.getDisplayedRowAtIndex && api.getDisplayedRowAtIndex(0);
        if(first){
          first.setSelected(true,  true, true);
          first.setSelected(false, true, true);
        } else {
          api.dispatchEvent && api.dispatchEvent({type:"selectionChanged"});
        }
      }
    }catch(e){}
  }

  function nudgeIfDfPushed(){
    var v = dfPushVer();
    if (v!=null && v!==lastDfVer){
      lastDfVer = v;
      nudgeReturn();
    }
  }

  // ======================= recherche de node + scroll =======================
  function findNodeByUuid(id){
    if(id==null) return null;
    var node=api.getRowNode?api.getRowNode(String(id)):null;
    if(node) return node;
    var n=api.getDisplayedRowCount?api.getDisplayedRowCount():0;
    for(var i=0;i<n;i++){
      var r=api.getDisplayedRowAtIndex(i);
      if(r&&r.data&&String(r.data.__uuid)===String(id)) return r;
    }
    return null;
  }

  function ensureVisible(node){
    if(!node) return;
    if(typeof node.rowIndex==="number" && api.ensureIndexVisible){
      api.ensureIndexVisible(node.rowIndex,"middle");
    } else if(api.ensureNodeVisible){
      api.ensureNodeVisible(node,"middle");
    }
  }

  // ======================= scheduler léger =======================
  var schedPending=false;
  function sched(){ if(schedPending) return; schedPending=true; setTimeout(function(){schedPending=false; run();},30); }

  // ======================= coeur (dé)sélection : priorité à la désélection =======================
  function run(){
    nudgeIfDfPushed();  // vérifier les pushes DF à chaque passage
    var m=readMeta(); if(!m) return;

    // 1) déselection programmée
    if(m.deselVer!=null && deselCache[instId]!==m.deselVer){
      updateMetaIfChanged("__sel_source","api");
      api.deselectAll && api.deselectAll();
      ensureVisible(findNodeByUuid(m.deselId));
      deselCache[instId]=m.deselVer;
    }

    // 2) sélection programmée (once)
    if(m.selId!=null && m.selVer!=null && selCache[instId]!==m.selVer){
      updateMetaIfChanged("__sel_source","api");
      var node=findNodeByUuid(m.selId);
      if(node){
        api.deselectAll && api.deselectAll();
        node.setSelected && node.setSelected(true,true,true); // source=api
        ensureVisible(node);
      }
      selCache[instId]=m.selVer;
    }
  }

  // ======================= marquer la source user/api =======================
  function onRowSelected(ev){
    var src=(ev && ev.source) ? String(ev.source) : "api";
    var human=(src==="rowClicked" || src==="checkboxSelected" || src==="touch" || src==="selectAll") ? "user" : "api";
    updateMetaIfChanged("__sel_source", human);
  }

  // ======================= wiring AG Grid =======================
  if(p.type==="gridReady"){
    delete deselCache[instId];
    delete selCache[instId];

    api.addEventListener && api.addEventListener('rowSelected', onRowSelected);

    ["firstDataRendered","modelUpdated","rowDataUpdated"].forEach(function(e){
      api.addEventListener && api.addEventListener(e, function(){
        nudgeIfDfPushed();  // ← déclenche le nudge si push détecté
        sched();
      });
    });

    setTimeout(function(){
      // init
      updateMetaIfChanged("__sel_source","api"); // état neutre au boot
      nudgeIfDfPushed();                         // ← check tout de suite à l'init
      sched();
    }, 0);
  } else {
    nudgeIfDfPushed();
    sched();
  }
}
""")

# JS Code chargé de lancer la recherche Web sur la colonne Activité via l'icône loupe
JS_ACTIVITE_ICON_RENDERER = JsCode("""
class ActiviteRenderer {
  init(params){
    const e = document.createElement('div');
    e.style.display='flex'; e.style.alignItems='center'; e.style.gap='0.4rem';
    e.style.width='100%'; e.style.overflow='hidden';

    const label = (params.value ?? '').toString();
    const raw = params.data ? (params.data['Hyperlien'] ?? params.data['Hyperliens'] ?? '') : '';
    const href = String(raw || ("https://www.festivaloffavignon.com/resultats-recherche?recherche="+encodeURIComponent(label))).trim();

    const txt = document.createElement('span');
    txt.style.flex='1 1 auto'; txt.style.overflow='hidden'; txt.style.textOverflow='ellipsis';
    txt.textContent = label;
    // 🔸 pas de handler dblclick ici → AG Grid capte tout seul le double-clic
    e.appendChild(txt);

    const a = document.createElement('a');
    a.textContent = '🔎';
    a.href = href;
    a.target = '_blank';
    a.rel = 'noopener,noreferrer';
    a.title = 'Rechercher / Ouvrir le lien';
    a.style.flex='0 0 auto'; a.style.textDecoration='none'; a.style.userSelect='none';
    // on bloque juste la propagation pour ne pas déclencher sélection/édition
    a.addEventListener('click', ev=>ev.stopPropagation());
    e.appendChild(a);

    this.eGui = e;
  }
  getGui(){ return this.eGui; }
  refresh(){ return false; }
}
""")

# JS Code chargé de lancer la recherche d'itinéraire sur la colonne Lieu via l'icône épingle
JS_LIEU_ICON_RENDERER = JsCode("""
class LieuRenderer {
  init(params){
    const e = document.createElement('div');
    e.style.display='flex'; e.style.alignItems='center'; e.style.gap='0.4rem';
    e.style.width='100%'; e.style.overflow='hidden';

    const label = (params.value ?? '').toString().trim();

    // ---- adresse résolue (si dispo) ----
    const addrEnc = (params.data && params.data.__addr_enc)
      ? String(params.data.__addr_enc).trim()
      : encodeURIComponent(label || "");

    // ---- préférences + plateforme (depuis gridOptions.context) ----
    const ctx  = params.context || {};
    const app  = ctx.itineraire_app || "Google Maps";
    const plat = ctx.platform || (
      /iPad|iPhone|iPod/.test(navigator.userAgent) ? "iOS"
      : /Android/.test(navigator.userAgent) ? "Android" : "Desktop"
    );

    // ---- construire l'URL comme ton bouton ----
    let url = "#";
    if (addrEnc) {
      if (app === "Apple Maps" && plat === "iOS") {
        url = `http://maps.apple.com/?daddr=${addrEnc}`;
      } else if (app === "Google Maps") {
        if (plat === "iOS")       url = `comgooglemaps://?daddr=${addrEnc}`;
        else if (plat === "Android") url = `geo:0,0?q=${addrEnc}`;
        else                      url = `https://www.google.com/maps/dir/?api=1&destination=${addrEnc}`;
      } else {
        url = `https://www.google.com/maps/dir/?api=1&destination=${addrEnc}`;
      }
    }

    // ---- texte cellule (double-clic géré nativement par AG Grid) ----
    const txt = document.createElement('span');
    txt.style.flex='1 1 auto'; txt.style.overflow='hidden'; txt.style.textOverflow='ellipsis';
    txt.textContent = label;
    e.appendChild(txt);

    // ---- icône itinéraire (épingle) ----
    const a = document.createElement('a');
    a.textContent = '📍';
    a.href = url;
    a.target = (url === '#') ? '_self' : '_blank';
    a.rel = 'noopener,noreferrer';
    a.title = 'Itinéraire vers ce lieu';
    a.style.flex='0 0 auto'; a.style.textDecoration='none'; a.style.userSelect='none';
    if (url === '#') { a.style.opacity = 0.4; a.style.pointerEvents = 'none'; }
    a.addEventListener('click', ev=>ev.stopPropagation()); // ne pas sélectionner la ligne
    e.appendChild(a);

    this.eGui = e;
  }
  getGui(){ return this.eGui; }
  refresh(){ return false; }
}
""")

# JS Code chargé de lancer la recherche Web sur la colonne Activité via appui long
JS_ACTIVITE_LONGPRESS_RENDERER = JsCode("""
class ActiviteRenderer {
  init(params){
    // ---- conteneur + texte ----
    var e = document.createElement('div');
    e.style.display='flex'; e.style.alignItems='center'; e.style.gap='0.4rem';
    e.style.width='100%'; e.style.overflow='hidden';

    var label = (params.value != null ? String(params.value) : '');
    var raw   = (params.data && (params.data['Hyperlien'] || params.data['Hyperliens'])) ? (params.data['Hyperlien'] || params.data['Hyperliens']) : '';
    var href  = String(raw || ("https://www.festivaloffavignon.com/resultats-recherche?recherche="+encodeURIComponent(label))).trim();

    var txt = document.createElement('span');
    txt.style.flex='1 1 auto'; txt.style.overflow='hidden'; txt.style.textOverflow='ellipsis';
    txt.style.cursor='pointer';
    txt.textContent = label;
    e.appendChild(txt);

    // ---- helper: simuler un vrai clic cellule AG Grid (sélection propre) ----
    function tapSelectViaSyntheticClick(el){
      var cell = el.closest ? el.closest('.ag-cell') : null;
      if (!cell) return;
      try {
        cell.dispatchEvent(new MouseEvent('mousedown', {bubbles:true}));
        cell.dispatchEvent(new MouseEvent('mouseup',   {bubbles:true}));
        cell.dispatchEvent(new MouseEvent('click',     {bubbles:true}));
      } catch(_){}
    }

    // ---- fallback long-press (si window.attachLongPress absent) ----
    var attachLongPress = (typeof window !== 'undefined' && window.attachLongPress) || function(el, opts){
      var DELAY  = (opts && opts.delay)  != null ? opts.delay  : 550;
      var THRESH = (opts && opts.thresh) != null ? opts.thresh : 8;
      var TAP_MS = (opts && opts.tapMs)  != null ? opts.tapMs  : 220;
      var onUrl  = opts && opts.onUrl;
      var onTap  = opts && opts.onTap;
      var isIOS  = /iPad|iPhone|iPod/.test(navigator.userAgent);

      var sx=0, sy=0, moved=false, pressed=false, timer=null, startT=0, firedLong=false;
      var hadTouchTs = 0;

      function clearT(){ if (timer){ clearTimeout(timer); timer=null; } }
      function now(){ return Date.now(); }
      function withinTouchGrace(){ return (now() - hadTouchTs) < 800; }

      function openSameTab(u){
        if (!u) return;
        try { window.top.location.assign(u); } catch(e){ window.location.assign(u); }
      }
                                     
      function openNewTab(u){
        if (!u) return;
        try {
          var a=document.createElement('a');
          a.href=u; a.target='_blank'; a.rel='noopener,noreferrer';
          a.style.position='absolute'; a.style.left='-9999px'; a.style.top='-9999px';
          document.body.appendChild(a); a.click(); a.remove(); return;
        } catch(e){}
        try { var w=window.open(u,'_blank','noopener'); if (w) return; } catch(e){}
        try { window.location.assign(u); } catch(e){}
      }

      function onDown(ev){
        if (ev.type === 'mousedown' && withinTouchGrace()) return;
        var t = ev.touches ? ev.touches[0] : ev;
        sx = (t && t.clientX) || 0; sy = (t && t.clientY) || 0;
        moved=false; pressed=true; firedLong=false; startT=now();

        clearT();
        timer = setTimeout(function(){
          if (pressed && !moved){
            firedLong = true;
            var u = onUrl ? onUrl() : null;
            if (isIOS) openSameTab(u); else openNewTab(u);
            pressed=false;
          }
        }, DELAY);
      }

      function onMove(ev){
        if (!pressed) return;
        var t = ev.touches ? ev.touches[0] : ev;
        var dx = Math.abs(((t && t.clientX) || 0) - sx);
        var dy = Math.abs(((t && t.clientY) || 0) - sy);
        if (dx>THRESH || dy>THRESH){ moved=true; clearT(); }
      }

      function onUp(ev){
        if (ev.type === 'mouseup' && withinTouchGrace()) return;
        if (!pressed){ clearT(); return; }
        var dur = now() - startT;
        var isTap = (dur < TAP_MS) && !moved;
        pressed=false; clearT();

        if (isTap && !firedLong){
          if (typeof onTap === 'function'){
            // sélection via clic synthétique (pas de sélection "programmée")
            requestAnimationFrame(function(){ try { onTap(); } catch(_){ } });
          }
        }
      }

      function onCancel(){ pressed=false; clearT(); }

      if (window.PointerEvent){
        el.addEventListener('pointerdown', onDown, true);
        el.addEventListener('pointermove', onMove,  true);
        el.addEventListener('pointerup',   onUp,    true);
        el.addEventListener('pointercancel', onCancel, true);
      } else {
        el.addEventListener('touchstart', function(e){ hadTouchTs = now(); onDown(e); }, true);
        el.addEventListener('touchmove',  onMove, true);
        el.addEventListener('touchend',   onUp,   false);
        el.addEventListener('touchcancel', onCancel, true);
        el.addEventListener('mousedown', onDown, true);
        el.addEventListener('mousemove', onMove, true);
        el.addEventListener('mouseup',   onUp,   true);
      }

      el.addEventListener('contextmenu', function(e){ e.preventDefault(); }, true);
      el.style.webkitTouchCallout='none';
      el.style.webkitUserSelect='none';
      el.style.userSelect='none';
      el.style.touchAction='manipulation';
    };

    // ---- branchement unique ----
    attachLongPress(txt, {
      delay: 550,
      thresh: 8,
      tapMs: 220,
      onUrl: function(){ return href; },
      onTap: function(){ tapSelectViaSyntheticClick(txt); }
    });

    this.eGui = e;
  }
  getGui(){ return this.eGui; }
  refresh(){ return false; }
}
""")

# JS Code chargé de lancer la recherche d'itinéraire sur la colonne Lieu via appui long
JS_LIEU_LONGPRESS_RENDERER = JsCode("""
class LieuRenderer {
  init(params){
    // ---- conteneur + texte ----
    var e = document.createElement('div');
    e.style.display='flex'; e.style.alignItems='center'; e.style.gap='0.4rem';
    e.style.width='100%'; e.style.overflow='hidden';

    var label = (params.value != null ? String(params.value) : '').trim();

    var addrEnc = (params.data && params.data.__addr_enc != null)
      ? String(params.data.__addr_enc).trim()
      : encodeURIComponent(label || "");

    var ctx  = params.context || {};
    var app  = ctx.itineraire_app || "Google Maps";
    var plat = ctx.platform || (
      /iPad|iPhone|iPod/.test(navigator.userAgent) ? "iOS"
      : (/Android/.test(navigator.userAgent) ? "Android" : "Desktop")
    );

    var url = "#";
    if (addrEnc) {
      if (app === "Apple Maps" && plat === "iOS") {
        url = "http://maps.apple.com/?daddr=" + addrEnc;
      } else if (app === "Google Maps") {
        if (plat === "iOS")          url = "https://www.google.com/maps/dir/?api=1&destination=" + addrEnc;
        else if (plat === "Android") url = "geo:0,0?q=" + addrEnc;
        else                         url = "https://www.google.com/maps/dir/?api=1&destination=" + addrEnc;
      } else {
        url = "https://www.google.com/maps/dir/?api=1&destination=" + addrEnc;
      }
    }

    var txt = document.createElement('span');
    txt.style.flex='1 1 auto'; txt.style.overflow='hidden'; txt.style.textOverflow='ellipsis';
    txt.style.cursor='pointer';
    txt.textContent = label;
    e.appendChild(txt);

    // ---- helper: clic synthétique cellule ----
    function tapSelectViaSyntheticClick(el){
      var cell = el.closest ? el.closest('.ag-cell') : null;
      if (!cell) return;
      try {
        cell.dispatchEvent(new MouseEvent('mousedown', {bubbles:true}));
        cell.dispatchEvent(new MouseEvent('mouseup',   {bubbles:true}));
        cell.dispatchEvent(new MouseEvent('click',     {bubbles:true}));
      } catch(_){}
    }

    // ---- fallback long-press ----
    var attachLongPress = (typeof window !== 'undefined' && window.attachLongPress) || function(el, opts){
      var DELAY  = (opts && opts.delay)  != null ? opts.delay  : 550;
      var THRESH = (opts && opts.thresh) != null ? opts.thresh : 8;
      var TAP_MS = (opts && opts.tapMs)  != null ? opts.tapMs  : 220;
      var onUrl  = opts && opts.onUrl;
      var onTap  = opts && opts.onTap;
      var isIOS  = /iPad|iPhone|iPod/.test(navigator.userAgent);

      var sx=0, sy=0, moved=false, pressed=false, timer=null, startT=0, firedLong=false;
      var hadTouchTs = 0;

      function clearT(){ if (timer){ clearTimeout(timer); timer=null; } }
      function now(){ return Date.now(); }
      function withinTouchGrace(){ return (now() - hadTouchTs) < 800; }

      function openSameTab(u){
        if (!u) return;
        try { window.top.location.assign(u); } catch(e){ window.location.assign(u); }
      }
                                 
      function openNewTab(u){
        if (!u) return;
        try {
          var a=document.createElement('a');
          a.href=u; a.target='_blank'; a.rel='noopener,noreferrer';
          a.style.position='absolute'; a.style.left='-9999px'; a.style.top='-9999px';
          document.body.appendChild(a); a.click(); a.remove(); return;
        } catch(e){}
        try { var w=window.open(u,'_blank','noopener'); if (w) return; } catch(e){}
        try { window.location.assign(u); } catch(e){}
      }

      function onDown(ev){
        if (ev.type === 'mousedown' && withinTouchGrace()) return;
        var t = ev.touches ? ev.touches[0] : ev;
        sx = (t && t.clientX) || 0; sy = (t && t.clientY) || 0;
        moved=false; pressed=true; firedLong=false; startT=now();

        clearT();
        timer = setTimeout(function(){
          if (pressed && !moved){
            firedLong = true;
            var u = onUrl ? onUrl() : null;
            if (isIOS) openSameTab(u); else openNewTab(u);
            pressed=false;
          }
        }, DELAY);
      }

      function onMove(ev){
        if (!pressed) return;
        var t = ev.touches ? ev.touches[0] : ev;
        var dx = Math.abs(((t && t.clientX) || 0) - sx);
        var dy = Math.abs(((t && t.clientY) || 0) - sy);
        if (dx>THRESH || dy>THRESH){ moved=true; clearT(); }
      }

      function onUp(ev){
        if (ev.type === 'mouseup' && withinTouchGrace()) return;
        if (!pressed){ clearT(); return; }
        var dur = now() - startT;
        var isTap = (dur < TAP_MS) && !moved;
        pressed=false; clearT();

        if (isTap && !firedLong){
          if (typeof onTap === 'function'){
            requestAnimationFrame(function(){ try { onTap(); } catch(_){ } });
          }
        }
      }

      function onCancel(){ pressed=false; clearT(); }

      if (window.PointerEvent){
        el.addEventListener('pointerdown', onDown, true);
        el.addEventListener('pointermove', onMove,  true);
        el.addEventListener('pointerup',   onUp,    true);
        el.addEventListener('pointercancel', onCancel, true);
      } else {
        el.addEventListener('touchstart', function(e){ hadTouchTs = now(); onDown(e); }, true);
        el.addEventListener('touchmove',  onMove, true);
        el.addEventListener('touchend',   onUp,   false);
        el.addEventListener('touchcancel', onCancel, true);
        el.addEventListener('mousedown', onDown, true);
        el.addEventListener('mousemove', onMove, true);
        el.addEventListener('mouseup',   onUp,   true);
      }

      el.addEventListener('contextmenu', function(e){ e.preventDefault(); }, true);
      el.style.webkitTouchCallout='none';
      el.style.webkitUserSelect='none';
      el.style.userSelect='none';
      el.style.touchAction='manipulation';
    };

    // ---- branchement unique ----
    attachLongPress(txt, {
      delay: 550,
      thresh: 8,
      tapMs: 220,
      onUrl: function(){ return url; },
      onTap: function(){ tapSelectViaSyntheticClick(txt); }
    });

    this.eGui = e;
  }
  getGui(){ return this.eGui; }
  refresh(){ return false; }
}
""")
        # if (plat === "iOS")          url = "comgooglemaps://?daddr=" + addrEnc; # l'ouverture directe de l'appli depuis le cellRenderer ne marche pas sur IOS -> fallback sur GoogleMaps Web

# JS Code permettant de régler le probleme de blocage de l'UI au retour d'une page Web sur IOS en complément de inject_ios_soft_revive_global
JS_IOS_SOFT_REVIVE = JsCode("""
    function(params){
    try { params.api.sizeColumnsToFit(); } catch(e){}

    if (window.__iosSoftReviveInstalled) return;
    window.__iosSoftReviveInstalled = true;

    var isIOS = /iPad|iPhone|iPod/.test(navigator.userAgent);
    function cameFromBackForward(){
        try {
        var nav = performance.getEntriesByType && performance.getEntriesByType('navigation');
        return !!(nav && nav[0] && nav[0].type === 'back_forward');
        } catch(e){ return false; }
    }

    window.addEventListener('pageshow', function(e){
        if (!isIOS) return;
        if (e.persisted || cameFromBackForward()){
        // “soft revive” côté grille (pas de reload)
        try { params.api.deselectAll(); } catch(_) {}
        try { params.api.refreshCells({ force: true }); } catch(_) {}
        try { params.api.redrawRows(); } catch(_) {}
        try { window.dispatchEvent(new Event('resize')); } catch(_) {}

        // astuce : micro reflow de l’iframe
        try { 
            var root = document.documentElement;
            var prev = root.style.webkitTransform;
            root.style.webkitTransform = 'translateZ(0)';
            void root.offsetHeight;
            root.style.webkitTransform = prev || '';
        } catch(_) {}
        }
    }, false);
    }
    """)

####################
# Contexte Manager #
####################

def charger_contexte_depuis_sql():
    def to_timedelta(value, default):
        try:
            minutes = int(str(value).strip())
            return datetime.timedelta(minutes=minutes)    
        except (ValueError, TypeError, AttributeError):
            return default
    
    if "df" not in st.session_state:

        # Récupération df, meta, ca à partir de la base SQLite
        df, meta, ca = sql.charger_contexte()

        # Mise à jour wb, fn, fp
        try:
            wb = None
            fn  = meta["fn"]
            fp  = meta["fp"]
            if not (fp is None or str(fp).strip() == ""):
                wb = download_excel_from_dropbox(fp)
        except Exception as e:
            print(f"Erreur au chargement du modèle Excel depuis DropBox : {e}")

        # Mise à jour paramètres
        try:
            st.session_state.MARGE = to_timedelta(meta["MARGE"], default=MARGE)
            if meta["MARGE"] is None:
                sql.sauvegarder_param("MARGE")
            st.session_state.DUREE_REPAS = to_timedelta(meta["DUREE_REPAS"], default=DUREE_REPAS)
            if meta["DUREE_REPAS"] is None:
                sql.sauvegarder_param("DUREE_REPAS")
            st.session_state.DUREE_CAFE = to_timedelta(meta["DUREE_CAFE"], default=DUREE_CAFE)
            if meta["DUREE_CAFE"] is None:
                sql.sauvegarder_param("DUREE_CAFE")

            st.session_state.itineraire_app = meta["itineraire_app"]
            st.session_state.city_default = meta["city_default"]
            st.session_state.traiter_pauses = str(meta["traiter_pauses"]).strip().lower() == "true"
        except Exception as e:
            print(f"Erreur au chargement des paramètres depuis SQLite : {e}")

        # Mise à jour période de programmation
        try:
            val = meta["periode_a_programmer_debut"]
            if val is not None and str(val).strip() != "":
                st.session_state.periode_a_programmer_debut = datetime.date.fromisoformat(val.split(" ")[0])
            val = meta["periode_a_programmer_fin"]
            if val is not None and str(val).strip() != "":
                st.session_state.periode_a_programmer_fin = datetime.date.fromisoformat(val.split(" ")[0])
        except Exception as e:
            print(f"Erreur au chargement de la période de programmation depuis SQLite : {e}")
        
        # Si période programmation absente des meta rattrapage via init standard à partir des activités programmées du contexte que l'on vient de charger  
        if "periode_a_programmer_debut" not in st.session_state or "periode_a_programmer_fin" not in st.session_state:
            initialiser_periode_programmation(df) 
            sql.sauvegarder_param("periode_a_programmer_debut")
            sql.sauvegarder_param("periode_a_programmer_fin")
    
        st.session_state["push_periode_programmation_modele_values"] = True 

        df = nettoyer_donnees(df, fn)
        initialiser_etat_contexte(df, wb, fn, fp, ca)
        undo.init(verify=False)
        bd_maj_contexte(maj_donnees_calculees=True, maj_options_date=False) 

        st.session_state.activites_programmees = st.session_state.activites_programmees.drop(columns="__options_date", errors="ignore")
        st.session_state.activites_non_programmees = st.session_state.activites_non_programmees.drop(columns="__options_date", errors="ignore")
        
        selection = st.session_state.activites_non_programmees.index[0] if len(st.session_state.activites_non_programmees) > 0 else None
        demander_selection("activites_non_programmees", selection, deselect="activites_programmees")
        st.session_state.menu_activites = {
            "menu": "menu_activites_non_programmees",
            "index_df": selection
        }

        wk.enqueue_save_full(df, meta, ca)

def charger_contexte_depuis_gsheet():

    def to_timedelta(value, default):
        try:
            minutes = int(str(value).strip())
            return datetime.timedelta(minutes=minutes)    
        except (ValueError, TypeError, AttributeError):
            return default

    if "gsheets" in st.session_state:
        
        curseur_attente()

        try:

            # Récupération df, meta, ca à partir de la base SQLite
            df, meta, ca = gs.charger_contexte()
        
            # Mise à jour wb, fn, fp
            try:
                wb = None
                fn  = meta["fn"]
                fp  = meta["fp"]
                if not (fp is None or str(fp).strip() == ""):
                    wb = download_excel_from_dropbox(fp)
            except Exception as e:
                print(f"Erreur au chargement du modèle Excel depuis DropBox : {e}")

            # Mise à jour paramètres
            try:
                st.session_state.MARGE = to_timedelta(meta["MARGE"], default=MARGE)
                if meta["MARGE"] is None:
                    sql.sauvegarder_param("MARGE")
                st.session_state.DUREE_REPAS = to_timedelta(meta["DUREE_REPAS"], default=DUREE_REPAS)
                if meta["DUREE_REPAS"] is None:
                    sql.sauvegarder_param("DUREE_REPAS")
                st.session_state.DUREE_CAFE = to_timedelta(meta["DUREE_CAFE"], default=DUREE_CAFE)
                if meta["DUREE_CAFE"] is None:
                    sql.sauvegarder_param("DUREE_CAFE")

                st.session_state.itineraire_app = meta["itineraire_app"]
                st.session_state.city_default = meta["city_default"]
                st.session_state.traiter_pauses = str(meta["traiter_pauses"]).strip().lower() == "true"
            except Exception as e:
                print(f"Erreur au chargement des paramètres depuis SQLite : {e}")

            # Mise à jour période de programmation
            try:
                val = meta["periode_a_programmer_debut"]
                if val is not None and str(val).strip() != "":
                    st.session_state.periode_a_programmer_debut = datetime.date.fromisoformat(val.split(" ")[0])
                val = meta["periode_a_programmer_fin"]
                if val is not None and str(val).strip() != "":
                    st.session_state.periode_a_programmer_fin = datetime.date.fromisoformat(val.split(" ")[0])
            except Exception as e:
                print(f"Erreur au chargement de la période de programmation depuis SQLite : {e}")
    
            # Si période programmation absente des meta rattrapage via init standard à partir des activités programmées du contexte que l'on vient de charger  
            if "periode_a_programmer_debut" not in st.session_state or "periode_a_programmer_fin" not in st.session_state:
                initialiser_periode_programmation(df) 
                gs.sauvegarder_param("periode_a_programmer_debut")
                gs.sauvegarder_param("periode_a_programmer_fin")
    
            st.session_state["push_periode_programmation_modele_values"] = True 

            initialiser_dtypes(df)
            df = nettoyer_donnees(df, fn)
            df = add_persistent_uuid(df)
            df = add_hyperliens(df)
            initialiser_etat_contexte(df, wb, fn, fp, ca)
            undo.init(verify=False)
            bd_maj_contexte(maj_donnees_calculees=True, maj_options_date=False) 
            selection = st.session_state.activites_non_programmees.index[0] if len(st.session_state.activites_non_programmees) > 0 else None
            demander_selection("activites_non_programmees", selection, deselect="activites_programmees")
            st.session_state.menu_activites = {
                "menu": "menu_activites_non_programmees",
                "index_df": selection
            }
            curseur_normal()
        
        except Exception as e:
            print(f"Erreur au chargement des données depuis la Google Sheets : {e}")
            curseur_normal

###############
# API DropBox #
###############

import dropbox
from io import BytesIO

def get_dropbox_client() -> dropbox.Dropbox:
    """
    Retourne un client Dropbox qui renouvelle automatiquement
    l'access token à partir du refresh_token.
    """
    cfg = st.secrets["dropbox"]
    return dropbox.Dropbox(
        app_key=cfg["app_key"],
        app_secret=cfg["app_secret"],
        oauth2_refresh_token=cfg["refresh_token"]
    )

# Sauvegarde sur Dropbox le fichier Excel de l'utilisateur 
# Cette sauvegarde permet de garder une trace de la mise en page du fichier utilisateur
def upload_excel_to_dropbox(file_bytes, filename, dropbox_path="/uploads/"):
    dbx = get_dropbox_client()
    full_path = f"{dropbox_path}{filename}"

    try:
        dbx.files_upload(file_bytes, full_path, mode=dropbox.files.WriteMode("overwrite"))
        # st.success(f"✅ Fichier '{filename}' uploadé dans Dropbox à {full_path}")
        return full_path
    except Exception as e:
        print(f"upload_excel_to_dropbox : {e}")
        return ""

# Renvoie le fichier Excel de l'utilisateur sauvegardé sur DropBox
# Cette sauvegarde permet de garder une trace de la mise en page du fichier utilisateur
def download_excel_from_dropbox(file_path):
    dbx = get_dropbox_client()
    try:
        metadata, res = dbx.files_download(file_path)
        file_bytes = BytesIO(res.content)
        return load_workbook(file_bytes)
    except Exception as e:
        # st.error(f"❌ Erreur lors du téléchargement depuis Dropbox : {e}")
        return Workbook()

# Force le reaffichage d'un dataframe
def forcer_reaffichage_df(key):
    session_state_key_counter = key + "_key_counter"
    if session_state_key_counter in st.session_state:
        st.session_state[session_state_key_counter] += 1 
    session_state_forcer_reaffichage = key + "_forcer_reaffichage"
    if session_state_forcer_reaffichage in st.session_state:
        st.session_state[session_state_forcer_reaffichage] = True

# Affichage d'un dataframe
def afficher_df(label, df, hide=[], fixed_columns={}, header_names={}, key="affichage_df", colorisation=False, hide_label=False, background_color=None):

    # Calcul de la hauteur de l'aggrid
    nb_lignes = len(df)
    ligne_px = 30  # hauteur approximative d’une ligne dans AgGrid
    max_height = 250 #150
    height = min(nb_lignes * ligne_px + 50, max_height)

    # Initialisation du compteur qui permet de forcer le réaffichage de l'aggrid après une suppression de ligne 
    session_state_key_counter = key + "_key_counter"
    st.session_state.setdefault(session_state_key_counter, 0)
    
    # Initialisation du flag indiquant si l'on est en mode réaffichage complet de l'aggrid
    session_state_forcer_reaffichage = key + "_forcer_reaffichage"
    st.session_state.setdefault(session_state_forcer_reaffichage, )
       
    # Initialisation de la variable d'état contenant la requête de selection / déselection
    session_state_sel_request = key + "_sel_request"
    st.session_state.setdefault(session_state_sel_request, copy.deepcopy(SEL_REQUEST_DEFAUT))

    gb = GridOptionsBuilder.from_dataframe(df)

    # Configuration par défaut des colonnes
    gb.configure_default_column(resizable=True)

    # Colonnes à largeur fixe
    for col, width in fixed_columns.items():
        if col in df.columns:
            gb.configure_column(
                col,
                filter=False,
                resizable=False,
                width=width,
                minWidth=width,
                maxWidth=width,
                flex=0,
                suppressSizeToFit=True,
            )

    # header names
    for col, name in header_names.items():
        if col in df.columns:
            gb.configure_column(
                col,
                headerName=name
            )

    # Epinglage de la colonne Date
    if "Date" in df.columns:
        gb.configure_column(
            "Date",
            pinned=JsCode("'left'")
        )

    #Colonnes cachées
    for col in hide:
        if col in df.columns:
            gb.configure_column(col, hide=True)

    # Colorisation
    if colorisation:
        if "Date" in df.columns:
            df["__jour"] = df["Date"].apply(lambda x: int(str(int(float(x)))[-2:]) if pd.notna(x) else None)
            gb.configure_column("__jour", hide=True)
            gb.configure_grid_options(getRowStyle=JsCode(f"""
            function(params) {{
                const jour = params.data.__jour;
                const couleurs = {PALETTE_COULEURS_JOURS};
                if (jour && couleurs[jour]) {{
                    return {{ 'backgroundColor': couleurs[jour] }};
                }}
                return null;
            }}
            """))
    elif background_color is not None:
        gb.configure_grid_options(getRowStyle=JsCode(f"""
            function(params) {{
                return {{
                    'backgroundColor': '{background_color}'
                }}
            }}
            """)
        )

    # Configuration de la sélection
    gb.configure_selection(selection_mode="single", use_checkbox=False) #, pre_selected_rows=[current_selected_row_pos]) 

    # Gestion des sélections / désélections demandées via demander_selection() demander_deselection()
    # Utilise le JS code JS_SELECT_DESELECT_ONCE lequel exploite les colonnes de travail __sel_id, __sel_ver, __desel_id, __desel_ver
    # __sel_id = id de la ligne à sélectionner (None si pas de contrainte de sélection)
    # __sel_ver = version de la demande de sélection (doit être incrémentée à chaque demande)
    # __desel_id = id de la ligne devant rester visible lors de la déselection (None si aucune contrainte de visibilité lors de la désélection)
    # __desel_ver = version de la demande de désélection (doit être incrémentée à chaque demande)
    sel_request_key = key + "_sel_request"
    sel_request = st.session_state.get(sel_request_key)
    gb.configure_column("__desel_ver", hide=True)
    if "__desel_ver" not in df.columns:
        df["__desel_ver"] = sel_request["desel"]["ver"] if sel_request is not None else 0
    gb.configure_column("__desel_id", hide=True)
    if "__desel_id" not in df.columns:
        df["__desel_id"] =  get_uuid(df, sel_request["desel"]["id"]) if sel_request is not None else None
    gb.configure_column("__sel_ver", hide=True)
    if "__sel_ver" not in df.columns:
        df["__sel_ver"] = sel_request["sel"]["ver"] if sel_request is not None else 0
    gb.configure_column("__sel_id", hide=True)
    if "__sel_id" not in df.columns:
        df["__sel_id"] =  get_uuid(df, sel_request["sel"]["id"]) if sel_request is not None else None
    gb.configure_column("__sel_source", hide=True)
    if "__sel_source" not in df.columns:
        df["__sel_source"] = "api"
    
    row = None
    selection_demandee = False
    if sel_request is not None and sel_request["sel"]["pending"]:
        if sel_request["sel"]["id"] is not None:
            reqid = sel_request["sel"]["id"]
            # tracer.log(f"{key}: Traitement de la requête de sélection id {sel_request["sel"]["id"]} ver {sel_request["sel"]["ver"]}")
            df["__sel_id"] = get_uuid(df, reqid)
            df["__sel_ver"] = sel_request["sel"]["ver"]
            if reqid in df.index: 
                row = df.loc[reqid]
                # tracer.log(f"{key}: row = df.loc[{reqid}]")
            selection_demandee = True
        st.session_state[sel_request_key]["sel"]["pending"] = False

    deselection_demandee = False
    if sel_request is not None and sel_request["desel"]["pending"]:
        # tracer.log(f"{key}: Traitement de la requête de desélection ver {sel_request["desel"]["ver"]}")
        df["__desel_ver"] = sel_request["desel"]["ver"]
        df["__desel_id"] = get_uuid(df, sel_request["desel"]["id"]) # id visible après déselection, None si pas de contrainte de visibilité
        df["__sel_id"] = None
        deselection_demandee = True
        st.session_state[sel_request_key]["desel"]["pending"] = False

    gb.configure_grid_options(
        onGridReady=JS_SELECT_DESELECT_ONCE,
    )
    
    # Mise en page de la grille
    gb.configure_grid_options(onFirstDataRendered=JsCode(f"""
        function(params) {{
            params.api.sizeColumnsToFit();
        }}
    """))

    # Permet de gérer les modifications de df_display dans avoir à redessiner l'aggrid complètement par changement de key
    gb.configure_grid_options(
        immutableData=True,
        deltaRowDataMode=True,
        getRowId=JsCode("function (params) { return params.data.__uuid; }"),
    )

    grid_options = gb.build()
    grid_options["suppressMovableColumns"] = True

    if not hide_label:
        st.markdown(f"##### {label}")

    response = AgGrid(
        df,
        gridOptions=grid_options,
        allow_unsafe_jscode=True,
        height=height,
        reload_data=True,
        data_return_mode=DataReturnMode.AS_INPUT,
        update_mode=GridUpdateMode.MODEL_CHANGED | GridUpdateMode.SELECTION_CHANGED,
        key=f"_{key}",
    )

    event_data = response.get("event_data")
    event_type = event_data["type"] if isinstance(event_data, dict) else None
    tracer.log(f"{key}: event {event_type}", types=["gen", "event"])

    # Récupération du retour grille __sel_source
    # Cette information est passée à la valeur "user" par le JsCode JS_SELECT_DESELECT_ONCE si le cellValueChanged provient d'un click utilisateur.
    # Elle permet de n'effectuer les traitements de cellValueChanged que sur les seuls évènements utilisateurs et de bypasser ceux provenant d'une
    # demande de sélection programmée via demander_selection().
    df_dom = pd.DataFrame(response["data"])
    if not df_dom.empty:
        first_row = df_dom.iloc[0]
        sel_source = (first_row.get("__sel_source") or "api") # 'user' ou 'api'
        tracer.log(f"{key}: sel_source {sel_source}", types=["sel_source"])

    selected_row_key = key + "_selected_row"
    selected_row_pred = st.session_state.get(selected_row_key, df.iloc[0] if len(df) > 0 else None)

    selected_rows = response["selected_rows"]
    if not selection_demandee:
        if isinstance(selected_rows, pd.DataFrame) and not selected_rows.empty:
            # tracer.log("{key}: row = selected_rows.iloc[0]")
            row = selected_rows.iloc[0] 
        elif isinstance(selected_rows, list) and len(selected_rows) > 0:
            # tracer.log("{key}: row = selected_rows[0]")
            row = selected_rows[0]
        else:
            # tracer.log("{key}: row = selected_row_pred")
            row = selected_row_pred

    st.session_state[sel_request_key]["sel"]["id"] = get_index_from_uuid(df, row["__uuid"]) if row is not None else None
    st.session_state[selected_row_key] = row
    return row

##########################
# Fonctions applicatives #
##########################

# Renvoie un descripteur d'activité à partir d'une date et d'une ligne du df
def get_descripteur_activite(date, row):
    titre = f"{date} - [{row['Debut'].strip()} - {row['Fin'].strip()}] - {row['Activite']}"
    if not (pd.isna(row["Lieu"]) or str(row["Lieu"]).strip() == ""):
        titre = titre + f"( {row['Lieu']}) - P{formatter_cellule_int(row['Priorite'])}"
    return titre

# Affiche le titre de la page de l'application
def afficher_titre(title):
    # Réduire l’espace en haut de la page
    st.markdown(
        """
        <style>
            .block-container {
                padding-top: 2rem;
            }
        </style>
        """, unsafe_allow_html=True
    )

    # Titre de la page
    st.markdown(f"## {title}")

# Affiche l'aide de l'application
def afficher_aide():
    with st.expander("À propos"):
    
        with st.expander("Fonctionnalités générales"):
            st.markdown("""
            <div style='font-size: 14px;'>
            <p style="margin-bottom: 0.2em">Cette application offre les fonctionnalités suivantes:</p>
            <ul style="margin-top: 0em; margin-bottom: 2em">
            <li>Choix de la période à programmer</li>
            <li>Chargement des activités à programmer à partir d'un fichier Excel</li>
            <li>Gestion de la programmation des activités en respectant les règles décrites dans le paragraphe ci-dessous</li>
            <li>Gestion des créneaux disponibles</li>
            <li>Prise en compte optionnelle des pauses (déjeuner, dîner, café)</li>
            <li>Gestion des liens de recherche sur le net</li>
            <li>Sauvegarde des données modifiées dans un fichier téléchargeable</li>
            <li>Fonction défaire / refaire</li>
            <li>Vérification de cohérence des données d'entrée (chevauchements d'activités, marges trop courtes, formats de données)</li>
            </ul>            
            </div>
            """, unsafe_allow_html=True)  

        with st.expander("Règles de programmation des activités"):
            st.markdown("""
            <div style='font-size: 14px;'>
            <p style="margin-bottom: 0.2em">Règles adoptées pour la programmation des activités:</p>
            <ul style="margin-top: 0em; margin-bottom: 0.5em">
            <li>30 minutes de marge entre activités</li>
            <li>1 heure par pause repas</li>
            <li>1/2 heure par pause café sans marge avec l'activité précédente ou suivante</li>
            <li>Respect des jours de relâches</li>
            </ul>
            <p>Ces valeurs sont paramétrables via la rubrique Paramètres.</p>
            </div>
            """, unsafe_allow_html=True)  

        with st.expander("Utilisation"):
            st.markdown("""
            <div style='font-size: 14px;'>
            <p>Les activités à programmer sont présentées dans deux tableaux séparés, 
                l'un pour les activités déja programmées à une date donnée, l'autre pour les activités restant à programmer. 
                Dans ces deux tableaux les informations sont éditables, sauf les heures de fin (qui sont calculées automatiquement) 
                et les dates de programmation, heures de début et durées des activités réservées (celles dont la colonne 'Réservé' est à Oui). 
                Sur la colonne Date un menu permet de programmer / reprogrammer les activités en fonction du jour sélectionné, 
                voire de déprogrammer les activités du tableau des activités programmées par sélection de l'item vide du menu. 
                Dans le tableau des activités programmées la couleur de fond est fonction du jour de programmation 
                et les activités réservées sont écrite en rouge. Dans le tableau des activités non programmées la couleur de fond menthe 
                permet de repérer les activités programmables.</p>
            
            <p>Deux autres tableaux adressent la gestion des créneaux disponibles. 
                Le premier présente les créneaux encore disponibles sur la période considérée et le deuxième les activités programmables dans 
                le créneau sélectionné en tenant compte de leur durée et de la marge entre activités. 
                Un bouton Programmer permet de programmer l'activité programmable sélectionnée au jour dit du créneau sélectionné. 
                la couleur de fond est fonction du jour pour les créneaux disponibles et les activités programmables.</p>
            
            <p style="margin-bottom: 0.2em">Les menus sont regroupés dans une barre latérale escamotable:</p>
            <ul style="margin-top: 0em">
                <li>Menu Fichier: permet de charger un contexte à partir d'un fichier, initialiser un nouveau contexte, sauvegarder le contexte courant dans un fichier téléchargeable.</li>
                <li>Menu Edition: permet de défaire, refaire une opération.</li>
                <li>Menu Activités: permet sur l'activité séléctionnée dans les tableaux d'activites programmées et non programmées (vous pouvez passer de l'activité sélectionnée dans l'un ou l'autre des tableaux en cliquant sur le champ affichant l'activité courante) de:
                        <ul>
                        <li>rechercher de l'information sur le Web (via un lien Web éditable dans les propriétés),</li> 
                        <li>rechercher un itinaire, sur la base du lieu enregistré pour l'activité (l'application d'itinéraire et la ville de recherche par défaut sont réglables dans la section Paramètres et un carnet d'adresses avec colonnes Nom et Adresse peut être enregistré dans la feuille 2 du fichier Excel d'entrée),</li>
                        <li>supprimer l'activité (si elle n'est pas réservée),</li> 
                        <li>déprogrammer l'activité (si elle est déjà programmée sans être réservée),</li>
                        <li>programmer / reprogrammer l'activité (si elle n'est pas réservée et que d'autres dates de programmation sont possibles)</li>
                        <li>éditer les propriétés l'activité.</li>
                        </ul>
                </li>
            </ul>
                        
            <p style="margin-bottom: 0.2em">En haut de la page principale une rubrique escamotable 'Infos' présente:</p>
            <ul style="margin-top: 0em">
                <li>La présente aide.</li>
                <li>Une rubrique présentant les incohérences dans le fichier chargé (notamment les chevauchements de programmation en tenant compte des marges entre activités). 
                    Cette rubrique est mise à jour au fil de l'eau.</li>
                <li>La période programmation: elle est automatiquement déduite des activités renseignées dans le fichier chargé, mais peut être modifiée en cours d'édition. Par défaut l'application recherche les dates de début et de fin du festival de l'année courante.</li>
                <li>Les paramètres de l'application comprennant:
                        <ul>
                        <li>la marge entre activités</li>
                        <li>la durée des pauses repas et café</li>
                        <li>le nom de l'application d'itinéraire (Google Maps, Apple, etc.)</li>
                        <li>la ville de recherche par défaut pour la recherche d'itinéraire</li>
                        <li>la possibilité de choisir si les menus de gestion des activités sont dans la barre latérale ou la page principale.</li>
                        </ul>
                </li>
            </ul>
                        
            <p>A la première utilisation l'application propose à l'utilisateur de créer un espace personnel dans lequel est automatiquement sauvegardé le contexte de travail (l'adresse de cet espace est : adresse de l'application/?user_id=id utilisateur).
                En cas de rupture de connexion avec le serveur, le travail en cours est ainsi automatiquement restitué à la prochaine reconnexion.</p>
            </div>
            """, unsafe_allow_html=True)  

        with st.expander("Format des données"):
            st.markdown("""
            <div style='font-size: 14px;'>
            <p style="margin-bottom: 0.2em">Le fichier Excel d'entrée doit contenir en feuille 1 les colonnes suivantes:</p>
            <ul style="margin-top: 0em; margin-bottom: 2em">
            <li>Date : Date de l'activité (entier)</li>
            <li>Début : Heure de début de l'activité (format HHhMM)</li>
            <li>Fin : Heure de fin de l'activité (format HHhMM)</li>
            <li>Durée : Durée de l'activité (format HHhMM ou HHh)</li>
            <li>Activité : Nom de l'activité (nom de spectacle, pause, visite, ...)</li>
            <li>Lieu : Lieu de l'activité</li>
            <li>Relâche : Jours de relâche pour l'activité (liste d'entiers, peut être vide)</li>
            <li>Réservé : Indique si l'activité est réservée (Oui/Non, vide interpété comme Non)</li>
            </ul>

            <p>En feuille 2 peut être fourni un carnet d'adresses des lieux d'activités utilisé pour la recherche d'itinéraire. 
            Il doit comprendre au moins une colonne Nom et une colonne Adresse.</p>

            <p>📥Un modèle Excel est disponible <a href="https://github.com/jnicoloso-91/PlanifAvignon-05/raw/main/Mod%C3%A8le%20Excel.xlsx" download>
            ici
            </a></p>
            <p>ℹ️ Si le téléchargement ne démarre pas, faites un clic droit → "Enregistrer le lien sous...".</p>
            </div>
            """, unsafe_allow_html=True)  

def get_dates_festival():
    
    # 1️⃣ Tentative de récupération des dates du festival depuis le site officiel (recherche simple)
    def fetch_off_festival_dates():
        url = "https://www.festivaloffavignon.com/"
        r = requests.get(url, timeout=5)
        soup = BeautifulSoup(r.text, "html.parser")
        # Recherche dans le texte "du 5 au 26 juillet 2025"
        text = soup.get_text()
        match = re.search(r"du\s+(\d{1,2})\s+juillet\s+au\s+(\d{1,2})\s+juillet\s+2025", text, re.IGNORECASE)
        if match:
            d1, d2 = map(int, match.groups())
            base_year = 2025
            base_month = 7
            return datetime.date(base_year, base_month, d1), datetime.date(base_year, base_month, d2)
        return None, None

    if "festival_debut" not in st.session_state or "festival_fin" not in st.session_state:
        debut, fin = fetch_off_festival_dates()
        if debut and fin:
            st.session_state.festival_debut = debut
            st.session_state.festival_fin = fin
        else:
            # Valeurs de secours (manuelles)
            st.session_state.festival_debut = datetime.date(2025, 7, 5)
            st.session_state.festival_fin = datetime.date(2025, 7, 26)
    return {
        "debut": st.session_state.festival_debut,
        "fin": st.session_state.festival_fin
    }

# Retourne les valeurs non nulles et convertibles de la colonne Date d'un df
def get_dates_from_df(df):
    return df["Date"].dropna().apply(lambda x: int(float(x)) if str(x).strip() != "" else None).dropna().astype(int)
     
# Affichage de la période à programmer
def initialiser_periode_programmation(df):

    if "nouveau_fichier" not in st.session_state:
        st.session_state.nouveau_fichier = True
    
    # Initialisation de la periode si nouveau fichier
    if st.session_state.nouveau_fichier == True:
        # Reset du flag déclenché par uploader
        st.session_state.nouveau_fichier = False

        # Initialisation des variables de début et de fin de période à programmer
        periode_a_programmer_debut = None 
        periode_a_programmer_fin = None

        dates_valides = get_dates_from_df(df)

        if not dates_valides.empty:
            # Conversion en datetime
            base_date = datetime.date(datetime.date.today().year, 7, 1)
            dates_datetime = dates_valides.apply(lambda j: datetime.datetime.combine(base_date, datetime.datetime.min.time()) + datetime.timedelta(days=j - 1))
            if not dates_datetime.empty:
                periode_a_programmer_debut = dates_datetime.min()
                periode_a_programmer_fin = dates_datetime.max()

        if periode_a_programmer_debut is None or periode_a_programmer_fin is None:
            dates_festival = get_dates_festival()
            periode_a_programmer_debut = dates_festival["debut"]
            periode_a_programmer_fin = dates_festival["fin"]
        
        st.session_state.periode_a_programmer_debut = periode_a_programmer_debut
        st.session_state.periode_a_programmer_fin = periode_a_programmer_fin
    
    if "periode_a_programmer_debut" not in st.session_state or "periode_a_programmer_fin" not in st.session_state:
        dates_festival = get_dates_festival()
        st.session_state.periode_a_programmer_debut = dates_festival["debut"]
        st.session_state.periode_a_programmer_fin = dates_festival["fin"]
    
def afficher_periode_programmation():
    with st.expander("Période de programmation", expanded=False):

        changed_keys = []
        need_maj_contexte = False

        if st.session_state.get("periode_programmation_abandon_pending", False) == True:
            st.session_state.periode_debut_input = st.session_state.periode_a_programmer_debut
            st.session_state.periode_fin_input = st.session_state.periode_a_programmer_fin
            st.session_state.periode_programmation_abandon_pending = False

        with st.form("periode_programmation_form"):
            base_deb = st.session_state.periode_a_programmer_debut
            base_fin = st.session_state.periode_a_programmer_fin

            deb_kwargs = dict(key="periode_debut_input", format="DD/MM/YYYY")
            fin_kwargs = dict(key="periode_fin_input",   format="DD/MM/YYYY")

            # init une seule fois
            if "periode_debut_input" not in st.session_state:
                st.session_state.periode_debut_input = base_deb
            if "periode_fin_input" not in st.session_state:
                st.session_state.periode_fin_input = base_fin

            # Prise en compte des valeurs du modèle si l'app les a recalculées par ailleurs
            push_modele_values = st.session_state.get("push_periode_programmation_modele_values", True)
            if push_modele_values and "periode_a_programmer_debut" in st.session_state and "periode_a_programmer_fin" in st.session_state:
                st.session_state.periode_debut_input = st.session_state.periode_a_programmer_debut
                st.session_state.periode_fin_input = st.session_state.periode_a_programmer_fin
                st.session_state["push_periode_programmation_modele_values"] = False

            # Surtout: ne PAS mettre deb_kwargs["value"] / fin_kwargs["value"]
            # -> st.date_input lira directement st.session_state[<key>]

            dates_valides = get_dates_from_df(st.session_state.df)  # doit renvoyer une série d'int (jours)
            date_min = int(dates_valides.min()) if not dates_valides.empty else None
            date_max = int(dates_valides.max()) if not dates_valides.empty else None

            if isinstance(date_min, int):
                try:
                    if date_min is not None:
                        deb_kwargs["max_value"] = base_deb.replace(day=date_min)
                except ValueError as e:
                    print(e)
            if isinstance(date_max, int):
                try:
                    if date_max is not None:
                        fin_kwargs["min_value"] = base_fin.replace(day=date_max)
                except ValueError as e:
                    print(e)

            try:
                col1, col2 = st.columns(2)
                with col1:
                    debut = st.date_input("Début", **deb_kwargs)
                with col2:
                    fin   = st.date_input("Fin", **fin_kwargs)

            except Exception as e:
                print(f"Erreur dans afficher_periode_programmation : {e}")
        

            col1, col2 = st.columns(2)
            appliquer = col1.form_submit_button("Appliquer", use_container_width=True)
            abandonner = col2.form_submit_button("Abandonner", use_container_width=True)

        if appliquer:
            if debut != st.session_state.periode_a_programmer_debut:
                st.session_state.periode_a_programmer_debut = debut
                changed_keys.append("periode_a_programmer_debut")
                need_maj_contexte = True

            if fin != st.session_state.periode_a_programmer_fin:
                st.session_state.periode_a_programmer_fin = fin
                changed_keys.append("periode_a_programmer_fin")
                need_maj_contexte = True
            
            # Ne forcer le réaffichage des grilles qu'une seule fois
            if need_maj_contexte:
                bd_maj_contexte(maj_donnees_calculees=False)
                forcer_reaffichage_df("creneaux_disponibles")

            # Sauvegarde en batch (une seule fois)
            if changed_keys:
                for k in changed_keys:
                    try:
                        sql.sauvegarder_param(k)  
                    except Exception  as e:
                        print(f"Erreur dans afficher_periode_programmation : {e}")

                # Pas de st.rerun() nécessaire : submit a déjà provoqué un rerun
                st.toast("Paramètres appliqués.", icon="✅")

        if abandonner:
            st.session_state.periode_programmation_abandon_pending = True
            st.rerun()

def afficher_parametres():

    def ajouter_sans_doublon(liste, val):
        if val not in liste:
            liste.append(val)

    def get_itin_options(platform):
        if platform == "iOS":
            itin_options = ["Apple Maps", "Google Maps"]
        elif platform == "Android":
            itin_options = ["Google Maps"]
        else:
            itin_options = ["Google Maps"]
        return itin_options

    with st.expander("Paramètres", expanded=False):

        # Recupération de la plateforme
        platform = get_platform()  # "iOS" | "Android" | "Desktop"/None

        changed_keys = []
        need_maj_contexte = False

        if st.session_state.get("param_abandon_pending", False) == True:
            st.session_state.param_marge_min = minutes(st.session_state.MARGE)
            st.session_state.param_repas_min = minutes(st.session_state.DUREE_REPAS)
            st.session_state.param_cafe_min  = minutes(st.session_state.DUREE_CAFE)
            st.session_state.itineraire_app_selectbox = st.session_state.itineraire_app
            st.session_state.city_default_input = st.session_state.city_default
            st.session_state.param_abandon_pending = False

        with st.form("params_form"):

            # Marge entre activités
            if st.session_state.get("MARGE") is None:
                st.session_state.MARGE = MARGE
                sql.sauvegarder_param("MARGE")  

            st.session_state.setdefault("param_marge_min", minutes(st.session_state.MARGE))
            st.slider(
                "Marge entre activités (minutes)",
                min_value=0, max_value=120, step=5,
                value=st.session_state.param_marge_min,
                key="param_marge_min",
                help="Marge utilisée pour le calcul des créneaux disponibles. Pour les pauses café, ne s’applique qu’à l’activité précédente OU suivante, la pause café étant supposée se tenir près du lieu de l'une ou de l'autre."
            )

            # Durée des pauses repas
            if st.session_state.get("DUREE_REPAS") is None:
                st.session_state.DUREE_REPAS = DUREE_REPAS
                sql.sauvegarder_param("DUREE_REPAS")  

            st.session_state.setdefault("param_repas_min", minutes(st.session_state.DUREE_REPAS))
            st.slider(
                "Durée des pauses repas (minutes)",
                min_value=0, max_value=120, step=5,
                value=st.session_state.param_repas_min,
                key="param_repas_min",
                help="Durée utilisée pour les pauses repas."
            )

            # Durée des pauses café
            if st.session_state.get("DUREE_CAFE") is None:
                st.session_state.DUREE_CAFE = DUREE_CAFE
                sql.sauvegarder_param("DUREE_CAFE")  

            st.session_state.setdefault("param_cafe_min",  minutes(st.session_state.DUREE_CAFE))
            st.slider(
                "Durée des pauses café (minutes)",
                min_value=0, max_value=120, step=5,
                value=st.session_state.param_cafe_min,
                key="param_cafe_min",
                help="Durée utilisée pour les pauses café."
            )

            # Application itinéraire
            itin_options = get_itin_options(platform)
            if st.session_state.get("itineraire_app") is None:
                st.session_state.itineraire_app = itin_options[0]
                sql.sauvegarder_param("itineraire_app")  
                        
            index = itin_options.index(st.session_state.itineraire_app) if "itineraire_app_selectbox" not in st.session_state else itin_options.index(st.session_state.itineraire_app_selectbox)
            st.selectbox(
                "Application itinéraire",
                options=itin_options,
                index=index, 
                key="itineraire_app_selectbox",
                help="Sur IOS : Apple/Google Maps. Sinon : Google Maps."
            )

            # Ville par défaut pour la recherche d'itinéraire
            if st.session_state.get("city_default") is None:
                st.session_state.city_default = "Avignon"
                sql.sauvegarder_param("city_default")  

            st.session_state.setdefault("city_default_input", st.session_state.city_default)
            st.text_input(
                "Ville par défaut pour la recherche d'itinéraire",
                value=st.session_state.city_default_input,
                key="city_default_input",
                help="Si vide, la ville du lieu de l’activité est utilisée pour la recherche d'itinéraire."
            )

            col1, col2 = st.columns(2)
            appliquer = col1.form_submit_button("Appliquer", use_container_width=True)
            abandonner = col2.form_submit_button("Abandonner", use_container_width=True)

        if appliquer:

            # MARGE
            new_marge = datetime.timedelta(minutes=st.session_state.param_marge_min)
            if st.session_state.MARGE != new_marge:
                st.session_state.MARGE = new_marge
                ajouter_sans_doublon(changed_keys, "MARGE")
                need_maj_contexte = True

            # DUREE_REPAS
            new_repas = datetime.timedelta(minutes=st.session_state.param_repas_min)
            if st.session_state.DUREE_REPAS != new_repas:
                st.session_state.DUREE_REPAS = new_repas
                ajouter_sans_doublon(changed_keys, "DUREE_REPAS")
                need_maj_contexte = True

            # DUREE_CAFE
            new_cafe = datetime.timedelta(minutes=st.session_state.param_cafe_min)
            if st.session_state.DUREE_CAFE != new_cafe:
                st.session_state.DUREE_CAFE = new_cafe
                ajouter_sans_doublon(changed_keys, "DUREE_CAFE")
                need_maj_contexte = True

            # Itinéraire
            new_itineraire = st.session_state.itineraire_app_selectbox
            if st.session_state.itineraire_app != new_itineraire:
                st.session_state.itineraire_app = new_itineraire
                ajouter_sans_doublon(changed_keys, "itineraire_app")

            # Ville par défaut
            new_city = st.session_state.city_default_input.strip()
            if st.session_state.city_default != new_city:
                st.session_state.city_default = new_city
                ajouter_sans_doublon(changed_keys, "city_default")

            # Mise à jour de contexte (seulement si nécessaire car opération lourde)
            if need_maj_contexte:
                bd_maj_contexte(maj_donnees_calculees=False)

            # Sauvegarde des paramètres modifiés
            if changed_keys:
                for k in changed_keys:
                    try:
                        sql.sauvegarder_param(k)  
                    except Exception  as e:
                        print(f"Erreur dans afficher_parametres : {e}")
            
            st.toast("Paramètres appliqués.", icon="✅")

        if abandonner:
            st.session_state.param_abandon_pending = True
            st.rerun()

# Nettoie les données du tableau Excel importé
def nettoyer_donnees(df, fn):
    try:
        # Nettoyage noms de colonnes : suppression espaces et accents
        df.columns = df.columns.str.strip().str.replace("\u202f", " ").str.normalize("NFKD").str.encode("ascii", errors="ignore").str.decode("utf-8")

        if not all(col in df.columns for col in COLONNES_ATTENDUES):
            st.session_state.contexte_invalide_message = f"Le fichier {fn} n'est pas au format Excel ou ne contient pas toutes les colonnes attendues: " + ", ".join(COLONNES_ATTENDUES_ACCENTUEES) + "."
        else:
            initialiser_dtypes(df)

            if (len(df) > 0):
                # Suppression des lignes presque vides i.e. ne contenant que des NaN ou des ""
                df = df[~df.apply(lambda row: all(pd.isna(x) or str(x).strip() == "" for x in row), axis=1)].reset_index(drop=True)

                # Nettoyage Heure (transforme les datetime, time et None en str mais ne garantit pas le format HHhMM, voir est_heure_valide pour cela)
                df["Debut"] = df["Debut"].apply(heure_str).astype("string")

                # Nettoyage Duree (transforme les timedelta et None en str mais ne garantit pas le format HhMM, voir est_duree_valide pour cela)
                df["Duree"] = df["Duree"].apply(duree_str).astype("string")

                # Colonne Relache castée en object avec NaN remplacés par "" et le reste en str
                df["Relache"] = df["Relache"].astype("object").fillna("").astype(str)

            # Valide le contexte si pas d'exception dans le traitement précédent
            if "contexte_invalide" in st.session_state:
                del st.session_state["contexte_invalide"]

        return df
            
    except Exception as e:
        st.error(f"Erreur lors du décodage du fichier : {e}")
        df = pd.DataFrame(columns=COLONNES_ATTENDUES)
        initialiser_dtypes(df)


# Renvoie les hyperliens de la colonne Activité 
def get_liens_activites(wb):
    liens_activites = {}
    try:
        ws = wb.worksheets[0]
        for cell in ws[1]:
            if cell.value and str(cell.value).strip().lower() in ["activité"]:
                col_activite_index = cell.column
        for row in ws.iter_rows(min_row=2, min_col=col_activite_index, max_col=col_activite_index):
            cell = row[0]
            if cell.hyperlink:
                liens_activites[cell.value] = cell.hyperlink.target
            else:
                # Construire l'URL de recherche par défaut
                if cell.value is not None:
                    url = f"https://www.festivaloffavignon.com/resultats-recherche?recherche={cell.value.replace(' ', '+')}"
                    liens_activites[cell.value] = url  # L'enregistrer dans la session
        return liens_activites
    except:
        return liens_activites

# Vérifie la cohérence des informations du dataframe et affiche le résultat dans un expander
def verifier_coherence(df):
    
    # @st.cache_data
    def get_log_verifier_coherence(df):
        # try:
        erreurs = []

        def est_entier(x):
            try:
                return not pd.isna(x) and str(x).strip() != "" and int(float(x)) == float(x)
            except Exception:
                return False
        
        if len(df) <= 0:
            return
        
        # 1. 🔁 Doublons
        df_valid = df[df["Activite"].notna() & (df["Activite"].astype(str).str.strip() != "")]
        df_valid = df_valid.copy()  # pour éviter SettingWithCopyWarning
        df_valid["_activite_clean"] = df_valid["Activite"].astype(str).str.strip().str.lower()
        doublons = df_valid[df_valid.duplicated(subset=["_activite_clean"], keep=False)]

        if not doublons.empty:
            bloc = []
            for _, row in doublons.iterrows():
                if not est_pause(row):
                    try:
                        date_str = str(int(float(row["Date"]))) if pd.notna(row["Date"]) else "Vide"
                    except (ValueError, TypeError):
                        date_str = "Vide"
                    heure_str = str(row["Debut"]).strip() if pd.notna(row["Debut"]) else "Vide"
                    duree_str = str(row["Duree"]).strip() if pd.notna(row["Duree"]) else "Vide"
                    
                    if not bloc:
                        bloc = ["🟠 Doublons d'activités :"]

                    bloc.append(f"{date_str} - {heure_str} - {row['Activite']} ({duree_str})")
            erreurs.append("\n".join(bloc))
            
        # 2. ⛔ Chevauchements
        chevauchements = []
        df_sorted = df.sort_values(by=["Date", "Debut_dt"])
        for i in range(1, len(df_sorted)):
            r1 = df_sorted.iloc[i - 1]
            r2 = df_sorted.iloc[i]
            if r1.isna().all() or r2.isna().all():
                continue
            if pd.notna(r1["Date"]) and pd.notna(r2["Date"]) and r1["Date"] == r2["Date"]:
                fin1 = r1["Debut_dt"] + r1["Duree_dt"]
                debut2 = r2["Debut_dt"]
                if debut2 < fin1:
                    chevauchements.append((r1, r2))
        if chevauchements:
            bloc = ["🔴 Chevauchements:"]
            for r1, r2 in chevauchements:
                bloc.append(
                    f"{r1['Activite']} ({r1['Debut']} / {r1['Duree']}) chevauche {r2['Activite']} ({r2['Debut']} / {r2['Duree']}) le {r1['Date']}"
                )
            erreurs.append("\n".join(bloc))

        # 3. 🕒 Erreurs de format
        bloc_format = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            # Date : uniquement si non NaN
            if pd.notna(row["Date"]) and not est_entier(row["Date"]):
                bloc_format.append(f"Date invalide à la ligne {idx + 2} : {row['Date']}")

            # Ne tester Heure/Duree que si Activite ou Autres est renseigné
            if str(row["Activite"]).strip() != "":
                if not re.match(r"^\d{1,2}h\d{2}$", str(row["Debut"]).strip()):
                    bloc_format.append(f"Heure invalide à la ligne {idx + 2} : {row['Debut']}")
                if not re.match(r"^\d{1,2}h\d{2}$", str(row["Duree"]).strip()):
                    bloc_format.append(f"Durée invalide à la ligne {idx + 2} : {row['Duree']}")
            
            # Test de la colonne Relache
            if not est_relache_valide(row["Relache"]):
                bloc_format.append(f"Relache invalide à la ligne {idx + 2} : {row['Relache']}")

        # 4. 📆 Spectacles un jour de relâche (Date == Relache)
        bloc_relache = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            if (
                est_entier(row["Date"]) and
                est_entier(row["Relache"]) and
                int(float(row["Date"])) == int(float(row["Relache"])) and
                str(row["Activite"]).strip() != ""
            ):
                bloc_relache.append(
                    f"{row['Activite']} prévu le jour de relâche ({int(row['Date'])}) à la ligne {idx + 2}"
                )
        if bloc_relache:
            erreurs.append("🛑 Spectacles programmés un jour de relâche:\n" + "\n".join(bloc_relache))

        # 5. 🕳️ Heures non renseignées
        bloc_heure_vide = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            if str(row["Activite"]).strip() != "":
                if pd.isna(row["Debut"]) or str(row["Debut"]).strip() == "":
                    bloc_heure_vide.append(f"Heure vide à la ligne {idx + 2}")
        if bloc_heure_vide:
            erreurs.append("⚠️ Heures non renseignées:\n" + "\n".join(bloc_heure_vide))

        # 6. 🕓 Heures au format invalide
        bloc_heure_invalide = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            if str(row["Activite"]).strip() != "":
                h = row["Debut"]
                if pd.notna(h) and str(h).strip() != "":
                    h_str = str(h).strip().lower()
                    is_time_like = isinstance(h, (datetime.datetime, datetime.time))
                    valid_format = bool(re.match(r"^\d{1,2}h\d{2}$", h_str) or re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", h_str))
                    if not is_time_like and not valid_format:
                        bloc_heure_invalide.append(f"Heure invalide à la ligne {idx + 2} : {h}")
        if bloc_heure_invalide:
            erreurs.append("⛔ Heures mal formatées:\n" + "\n".join(bloc_heure_invalide))

        # 7. 🕳️ Durées non renseignées ou nulles
        bloc_duree_nulle = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            if isinstance(row["Duree_dt"], pd.Timedelta) and row["Duree_dt"] == pd.Timedelta(0):
                if pd.isna(row["Duree"]) or str(row["Duree"]).strip() == "":
                    msg = f"Durée vide à la ligne {idx + 2}"
                else:
                    msg = f"Durée égale à 0 à la ligne {idx + 2} : {row['Duree']}"
                bloc_duree_nulle.append(msg)
        if bloc_duree_nulle:
            erreurs.append("⚠️ Durées nulles ou vides:\n" + "\n".join(bloc_duree_nulle))

        # 8. ⏱️ Durées au format invalide
        bloc_duree_invalide = []
        for idx, row in df.iterrows():
            # ignorer si rien n'est programmé
            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in ["Activite", "Debut", "Duree"]):
                continue
            if row.isna().all():
                continue

            if str(row["Activite"]).strip() != "":
                d = row["Duree"]
                if pd.notna(d) and str(d).strip() != "":
                    d_str = str(d).strip().lower()
                    is_timedelta = isinstance(d, pd.Timedelta)
                    valid_format = bool(re.match(r"^\d{1,2}h\d{2}$", d_str) or re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", d_str))
                    if not is_timedelta and not valid_format:
                        bloc_duree_invalide.append(f"Durée invalide à la ligne {idx + 2} : {d}")
        if bloc_duree_invalide:
            erreurs.append("⛔ Durées mal formatées:\n" + "\n".join(bloc_duree_invalide))

        contenu = "<div style='font-size: 14px;'>"
        for bloc in erreurs:
            lignes = bloc.split("\n")
            if lignes[0].startswith(("🟠", "🔴", "⚠️", "🛑", "⛔")):
                contenu += f"<p><strong>{lignes[0]}</strong></p><ul>"
                for ligne in lignes[1:]:
                    contenu += f"<li>{ligne}</li>"
                contenu += "</ul>"
            else:
                contenu += f"<p>{bloc}</p>"
        contenu += "</div>"
        return contenu
        
    with st.expander("Cohérence des données"):
        st.markdown(get_log_verifier_coherence(df), unsafe_allow_html=True)

# Indique si une row est une activité programmée
def est_activite_programmee(row):
    if isinstance(row, pd.DataFrame):
        row=row.iloc[0] # sinon and plante car pd.isna et pd.notna renvoient des series
    return (est_float_valide(row["Date"]) and 
             pd.notna(row["Debut"]) and 
             pd.notna(row["Duree"]) and 
             pd.notna(row["Activite"]))


# Renvoie le dataframe des activités programmées
def get_activites_programmees(df):
    return df[
        df["Date"].apply(est_float_valide) & 
        df["Debut"].notna() & 
        df["Duree"].notna() &
        df["Activite"].notna()
    ].sort_values(by=["Date", "Debut_dt"], ascending=[True, True])

# Indique si une row est une activité non programmée
def est_activite_non_programmee(row):
    if isinstance(row, pd.DataFrame):
        row=row.iloc[0] # sinon and plante car pd.isna et pd.notna renvoient des series
    return (pd.isna(row["Date"]) and 
             pd.notna(row["Debut"]) and 
             pd.notna(row["Duree"]) and 
             pd.notna(row["Activite"]))

# Renvoie le dataframe des activités non programmées
def get_activites_non_programmees(df):
    return df[df["Date"].isna() & 
              df["Debut"].notna() & 
              df["Duree"].notna() &
              df["Activite"].notna()
    ].sort_values(by=["Date", "Debut_dt"], ascending=[True, True])

# Affiche le bouton de recharche sur le web
def afficher_bouton_web(nom_activite, disabled=False):    

    #Retour si nom activité vide
    if pd.isna(nom_activite):
        return
                
    # Initialiser le dictionnaire si nécessaire
    if "liens_activites" not in st.session_state:
        st.session_state["liens_activites"] = {}

    liens = st.session_state["liens_activites"]

    # Vérifier si un lien existe déjà
    if nom_activite in liens:
        url = liens[nom_activite]
    else:
        # Construire l'URL de recherche
        url = f"https://www.festivaloffavignon.com/resultats-recherche?recherche={nom_activite.replace(' ', '+')}"
        if nom_activite in liens:
            liens[nom_activite] = url  # L'enregistrer dans la session

    st.link_button(LABEL_BOUTON_CHERCHER_WEB, url, use_container_width=CENTRER_BOUTONS, disabled=disabled)

# Détection basique de plateforme
def get_platform():
    if "platform" in st.session_state:
        return st.session_state["platform"]

    user_agent = st_javascript("navigator.userAgent", key="user_agent_detect")
    if user_agent == 0 or user_agent is None:
        # tracer.log("Détection plateforme")
        st.stop()

    # Traitement une fois la valeur reçue
    ua = user_agent.lower()
    if "iphone" in ua or "ipad" in ua or "ipod" in ua:
        platform = "iOS"
    elif "android" in ua:
        platform = "Android"
    elif "windows" in ua:
        platform = "Windows"
    elif "macintosh" in ua:
        platform = "macOS"
    elif "linux" in ua:
        platform = "Linux"
    else:
        platform = "Autre"

    # tracer.log("Plateforme détectée")

    st.session_state["platform"] = platform
    st.rerun()   

from difflib import SequenceMatcher

@st.cache_data(show_spinner=False)
def prepare_carnet(carnet_df: pd.DataFrame) -> pd.DataFrame:
    """Ajoute une colonne normalisée (une seule fois, puis cache)."""
    df = carnet_df.copy()
    if "Nom" in df.columns:
        df["_Nom_norm"] = df["Nom"].astype(str).map(normalize_text)
    else:
        df["_Nom_norm"] = ""
    return df

def ensure_addr_cols(df):
    if "__addr_enc"   not in df.columns: df["__addr_enc"]   = None
    carnet = st.session_state.get("ca")
    city_default = st.session_state.get("city_default", "")
    mask = df["Lieu"].notna()
    for i in df.index[mask]:
        if pd.isna(df.at[i, "__addr_enc"]) or not str(df.at[i, "__addr_enc"]).strip():
            addr_h, addr_enc = resolve_address_fast(df.at[i, "Lieu"], carnet, city_default=city_default)
            df.at[i, "__addr_enc"] = addr_enc
    return df

def set_addr_cols(df, idx, lieu):
    carnet = st.session_state.get("ca")
    city_default = st.session_state.get("city_default", "")
    matches = df[df["__index"].astype(str) == str(idx)]
    if not matches.empty:
        addr_h, addr_enc = resolve_address_fast(lieu, carnet, city_default=city_default)
        df.at[matches.index[0], "__addr_enc"] = addr_enc


def resolve_address_fast(lieu: str, carnet_df: pd.DataFrame | None, city_default="Avignon"):
    """
    1) Cherche dans le carnet par égalité puis 'contains' (normalisé, sans accents).
    2) Si rien -> renvoie 'lieu, <city>'.
    Retourne (addr_humaine, addr_enc).
    """
    lieu = lieu if isinstance(lieu, str) else ""
    lieu = lieu.strip()
    key = normalize_text(lieu)

    addr = ""
    if carnet_df is not None and {"Nom","Adresse"}.issubset(carnet_df.columns):
        df = prepare_carnet(carnet_df)

        # match exact (rapide)
        hit = df.loc[df["_Nom_norm"].eq(key)]
        if hit.empty and key:
            # contains (vectorisé)
            hit = df.loc[df["_Nom_norm"].str.contains(re.escape(key), na=False)]

        if not hit.empty:
            val = hit.iloc[0]["Adresse"]
            if pd.notna(val):
                addr = str(val).strip()

    if not addr:
        # fallback toujours: lieu + ville
        addr = f"{lieu}, {city_default}" if lieu else city_default

    return addr, quote_plus(addr)

def resolve_address(lieu: str, carnet_df: pd.DataFrame | None = None, default_city="Avignon"):
    """
    Retourne (addr_humaine, addr_enc) en essayant d'abord le carnet (Nom -> Adresse)
    avec recherche accent-insensible, partielle, et fuzzy.
    Si pas trouvé, ajoute toujours ", <city>" au lieu.
    """
    def _best_match_row(carnet_df: pd.DataFrame, key_norm: str):
        """
        Retourne l'index de la meilleure ligne matchée dans carnet_df
        selon l'ordre: égalité stricte > contains > fuzzy.
        Renvoie None si aucun candidat crédible.
        """
        if carnet_df.empty:
            return None

        # Prépare colonne normalisée
        if "_Nom_norm" not in carnet_df.columns:
            carnet_df["_Nom_norm"] = carnet_df["Nom"].astype(str).apply(normalize_text)

        noms = carnet_df["_Nom_norm"]

        # 1) égalité stricte
        exact = carnet_df.index[noms == key_norm]
        if len(exact):
            return exact[0]

        # 2) contains (key dans nom)
        contains_idx = [i for i, n in noms.items() if key_norm in n]
        if contains_idx:
            # si plusieurs, prend le plus proche via ratio fuzzy
            if len(contains_idx) == 1:
                return contains_idx[0]
            best = max(contains_idx, key=lambda i: SequenceMatcher(None, key_norm, noms[i]).ratio())
            return best

        # 3) fuzzy global (utile si fautes de frappe)
        # on prend les candidats avec ratio >= 0.75 et choisit le meilleur
        scored = [(i, SequenceMatcher(None, key_norm, n).ratio()) for i, n in noms.items()]
        scored = [x for x in scored if x[1] >= 0.75]
        if scored:
            scored.sort(key=lambda x: x[1], reverse=True)
            return scored[0][0]

        return None

    lieu = lieu if isinstance(lieu, str) else ""
    saisie = lieu.strip()
    key = normalize_text(saisie)

    addr = ""

    if carnet_df is not None and {"Nom", "Adresse"}.issubset(carnet_df.columns):
        try:
            row_idx = _best_match_row(carnet_df, key)
            if row_idx is not None:
                val = carnet_df.loc[row_idx, "Adresse"]
                if pd.notna(val):
                    addr = str(val).strip()
        except Exception:
            pass  # pas de blocage si carnet mal formé

    # Fallback : toujours ajouter la ville si rien trouvé
    if not addr:
        if saisie:
            addr = f"{saisie}, {default_city}"
        else:
            addr = default_city

    addr_enc = quote_plus(addr) if addr else ""
    return addr, addr_enc

# Affiche le bouton de recherche d'itinéraire
def afficher_bouton_itineraire(lieu, disabled=False):  

    # Bouton désactivé si lieu vide ou None
    if pd.isna(lieu) or not str(lieu).strip():
        st.link_button(
            LABEL_BOUTON_CHERCHER_ITINERAIRE,
            "#",  # pas de lien cliquable
            use_container_width=CENTRER_BOUTONS,
            disabled=True
        )
        return
    
     # Résolution depuis carnet + fallback
    addr_human, addr_enc = resolve_address_fast(lieu, st.session_state.ca, city_default=st.session_state.city_default)
    itineraire_app = st.session_state.get("itineraire_app", "Google Maps")
    platform = get_platform()  

    if itineraire_app == "Apple Maps" and platform == "iOS":
        url = f"http://maps.apple.com/?daddr={addr_enc}"

    elif itineraire_app == "Google Maps":
        if platform == "iOS":
            url = f"comgooglemaps://?daddr={addr_enc}"
        elif platform == "Android":
            url = f"geo:0,0?q={addr_enc}"
        else:
            # Sur desktop, on retombe sur la version web
            url = f"https://www.google.com/maps/dir/?api=1&destination={addr_enc}"

    else:  # Google Maps
        url = f"https://www.google.com/maps/dir/?api=1&destination={addr_enc}"

    st.link_button(
        LABEL_BOUTON_CHERCHER_ITINERAIRE,
        url,
        use_container_width=CENTRER_BOUTONS,
        disabled=disabled or not addr_enc
    )

# Indique si une activité donnée par son descripteur dans le df est réservée
def est_activite_reserve(ligne_df):
    return str(ligne_df["Reserve"]).strip().lower() == "oui"

# Renvoie les lignes modifées entre df1 et df2, l'index de df2 est supposé se trouver dans la colonne __index de df1
def get_lignes_modifiees(df1, df2, columns_to_drop=[]):
    lignes_modifiees = set()
    for i, row in df1.iterrows():
        idx = row["__index"]
        for col in df1.drop(columns=columns_to_drop).columns:
            if idx in df2.index:
                val_avant = df2.at[idx, col]
                val_apres = row[col]
                if pd.isna(val_avant) and pd.isna(val_apres):
                    continue
                if (pd.isna(val_avant) and pd.notna(val_apres)) or val_avant != val_apres:
                    lignes_modifiees.add((i, idx))
    return lignes_modifiees

# Renvoie la première ligne modifée entre df1 et df2, l'index de df2 est supposé se trouver dans la colonne __index de df1
def get_ligne_modifiee(df1, df2, columns_to_drop=[]):
    for i, row in df1.iterrows():
        idx = row["__index"]
        for col in df1.drop(columns=columns_to_drop).columns:
            if idx in df2.index:
                val_avant = df2.at[idx, col]
                val_apres = row[col]
                if pd.isna(val_avant) and pd.isna(val_apres):
                    continue
                if (pd.isna(val_avant) and pd.notna(val_apres)) or val_avant != val_apres:
                    return i, idx
    return None, None

# DialogBox de suppression d'activité
@st.dialog("Suppression activité")
def show_dialog_supprimer_activite(df, index_df, df_display):
    st.markdown("Voulez-vous supprimer cette activité ?")
    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button(LABEL_BOUTON_VALIDER, use_container_width=CENTRER_BOUTONS):
            undo.save()
            if est_activite_programmee(df.loc[index_df]):
                demander_selection("activites_programmees", ligne_voisine_index(df_display, index_df), deselect="activites_non_programmees")
            else:
                demander_selection("activites_non_programmees", ligne_voisine_index(df_display, index_df), deselect="activites_programmees")
            forcer_reaffichage_df("creneaux_disponibles")
            supprimer_activite(index_df)
            sql.sauvegarder_row(index_df)
            st.rerun()
    with col2:
        if st.button(LABEL_BOUTON_ANNULER, use_container_width=CENTRER_BOUTONS):
            st.rerun()

# DialogBox de reprogrammation d'activité programmée
@st.dialog("Reprogrammation activité")
def show_dialog_reprogrammer_activite_programmee(df, activites_programmees, index_df, df_display, jours_possibles):
    jour_escape = "Aucune" # escape pour déprogrammer l'activité
    jours_possibles = get_jours_possibles(df, activites_programmees, index_df) + [jour_escape]
    jours_label = [f"{int(jour):02d}" for jour in jours_possibles[:-1]] + [jours_possibles[-1]]
    jour_selection = st.selectbox("Choisissez une nouvelle date pour cette activité :", jours_label, key = "ChoixJourReprogrammationActiviteProgrammee")
    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button(LABEL_BOUTON_VALIDER, use_container_width=CENTRER_BOUTONS):
            if jour_selection == jour_escape:
                # Déprogrammation
                undo.save()
                demander_selection("activites_non_programmees", index_df, deselect="activites_programmees")
                bd_deprogrammer_activite_programmee(index_df)
                forcer_reaffichage_df("creneaux_disponibles")
                sql.sauvegarder_row(index_df)
                st.rerun()
            else:
                # Reprogrammation 
                jour_choisi = int(jour_selection) 
                undo.save()
                demander_selection("activites_programmees", index_df, deselect="activites_non_programmees")
                df.at[index_df, "Date"] = jour_choisi
                sql.sauvegarder_row(index_df)
                st.rerun()
    with col2:
        if st.button(LABEL_BOUTON_ANNULER, use_container_width=CENTRER_BOUTONS):
            st.rerun()

# DialogBox de programmation d'activité non programmée
@st.dialog("Programmation activité")
def show_dialog_programmer_activite_non_programmee(df, index_df, df_display, jours_possibles):
    jours_label = [f"{int(jour):02d}" for jour in jours_possibles]
    jour_selection = st.selectbox("Choisissez une date pour cette activité :", jours_label, key = "ChoixJourProgrammationActiviteNonProgrammee")
    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button(LABEL_BOUTON_VALIDER, use_container_width=CENTRER_BOUTONS):
            # Programmation à la date choisie
            jour_choisi = int(jour_selection.split()[-1])
            undo.save()
            demander_selection("activites_programmees", index_df, deselect="activites_non_programmees")
            df.at[index_df, "Date"] = jour_choisi
            forcer_reaffichage_df("creneaux_disponibles")
            sql.sauvegarder_row(index_df)
            st.rerun()
    with col2:
        if st.button(LABEL_BOUTON_ANNULER, use_container_width=CENTRER_BOUTONS):
            st.rerun()

# Demande de sélection d'une ligne sur une grille
def demander_selection(grid_name: str, target_id: str | None, deselect=None, visible_id: str | None=None):
    if grid_name is not None:
        tracer.log(f"{grid_name} {target_id}")
        k = f"{grid_name}_sel_request"
        st.session_state.setdefault(k, copy.deepcopy(SEL_REQUEST_DEFAUT))
        st.session_state[k]["sel"]["ver"] += 1
        st.session_state[k]["sel"]["id"] = target_id
        st.session_state[k]["sel"]["pending"] = True
        demander_deselection(deselect, visible_id=visible_id)

# Demande de désélection de la ligne sélectionnée sur une grille
def demander_deselection(grid_name: str, visible_id: str | None=None):
    if grid_name is not None:
        tracer.log(f"{grid_name}")
        k = f"{grid_name}_sel_request"
        st.session_state.setdefault(k, copy.deepcopy(SEL_REQUEST_DEFAUT))
        st.session_state[k]["desel"]["ver"] += 1
        st.session_state[k]["desel"]["id"] = visible_id
        st.session_state[k]["desel"]["pending"] = True
        st.session_state[k]["sel"]["id"] = None

# Initialisation des grid_options sur la grille des activités programmées
def init_activites_programmees_grid_options(df_display):

    gb = GridOptionsBuilder.from_dataframe(df_display)

    # Configuration par défaut des colonnes
    gb.configure_default_column(resizable=True) 

    # Colonnes à largeur fixe
    colonnes_fixes = {"Date": 55, "Début": 55, "Fin": 55, "Durée": 55}
    for col, width in colonnes_fixes.items():
        gb.configure_column(
            col,
            filter=False,
            resizable=False,
            width=width,
            minWidth=width,
            maxWidth=width,
            flex=0,
            suppressSizeToFit=True,
        )

    # Epinglage de la colonne Date
    gb.configure_column(
        "Date",
        pinned=JsCode("'left'")
    )

    # Masquage des colonnes de travail
    work_cols = ACTIVITES_PROGRAMMEES_WORK_COLS
    for c in work_cols:
        gb.configure_column(c, hide=True)

    # Colonnes editables
    non_editable_cols = ["Fin"] + work_cols
    for col in df_display.columns:
        gb.configure_column(col, editable=(col not in non_editable_cols))

    gb.configure_column(
        "Début",
        editable=JsCode("function(params) { return params.data.__non_reserve; }")
    )

    gb.configure_column(
        "Durée",
        editable=JsCode("function(params) { return params.data.__non_reserve; }")
    )

    # Configuration des menus de la colonne 
    gb.configure_column(
        "Date",
        editable=True,
        cellEditor="agSelectCellEditor",
        cellEditorParams=JsCode("""
            function(params) {
                let raw = params.data.__options_date;
                let values = [];

                try {
                    values = JSON.parse(raw);
                } catch (e) {
                    values = [];
                }

                return { values: values };
            }
        """)
    )

    # Configuration des icônes de colonnes pour la recherche Web et la recherche d'itinéraire
    # gb.configure_column("Activité", editable=True, cellRenderer=JS_ACTIVITE_ICON_RENDERER) #, minWidth=220)
    # gb.configure_column("Lieu", editable=True, cellRenderer=JS_LIEU_ICON_RENDERER) #, minWidth=200)

    # Configuration de l'appui long pour la recherche Web et la recherche d'itinéraire
    gb.configure_column("Activité", editable=True, cellRenderer=JS_ACTIVITE_LONGPRESS_RENDERER) #, minWidth=220)
    gb.configure_column("Lieu",     editable=True, cellRenderer=JS_LIEU_LONGPRESS_RENDERER) #, minWidth=200)

    # Colorisation
    gb.configure_grid_options(getRowStyle=JsCode(f"""
        function(params) {{
            const jour = params.data.__jour;
            const couleurs = {PALETTE_COULEURS_JOURS};
            let style = {{}};

            if (jour && couleurs[jour]) {{
                style.backgroundColor = couleurs[jour];
            }}

            if (params.data.__non_reserve === false) {{
                style.color = 'red';
            }}

            return style;
        }}
    """))

    # Configuration de la sélection
    gb.configure_selection(selection_mode="single", use_checkbox=False) 
    
    gb.configure_grid_options(
        getRowNodeId=JsCode("function(data) { return String(data.__uuid); }"),
        getRowId=JsCode("function(p){ return String(p.data.__uuid); }"),
        columnTypes={"textColumn": {}},  # évite l'erreur #36
        onGridReady=JS_SELECT_DESELECT_ONCE,
        onFirstDataRendered=JS_IOS_SOFT_REVIVE,
    )

    # Mise en page de la grille (repris dans JS_IOS_SOFT_REVIVE)
    # gb.configure_grid_options(onFirstDataRendered=JsCode(f"""
    #     function(params) {{
    #         params.api.sizeColumnsToFit();
    #     }}
    # """))

    grid_options = gb.build()

    # Empêche la possibilité de réorganiser les colonnes
    grid_options["suppressMovableColumns"] = True

    # Supprime le highlight de survol qui pose problème sur mobile et tablette
    grid_options["suppressRowHoverHighlight"] = True

    # Enregistre dans le contexte les paramètres nécessaires à la recherche d'itinéraire (voir JS_LIEU_ICON_RENDERER)
    grid_options["context"] = {
        "itineraire_app": st.session_state.get("itineraire_app", "Google Maps"),
        "platform": get_platform(),  # "iOS" / "Android" / "Desktop"
    }

    return grid_options

# Affiche les activités programmées dans un tableauflag allow_unsafe_jscode is on. AgGrid.tsx:124:15
def afficher_activites_programmees():

    df = st.session_state.get("df")
    if df is None :
        return

    df_display = st.session_state.get("activites_programmees_df_display")
    if df_display is None :
        return

    work_cols = ACTIVITES_PROGRAMMEES_WORK_COLS
    non_editable_cols = ["Fin"] + work_cols

    # Calcul de la hauteur de l'aggrid
    nb_lignes = len(df_display)
    ligne_px = 30  # hauteur approximative d’une ligne dans AgGrid
    max_height = 250
    height = min(nb_lignes * ligne_px + 50, max_height)

    # Initialisation du compteur qui permet de savoir si l'on doit forcer le réaffichage de l'aggrid après une suppression de ligne 
    st.session_state.setdefault("activites_programmees_key_counter", 0)

    # Initialisation de la variable d'état indiquant s'il convient de bypasser la section d'édition de cellule 
    st.session_state.setdefault("activites_programmees_bypass_cell_edit", False)

    # Initialisation de la variable d'état contenant la requête de selection / déselection
    st.session_state.setdefault("activites_programmees_sel_request", copy.deepcopy(SEL_REQUEST_DEFAUT))
   
    # Gestion des sélections / désélections demandées via demander_selection() demander_deselection()
    # Utilise le JS code JS_SELECT_DESELECT_ONCE lequel exploite les colonnes de travail suivantes:
    # __sel_id = id de la ligne à sélectionner (None si pas de contrainte de sélection)
    # __sel_ver = version de la demande de sélection (doit être incrémentée à chaque demande)
    # __desel_id = id de la ligne devant rester visible lors de la déselection (None si aucune contrainte de visibilité lors de la désélection)
    # __desel_ver = version de la demande de désélection (doit être incrémentée à chaque demande)
    # __sel_source = information renvoyée par le DOM (event.source exposé par onRowSelected) indiquant si la source de selection est "user" ou "api" selon que la demande de sélection provient d'un click utilisateur ou d'une requête python via JsCode
    # Ces colonnes sont configurées par les fonctions utilisateur demander_selection() et demander_deselection()
    # L'information de retour __sel_source est exploitée par le mecanisme de flip flop entre grille A et grille B
    # via le champ "data" de la réponse de l'aggrid (cf. fonction afficher_activites_programmees() et afficher_activites_non_programmees())
    row = None
    selection_demandee = False
    sel_request = st.session_state.get("activites_programmees_sel_request")
    if sel_request["sel"]["pending"]:
        if sel_request["sel"]["id"] is not None:
            reqid = sel_request["sel"]["id"]
            # tracer.log(f"Traitement de la requête de sélection id {sel_request["sel"]["id"]} ver {sel_request["sel"]["ver"]}")
            df_display["__sel_id"] = get_uuid(df_display, reqid)
            df_display["__sel_ver"] = sel_request["sel"]["ver"]
            if reqid in df_display.index: 
                row = df_display.loc[reqid]
                # tracer.log(f"row = df_display.loc[{reqid}]")
            selection_demandee = True
        st.session_state.activites_programmees_sel_request["sel"]["pending"] = False

    deselection_demandee = False
    if sel_request["desel"]["pending"]:
        # tracer.log(f"Traitement de la requête de desélection ver {sel_request["desel"]["ver"]}")
        df_display["__desel_ver"] = sel_request["desel"]["ver"]
        df_display["__desel_id"] = get_uuid(df_display, sel_request["desel"]["id"]) # id visible après déselection, None si pas de contrainte de visibilité
        df_display["__sel_id"] = None
        deselection_demandee = True
        st.session_state.activites_programmees_sel_request["desel"]["pending"] = False
        
    # if len(df_display) > 0:
    #     tracer.log(f"df_display['__sel_id'] {df_display.iloc[0]["__sel_id"]} df_display['__sel_ver'] {df_display.iloc[0]["__sel_ver"]} df_display['__desel_ver'] {df_display.iloc[0]["__desel_ver"]}")

    grid_options = init_activites_programmees_grid_options(df_display)

    # Affichage
    with st.expander("**Activités programmées**", expanded=True):
        response = AgGrid(
            df_display,
            gridOptions=grid_options,
            allow_unsafe_jscode=True,
            height=height,
            reload_data=True,
            data_return_mode=DataReturnMode.AS_INPUT,
            key=f"Activités programmées {st.session_state.activites_programmees_key_counter}"  # incrémentation de la clef permet de forcer le reaffichage 
        )

        # Affichage de l'erreur renvoyée par le précédent run
        erreur = st.session_state.get("aggrid_activites_programmees_erreur") 
        if erreur is not None:
            st.error(erreur)

        event_data = response.get("event_data")
        event_type = event_data["type"] if isinstance(event_data, dict) else None
        tracer.log(f"event {event_type}", types=["gen", "event"])

        # Pas d'event aggrid à traiter si event_type is None (i.e. le script python est appelé pour autre chose qu'un event aggrid)
        if event_type is None:
            if len(df_display) == 0:
                if st.session_state.menu_activites["menu"] == "menu_activites_programmees":
                    st.session_state.menu_activites = {
                        "menu": "menu_activites_programmees",
                        "index_df": None
                    }
            return

        # Récupération du retour grille __sel_source
        # Cette information est passée à la valeur "user" par le JsCode JS_SELECT_DESELECT_ONCE si le cellValueChanged provient d'un click utilisateur.
        # Elle permet de n'effectuer les traitements de cellValueChanged que sur les seuls évènements utilisateurs et de bypasser ceux provenant d'une
        # demande de sélection programmée via demander_selection().
        sel_source = "unknown"
        try:
            df_dom = pd.DataFrame(response["data"]) if "data" in response and isinstance(response["data"], pd.DataFrame) else pd.DataFrame()  
        except:
            df_dom = pd.DataFrame() 
        if not df_dom.empty:
            first_row = df_dom.iloc[0]
            sel_source = (first_row.get("__sel_source") or "api") # 'user' ou 'api'
            tracer.log(f"sel_source {sel_source}", types=["sel_source"])

        # Récupération de la ligne sélectionnée courante
        selected_rows = response["selected_rows"] if "selected_rows" in response else None
        if not selection_demandee:
            if isinstance(selected_rows, pd.DataFrame) and not selected_rows.empty:
                # tracer.log("row = selected_rows.iloc[0]")
                row = selected_rows.iloc[0] 
            elif isinstance(selected_rows, list) and len(selected_rows) > 0:
                # tracer.log("row = selected_rows[0]")
                row = selected_rows[0]

        # 🟡 Traitement si ligne sélectionnée et index correspondant non vide
        if row is not None:

            # Récupération de l'index de ligne sélectionnée
            index_df = row["__index"]

            # Evènement de type "selectionChanged" 
            if event_type == "selectionChanged":
                # tracer.log(f"Selected row {selected_rows.iloc[0]["__index"] if isinstance(selected_rows, pd.DataFrame) and not selected_rows.empty else (selected_rows[0]["__index"] if isinstance(selected_rows, list) and len(selected_rows) > 0 else None)}")
                if index_df != st.session_state.activites_programmees_sel_request["sel"]["id"] and not deselection_demandee and sel_source == "user":
                    # tracer.log(f"***activites_programmees_sel_request[id] de {st.session_state.activites_programmees_sel_request["sel"]["id"]} à {index_df}")
                    st.session_state.activites_programmees_sel_request["sel"]["id"] = index_df
                    # tracer.log(f"***demander_deselection activites_non_programmees")
                    demander_deselection("activites_non_programmees")
                    
                    # time.sleep(0.05) # Hack défensif pour éviter les erreurs Connection error Failed to process a Websocket message Cached ForwardMsg MISS

                    if not st.session_state.forcer_menu_activites_non_programmees:
                        st.session_state.editeur_activite_idx = index_df
                        st.session_state.menu_activites = {
                            "menu": "menu_activites_programmees",
                            "index_df": index_df
                        }
                    st.rerun()
                else:
                    if st.session_state.forcer_menu_activites_programmees or st.session_state.forcer_maj_menu_activites_programmees:
                        st.session_state.editeur_activite_idx = index_df
                        st.session_state.menu_activites = {
                            "menu": "menu_activites_programmees",
                            "index_df": index_df
                        }
                        st.session_state.forcer_maj_menu_activites_programmees = False
                        
            # Gestion des modifications de cellules
            # Attention : la modification de cellule uniquement sur "cellValueChanged" n'est pas suffisante, car lorsque l'on valide la modification
            # de cellule en cliquant sur une autre ligne, on reçoit un event de type "selectionChanged" et non "cellValueChanged". Mais cela implique 
            # que toute modification programmée de cellule (via l'éditeur d'activité ou les boutons de programmation) va engendrée un écart entre le 
            # df_display modifié par programmation et le df_dom revenant du DOM, ce qui, via le code ci-dessous, va déclencher une modification inverse 
            # à celle souhaitée. Pour eviter cela il faut :
            # 1. Mettre en place un mécanisme de requête de modification qui bypasse la modification de cellule tant que le DOM n'a pas enregistré les 
            #    modifications demandées via le df_display (voir reprogrammation_request et row_modification_request).
            # 2. S'assurer que le DOM renvoie bien via response["data"] les modifications enregistrées. Ceci est réalisé par l'incrémentation de la 
            #    colonne de travail __df_push_ver qui via le JsCode 
            if not df_dom.empty:
            # if isinstance(response["data"], pd.DataFrame):

                bypass_cell_edit = False

                # Si une requete de reprogrammation est en cours sur index_df on bypasse la gestion de modification de cellules
                # jusquà ce que le DOM ait enregistré la reprogrammation. Sinon une modification de valeur sur index_df est détectée 
                # et déclenche une reprogrammation inverse à celle demandée.
                reprogrammation_request = reprogrammation_request_get()
                if reprogrammation_request is not None:
                    if reprogrammation_request["idx"] == index_df:
                        matching = df_dom.index[df_dom["__index"] == index_df]
                        if not matching.empty:
                            if reprogrammation_request["jour"] == safe_int(df_dom.at[matching[0], "Date"]): # la modification de date a été prise en compte par le DOM
                                reprogrammation_request_del()
                            else:
                                bypass_cell_edit = True

                # Si une requete de modification de ligne est en cours sur index_df on bypasse la gestion de modification de cellules
                # jusquà ce que le DOM ait enregistré la modification de ligne. Sinon une modification de valeur sur index_df est détectée 
                # et déclenche une modification inverse à celle demandée.
                row_modification_request = row_modification_request_get()
                if row_modification_request is not None:
                    if row_modification_request["idx"] == index_df:
                        matching = df_dom.index[df_dom["__index"] == index_df]
                        if not matching.empty:
                            for col, val in row_modification_request["cols"].items():
                                val_dom = df_dom.at[matching[0], df_display_col_nom(col)]
                                if (pd.isna(val_dom) and pd.notna(val)) or str(val_dom) != str(val): # la modification de date a été prise en compte par le DOM
                                    bypass_cell_edit = True
                            if not bypass_cell_edit:
                                row_modification_request_del()

                if not bypass_cell_edit:
                    i, idx = get_ligne_modifiee(df_dom, st.session_state.activites_programmees_df_display_copy, columns_to_drop=work_cols)
                    if i is not None:
                        if idx == index_df: # on ne considère que les modifications sur la ligne ayant généré l'event
                            st.session_state.aggrid_activites_programmees_erreur = None
                            for col in [col for col in df_dom.columns if col not in non_editable_cols]:
                                col_df = RENOMMAGE_COLONNES_INVERSE[col] if col in RENOMMAGE_COLONNES_INVERSE else col
                                if pd.isna(df.at[idx, col_df]) and pd.isna(df_dom.at[i, col]):
                                    continue
                                if col == "Date":
                                    if df_dom.at[i, col] == "":
                                        # Déprogrammation de l'activité (Suppression de l'activité des activités programmées)
                                        undo.save()
                                        demander_selection("activites_non_programmees", idx, deselect="activites_programmees", visible_id=ligne_voisine_index(df_display, idx))
                                        activites_programmees_deprogrammer(idx)
                                        demander_selection("creneaux_disponibles", get_creneau_proche(st.session_state.get("creneaux_disponibles"), df.loc[idx])[0])
                                        st.session_state["activites_programmables_selected_row"] = df.loc[idx]
                                        st.rerun()
                                    elif pd.isna(df.at[idx, "Date"]) or df_dom.at[i, col] != str(int(df.at[idx, "Date"])):
                                        # Reprogrammation de l'activité à la date choisie
                                        jour_choisi = int(df_dom.at[i, col])
                                        undo.save()
                                        demander_selection("activites_programmees", idx, deselect="activites_non_programmees")
                                        activites_programmees_reprogrammer(idx, jour_choisi)
                                        demander_selection("creneaux_disponibles", get_creneau_proche(st.session_state.get("creneaux_disponibles"), df.loc[idx])[0])
                                        st.rerun()
                                else:
                                    if (pd.isna(df.at[idx, col_df]) and pd.notna(df_dom.at[i, col])) or df.at[idx, col_df] != df_dom.at[i, col]:
                                        demander_selection("activites_programmees", idx, deselect="activites_non_programmees")
                                        activites_programmees_modifier_cellule(idx, col_df, df_dom.at[i, col])
                                        demander_selection("creneaux_disponibles", get_creneau_proche(st.session_state.get("creneaux_disponibles"), df.loc[idx])[0])
                                        st.rerun()

# Section critique pour la déprogrammation d'une activité programmée.
# Section critique car la modification de cellule depuis la grille est validée par un click row 
# qui peut entraîner une interruption du script python et donc une incohérence de contexte.
# Le mécanisme de section critique permet une relance automatique du traitement jusqu'à complétion 
# en cas d'interruption par un rerun Streamlit : une commande est enregistrée dans st.session_state 
# et est automatiquement relancée en début de rerun tant qu'elle n'est pas terminée.
def activites_programmees_deprogrammer(idx):
    
    st.session_state.setdefault("activites_programmees_deprogrammer_cmd", 
        {
            "idx": idx,
            "step": 0,
        }
    )

    tracer.log(f"Début {idx}")

    st.session_state.forcer_menu_activites_non_programmees = True
    bd_deprogrammer_activite_programmee(idx)

    # Workaround pour forcer le réaffichage de la grille.
    # Sinon figeage grille après modification de cellule.
    forcer_reaffichage_activites_programmees() 

    forcer_reaffichage_df("creneaux_disponibles")
    sql.sauvegarder_row(idx)

    tracer.log(f"Fin {idx}")
    del st.session_state["activites_programmees_deprogrammer_cmd"]

# Section critique pour la reprogrammation d'une activité programmée.
# Section critique car la modification de cellule depuis la grille est validée par un click row 
# qui peut entraîner une interruption du script python et donc une incohérence de contexte.
# Le mécanisme de section critique permet une relance automatique du traitement jusqu'à complétion 
# en cas d'interruption par un rerun Streamlit : une commande est enregistrée dans st.session_state 
# et est automatiquement relancée en début de rerun tant qu'elle n'est pas terminée.
def activites_programmees_reprogrammer(idx, jour):
    
    st.session_state.setdefault("activites_programmees_reprogrammer_cmd", 
        {
            "idx": idx,
            "jour": jour,
            "step": 0,
        }
    )

    tracer.log(f"Début {idx} {jour}")

    bd_modifier_cellule(idx, "Date", jour)

    # Workaround pour forcer le réaffichage de la grille.
    # Sinon figeage grille après modification de cellule.
    forcer_reaffichage_activites_programmees() 

    sql.sauvegarder_row(idx)

    tracer.log(f"Fin {idx} {jour}")
    del st.session_state["activites_programmees_reprogrammer_cmd"]

# Section critique pour la modification de cellules d'une activité programmée.
# Section critique car la modification de cellule depuis la grille est validée par un click row 
# qui peut entraîner une interruption du script python et donc une incohérence de contexte.
# Le mécanisme de section critique permet une relance automatique du traitement jusqu'à complétion 
# en cas d'interruption par un rerun Streamlit : une commande est enregistrée dans st.session_state 
# et est automatiquement relancée en début de rerun tant qu'elle n'est pas terminée.
def activites_programmees_modifier_cellule(idx, col, val):
    
    st.session_state.setdefault("activites_programmees_modifier_cellule_cmd", 
        {
            "idx": idx,
            "col": col,
            "val": val,
            "step": 0,
        }
    )

    tracer.log(f"Début {idx} {col} {val}")

    erreur = affecter_valeur_df(idx, col, val, section_critique=st.session_state.activites_programmees_modifier_cellule_cmd)

    # Workaround pour forcer le réaffichage de la grille.
    # Sinon figeage grille après modification de cellule.
    forcer_reaffichage_activites_programmees() 

    if not erreur:
        if col in ["Debut", "Duree", "Activité"]:
            forcer_reaffichage_df("creneaux_disponibles")
    else:
        st.session_state.aggrid_activites_programmees_erreur = erreur

    tracer.log(f"Fin {idx} {col} {val}")
    del st.session_state["activites_programmees_modifier_cellule_cmd"]

def reprogrammation_request_set(idx, jour):
    st.session_state.setdefault("reprogrammation_request", 
        {
            "idx": idx,
            "jour": jour,
        }
    )

def reprogrammation_request_get():
    return st.session_state.get("reprogrammation_request")

def reprogrammation_request_del():
    if "reprogrammation_request" in st.session_state:
        del st.session_state["reprogrammation_request"]

def row_modification_request_set(idx, cols):
    st.session_state.setdefault("row_modification_request", 
        {
            "idx": idx,
            "cols": cols,
        }
    )

def row_modification_request_get():
    return st.session_state.get("row_modification_request")

def row_modification_request_del():
    if "row_modification_request" in st.session_state:
        del st.session_state["row_modification_request"]

# Menu activité à afficher dans la sidebar si click dans aggrid d'activités programmées         }
def menu_activites_programmees(index_df):

    df = st.session_state.df
    df_display = st.session_state.activites_programmees_df_display
    nom_activite = df.at[index_df, "Activite"] if  isinstance(df, pd.DataFrame) and index_df is not None and index_df in df.index else ""
    nom_activite = nom_activite.strip() if pd.notna(nom_activite) else ""

    boutons_disabled = nom_activite == "" or pd.isna(index_df) or not isinstance(df, pd.DataFrame) or (isinstance(df, pd.DataFrame) and len(df) == 0)
    activite_reservee = est_activite_reserve(df.loc[index_df]) if pd.notna(index_df) else True 
    jours_possibles = get_jours_possibles(df, st.session_state.activites_programmees, index_df)

    # Affichage du label d'activité
    afficher_nom_activite(df, index_df, nom_activite)

    # Affichage du contrôle recherche sur le Web
    afficher_bouton_web(nom_activite, disabled=boutons_disabled or est_nom_pause(nom_activite))

    # Affichage du contrôle recherche itinéraire
    afficher_bouton_itineraire(df.loc[index_df, "Lieu"] if pd.notna(index_df) and len(df) > 0 else "")

    # Affichage contrôle Supprimer
    if st.button(LABEL_BOUTON_SUPPRIMER, use_container_width=CENTRER_BOUTONS, disabled=boutons_disabled or activite_reservee, key="menu_activite_supprimer"):
        undo.save()
        demander_selection("activites_programmees", ligne_voisine_index(df_display, index_df), deselect="activites_non_programmees")
        demander_selection("creneaux_disponibles", get_creneau_proche(st.session_state.get("creneaux_disponibles"), df.loc[index_df])[0])
        st.session_state.forcer_maj_menu_activites_programmees = True
        supprimer_activite(index_df)
        forcer_reaffichage_df("creneaux_disponibles")
        sql.sauvegarder_row(index_df)
        st.rerun()

    # Affichage contrôle Déprogrammer
    if st.button(LABEL_BOUTON_DEPROGRAMMER, use_container_width=CENTRER_BOUTONS, disabled=boutons_disabled or activite_reservee or est_nom_pause(nom_activite), key="menu_activite_deprogrammer"):
        undo.save()
        st.session_state.forcer_menu_activites_non_programmees = True
        demander_selection("activites_non_programmees", index_df, deselect="activites_programmees")
        bd_deprogrammer_activite_programmee(index_df)
        demander_selection("creneaux_disponibles", get_creneau_proche(st.session_state.get("creneaux_disponibles"), df.loc[index_df])[0])
        st.session_state["activites_programmables_selected_row"] = df.loc[index_df]
        forcer_reaffichage_df("creneaux_disponibles")
        sql.sauvegarder_row(index_df)
        st.rerun()

    # Affichage contrôle Reprogrammer
    if st.button(LABEL_BOUTON_REPROGRAMMER, use_container_width=True, disabled=boutons_disabled or activite_reservee or est_nom_pause(nom_activite) or not jours_possibles, key="menu_activite_programmer"):
        if "activites_programmees_jour_choisi" in st.session_state:
            jour_choisi = st.session_state.activites_programmees_jour_choisi
            undo.save()
            demander_selection("activites_programmees", index_df, deselect="activites_non_programmees")
            reprogrammation_request_set(index_df, int(jour_choisi)) # inhibe les cellValuChanged résultant de cette modification et qui inverseraient l'opération
            bd_modifier_cellule(index_df, "Date", int(jour_choisi))
            demander_selection("creneaux_disponibles", get_creneau_proche(st.session_state.get("creneaux_disponibles"), df.loc[index_df])[0])
            sql.sauvegarder_row(index_df)
            st.rerun()
    
    # Affichage Liste des jours possibles
    jours_label = [f"{int(jour):02d}" for jour in jours_possibles]
    if jours_label and (not st.session_state.get("menu_activite_choix_jour_programmation") or st.session_state.menu_activite_choix_jour_programmation not in jours_label):
            st.session_state.menu_activite_choix_jour_programmation = jours_label[0]
    st.session_state.activites_programmees_jour_choisi = st.selectbox("Jours possibles", jours_label, label_visibility="visible", disabled=boutons_disabled or activite_reservee or not jours_possibles, key = "menu_activite_choix_jour_programmation") 
        
    # Affichage de l'éditeur d'activité
    if st.button(LABEL_BOUTON_EDITER, use_container_width=CENTRER_BOUTONS, disabled=boutons_disabled, key="menu_activite_bouton_editer"):
        if "editeur_activite_etat" in st.session_state:
            del st.session_state["editeur_activite_etat"]
        show_dialog_editeur_activite(df, index_df)
                               
    # Affichage du contrôle Ajouter
    afficher_bouton_nouvelle_activite(key="menu_activite_bouton_nouvelle_activite")

# Initialisation des grid_options sur la grille des activités non programmées
def init_activites_non_programmees_grid_options(df_display):

    gb = GridOptionsBuilder.from_dataframe(df_display)

    # Configuration par défaut des colonnes
    gb.configure_default_column(resizable=True)

    # Colonnes à largeur fixe
    colonnes_fixes = {"Date": 55, "Début": 55, "Fin": 55, "Durée": 55}
    for col, width in colonnes_fixes.items():
        gb.configure_column(
            col,
            filter=False,
            resizable=False,
            width=width,
            minWidth=width,
            maxWidth=width,
            flex=0,
            suppressSizeToFit=True,
        )

    # Epinglage de la colonne Date
    gb.configure_column(
        "Date",
        pinned=JsCode("'left'")
    )

    # Masquage des colonnes de travail
    work_cols = ACTIVITES_NON_PROGRAMMEES_WORK_COLS
    for col in work_cols:
        gb.configure_column(col, hide=True)

    # Colonnes editables
    non_editable_cols = ["Fin"] + work_cols
    for col in df_display.columns:
        gb.configure_column(col, editable=(col not in non_editable_cols))

    # Configuration des menus de la colonne Date
    gb.configure_column(
        "Date",
        editable=True,
        cellEditor="agSelectCellEditor",
        cellEditorParams=JsCode("""
            function(params) {
                let raw = params.data.__options_date;
                let values = [];

                try {
                    values = JSON.parse(raw);
                } catch (e) {
                    values = [];
                }

                return { values: values };
            }
        """)
    )

    # Configuration des icônes de colonnes pour la recherche Web et la recherche d'itinéraire
    # gb.configure_column("Activité", editable=True, cellRenderer=JS_ACTIVITE_ICON_RENDERER) #, minWidth=220)
    # gb.configure_column("Lieu", editable=True, cellRenderer=JS_LIEU_ICON_RENDERER) #, minWidth=200)

    # Configuration de l'appui long pour la recherche Web et la recherche d'itinéraire
    gb.configure_column("Activité", editable=True, cellRenderer=JS_ACTIVITE_LONGPRESS_RENDERER) #, minWidth=220)
    gb.configure_column("Lieu",     editable=True, cellRenderer=JS_LIEU_LONGPRESS_RENDERER) #, minWidth=200)

    # Colorisation 
    gb.configure_grid_options(getRowStyle= JsCode(f"""
        function(params) {{
            if (params.data.__options_date !== "[]") {{
                return {{
                    'backgroundColor': '{COULEUR_ACTIVITE_PROGRAMMABLE}'
                }}
            }}
            return null;
        }}
        """))

    # Configuration de la sélection
    gb.configure_selection(selection_mode="single", use_checkbox=False) 
    
    gb.configure_grid_options(
        getRowNodeId=JsCode("function(data) { return String(data.__uuid); }"),
        getRowId=JsCode("function(p){ return String(p.data.__uuid); }"),
        columnTypes={"textColumn": {}},  # évite l'erreur #36
        onGridReady=JS_SELECT_DESELECT_ONCE,
        onFirstDataRendered=JS_IOS_SOFT_REVIVE,
    )

    # Mise en page de la grille (repris dans JS_IOS_SOFT_REVIVE)
    # gb.configure_grid_options(onFirstDataRendered=JsCode(f"""
    #     function(params) {{
    #         params.api.sizeColumnsToFit();
    #     }}
    # """))

    grid_options = gb.build()

    # Empêche la possibilité de réorganiser les colonnes
    grid_options["suppressMovableColumns"] = True

    # Supprime le highlight de survol qui pose problème sur mobile et tablette
    grid_options["suppressRowHoverHighlight"] = True

    # Enregistre dans le contexte les paramètres nécessaires à la recherche d'itinéraire (voir JS_LIEU_ICON_RENDERER)
    grid_options["context"] = {
        "itineraire_app": st.session_state.get("itineraire_app", "Google Maps"),
        "platform": get_platform(),  # "iOS" / "Android" / "Desktop"
    }

    return grid_options

# Affiche les activités non programmées dans un tableau
def afficher_activites_non_programmees():

    df = st.session_state.get("df")
    if df is None:
        return
    
    df_display = st.session_state.get("activites_non_programmees_df_display")
    if df_display is None:
        return
    
    work_cols = ACTIVITES_NON_PROGRAMMEES_WORK_COLS
    non_editable_cols = ["Fin"] + work_cols

    # Calcul de la hauteur de l'aggrid
    nb_lignes = len(df_display)
    ligne_px = 30  # hauteur approximative d’une ligne dans AgGrid
    max_height = 250
    height = min(nb_lignes * ligne_px + 50, max_height)

    # Initialisation du compteur qui permet de savoir si l'on doit forcer le réaffichage de l'aggrid après une suppression de ligne 
    st.session_state.setdefault("activites_non_programmees_key_counter", 0)
    
    # Initialisation de la variable d'état indiquant s'il convient de bypasser la section d'édition de cellule 
    st.session_state.setdefault("activites_non_programmees_bypass_cell_edit", False)

    # Initialisation de la variable d'état contenant la requête de selection / déselection
    st.session_state.setdefault("activites_non_programmees_sel_request", copy.deepcopy(SEL_REQUEST_DEFAUT))

    # Gestion des sélections / désélections demandées via demander_selection() demander_deselection()
    # Utilise le JS code JS_SELECT_DESELECT_ONCE lequel exploite les colonnes de travail suivantes:
    # __sel_id = id de la ligne à sélectionner (None si pas de contrainte de sélection)
    # __sel_ver = version de la demande de sélection (doit être incrémentée à chaque demande)
    # __desel_id = id de la ligne devant rester visible lors de la déselection (None si aucune contrainte de visibilité lors de la désélection)
    # __desel_ver = version de la demande de désélection (doit être incrémentée à chaque demande)
    # __sel_source = information renvoyée par le DOM (event.source exposé par onRowSelected) indiquant si la source de selection est "user" ou "api" selon que la demande de sélection provient d'un click utilisateur ou d'une requête python via JsCode
    # Ces colonnes sont configurées par les fonctions utilisateur demander_selection() et demander_deselection()
    # L'information de retour __sel_source est exploitée par le mecanisme de flip flop entre grille A et grille B
    # via le champ "data" de la réponse de l'aggrid (cf. fonction afficher_activites_programmees() et afficher_activites_non_programmees())
    row = None
    selection_demandee = False
    sel_request = st.session_state.get("activites_non_programmees_sel_request")
    if sel_request["sel"]["pending"]:
        if sel_request["sel"]["id"] is not None:
            reqid = sel_request["sel"]["id"]
            # tracer.log(f"Traitement de la requête de sélection {sel_request["sel"]["id"]} {sel_request["sel"]["ver"]}")
            df_display["__sel_id"] = get_uuid(df_display, reqid)
            df_display["__sel_ver"] = sel_request["sel"]["ver"]
            if reqid in df_display.index: 
                row = df_display.loc[reqid]
                # tracer.log(f"row = df_display.loc[{reqid}]")
            selection_demandee = True
        st.session_state.activites_non_programmees_sel_request["sel"]["pending"] = False

    deselection_demandee = False
    if sel_request["desel"]["pending"]:
        # tracer.log(f"Traitement de la requête de desélection {sel_request["desel"]["ver"]}")
        df_display["__desel_ver"] = sel_request["desel"]["ver"]
        df_display["__desel_id"] = get_uuid(df_display, sel_request["desel"]["id"]) # id visible après déselection, None si pas de contrainte de visibilité
        df_display["__sel_id"] = None
        deselection_demandee = True
        st.session_state.activites_non_programmees_sel_request["desel"]["pending"] = False

    # if len(df_display) > 0:
    #     tracer.log(f"df_display['__sel_id'] {df_display.iloc[0]["__sel_id"]} df_display['__sel_ver'] {df_display.iloc[0]["__sel_ver"]} df_display['__desel_ver'] {df_display.iloc[0]["__desel_ver"]}")

    grid_options = init_activites_non_programmees_grid_options(df_display)

    # Affichage
    with st.expander("**Activités non programmées**", expanded=True):
        response = AgGrid(
            df_display,
            gridOptions=grid_options,
            allow_unsafe_jscode=True,
            height=height,
            reload_data=True,
            data_return_mode=DataReturnMode.AS_INPUT,
            update_mode=(GridUpdateMode.MODEL_CHANGED | GridUpdateMode.VALUE_CHANGED
                        | GridUpdateMode.SELECTION_CHANGED),
            key=f"Activités non programmées {st.session_state.activites_non_programmees_key_counter}",  # incrémentation de la clef permet de forcer le reaffichage
        )

        # Affichage de l'erreur renvoyée par le précédent run
        erreur = st.session_state.get("aggrid_activites_non_programmees_erreur") 
        if erreur is not None:
            st.error(erreur)

        event_data = response.get("event_data")
        event_type = event_data["type"] if isinstance(event_data, dict) else None
        tracer.log(f"event {event_type}", types=["gen", "event"])

        # Pas d'event aggrid à traiter si event_type is None (i.e. le script python est appelé pour autre chose qu'un event aggrid)
        if event_type is None:
            if len(df_display) == 0:
                if st.session_state.menu_activites["menu"] == "menu_activites_non_programmees":
                    st.session_state.menu_activites = {
                        "menu": "menu_activites_non_programmees",
                        "index_df": None
                    }
            return
        
        # Récupération du retour grille __sel_source
        # Cette information est passée à la valeur "user" par le JsCode JS_SELECT_DESELECT_ONCE si le cellValueChanged provient d'un click utilisateur.
        # Elle permet de n'effectuer les traitements de cellValueChanged que sur les seuls évènements utilisateurs et de bypasser ceux provenant d'une
        # demande de sélection programmée via demander_selection().
        sel_source = "unknown"
        try:
            df_dom = pd.DataFrame(response["data"]) if "data" in response and isinstance(response["data"], pd.DataFrame) else pd.DataFrame()  
        except:
            df_dom = pd.DataFrame() 
        if not df_dom.empty:
            first_row = df_dom.iloc[0]
            sel_source = (first_row.get("__sel_source") or "api") # 'user' ou 'api'
            tracer.log(f"sel_source {sel_source}", types=["sel_source"])

        # Récupération de la ligne sélectionnée
        selected_rows = response["selected_rows"] if "selected_rows" in response else None
        row = None
        if not selection_demandee:
            if isinstance(selected_rows, pd.DataFrame) and not selected_rows.empty:
                # tracer.log("row = selected_rows.iloc[0]")
                row = selected_rows.iloc[0] 
            elif isinstance(selected_rows, list) and len(selected_rows) > 0:
                # tracer.log("row = selected_rows[0]")
                row = selected_rows[0]

        # 🟡 Traitement si ligne sélectionnée et index correspondant non vide
        if row is not None:

            # Récupération de l'index de ligne sélectionnée
            index_df = row["__index"]

            # Evènement de type "selectionChanged"
            if event_type == "selectionChanged":
                # tracer.log(f"Selected row {selected_rows.iloc[0]["__index"] if isinstance(selected_rows, pd.DataFrame) and not selected_rows.empty else (selected_rows[0]["__index"] if isinstance(selected_rows, list) and len(selected_rows) > 0 else None)}")
                if index_df != st.session_state.activites_non_programmees_sel_request["sel"]["id"] and not deselection_demandee and sel_source == "user":
                    # tracer.log(f"***activites_non_programmees_sel_request[id] de {st.session_state.activites_non_programmees_sel_request["sel"]["id"]} à {index_df}")
                    st.session_state.activites_non_programmees_sel_request["sel"]["id"] = index_df
                    # tracer.log(f"***demander_deselection activites_programmees")
                    demander_deselection("activites_programmees")

                    # time.sleep(0.05) # Hack défensif pour éviter les erreurs Connection error Failed to process a Websocket message Cached ForwardMsg MISS

                    if not st.session_state.forcer_menu_activites_programmees:
                        st.session_state.editeur_activite_idx = index_df
                        st.session_state.menu_activites = {
                            "menu": "menu_activites_non_programmees",
                            "index_df": index_df
                        }
                    st.rerun()
                else:
                    if st.session_state.forcer_menu_activites_non_programmees or st.session_state.forcer_maj_menu_activites_non_programmees:
                        st.session_state.editeur_activite_idx = index_df
                        st.session_state.menu_activites = {
                            "menu": "menu_activites_non_programmees",
                            "index_df": index_df
                        }
                        st.session_state.forcer_maj_menu_activites_non_programmees = False

            # Gestion des modifications de cellules
            # Attention : la modification de cellule uniquement sur "cellValueChanged" n'est pas suffisante, car lorsque l'on valide la modification
            # de cellule en cliquant sur une autre ligne, on reçoit un event de type "selectionChanged" et non "cellValueChanged". Mais cela implique 
            # que toute modification programmée de cellule (via l'éditeur d'activité ou les boutons de programmation) va engendrée un écart entre le 
            # df_display modifié par programmation et le df_dom revenant du DOM, ce qui, via le code ci-dessous, va déclencher une modification inverse 
            # à celle souhaitée. Pour eviter cela il faut :
            # 1. Mettre en place un mécanisme de requête de modification qui bypasse la modification de cellule tant que le DOM n'a pas enregistré les 
            #    modifications demandées via le df_display (voir reprogrammation_request et row_modification_request).
            # 2. S'assurer que le DOM renvoie bien via response["data"] les modifications enregistrées. Ceci est réalisé par l'incrémentation de la 
            #    colonne de travail __df_push_ver qui via le JsCode 
            if not df_dom.empty:
            # if isinstance(response["data"], pd.DataFrame):

                bypass_cell_edit = False

                # Si une requete de modification de ligne est en cours sur index_df on bypasse la gestion de modification de cellules
                # jusquà ce que le DOM ait enregistré la modification de ligne. Sinon une modification de valeur sur index_df est détectée 
                # et déclenche une modification inverse à celle demandée.
                row_modification_request = row_modification_request_get()
                if row_modification_request is not None:
                    if row_modification_request["idx"] == index_df:
                        matching = df_dom.index[df_dom["__index"] == index_df]
                        if not matching.empty:
                            for col, val in row_modification_request["cols"].items():
                                val_dom = df_dom.at[matching[0], df_display_col_nom(col)]
                                if (pd.isna(val_dom) and pd.notna(val)) or str(val_dom) != str(val): # la modification de date a été prise en compte par le DOM
                                    bypass_cell_edit = True
                            if not bypass_cell_edit:
                                row_modification_request_del()

                if not bypass_cell_edit:
                    i, idx = get_ligne_modifiee(df_dom, st.session_state.activites_non_programmees_df_display_copy, columns_to_drop=work_cols)
                    if i is not None:
                        if idx == index_df: # on ne considère que les modifications sur la ligne ayant généré l'event
                            st.session_state.aggrid_activites_non_programmees_erreur = None
                            for col in [col for col in df_dom.columns if col not in non_editable_cols]:
                                col_df = RENOMMAGE_COLONNES_INVERSE[col] if col in RENOMMAGE_COLONNES_INVERSE else col
                                if pd.isna(df.at[idx, col_df]) and pd.isna(df_dom.at[i, col]):
                                    continue
                                if col == "Date":
                                    if df_dom.at[i, col] != "":
                                        # Programmation de l'activité à la date choisie
                                        jour_choisi = int(df_dom.at[i, col])
                                        undo.save()
                                        demander_selection("activites_programmees", idx, deselect="activites_non_programmees")
                                        activites_non_programmees_programmer(idx, jour_choisi)
                                        demander_selection("creneaux_disponibles", get_creneau_proche(st.session_state.get("creneaux_disponibles"), df.loc[idx])[0])
                                        st.rerun()
                                else:
                                    if (pd.isna(df.at[idx, col_df]) and pd.notna(df_dom.at[i, col])) or df.at[idx, col_df] != df_dom.at[i, col]:
                                        demander_selection("activites_non_programmees", idx, deselect="activites_programmees")
                                        activites_non_programmees_modifier_cellule(idx, col_df, df_dom.at[i, col])
                                        demander_selection("creneaux_disponibles", get_creneau_proche(st.session_state.get("creneaux_disponibles"), df.loc[idx])[0])
                                        st.rerun()

        elif len(df_display) == 0:
            if st.session_state.menu_activites["menu"] == "menu_activites_non_programmees":
                st.session_state.menu_activites = {
                    "menu": "menu_activites_non_programmees",
                    "index_df": None
                }

# Section critique pour la programmation d'une activité non programmée.
# Section critique car la modification de cellule depuis la grille est validée par un click row 
# qui peut entraîner une interruption du script python et donc une incohérence de contexte.
# Le mécanisme de section critique permet une relance automatique du traitement jusqu'à complétion 
# en cas d'interruption par un rerun Streamlit : une commande est enregistrée dans st.session_state 
# et est automatiquement relancée en début de rerun tant qu'elle n'est pas terminée.
def activites_non_programmees_programmer(idx, jour):
    
    st.session_state.setdefault("activites_non_programmees_programmer_cmd", 
        {
            "idx": idx,
            "jour": jour,
            "step": 0,
        }
    )

    tracer.log(f"Début {idx} {jour}")

    st.session_state.forcer_menu_activites_programmees = True
    bd_modifier_cellule(idx, "Date", int(jour))

    # Workaround pour forcer le réaffichage de la grille.
    # Sinon figeage grille après modification de cellule.
    forcer_reaffichage_activites_non_programmees() 

    forcer_reaffichage_df("creneaux_disponibles")
    sql.sauvegarder_row(idx)

    tracer.log(f"Fin {idx} {jour}")
    del st.session_state["activites_non_programmees_programmer_cmd"]

# Section critique pour la modification de cellules d'une activité non programmée.
# Section critique car la modification de cellule depuis la grille est validée par un click row 
# qui peut entraîner une interruption du script python et donc une incohérence de contexte.
# Le mécanisme de section critique permet une relance automatique du traitement jusqu'à complétion 
# en cas d'interruption par un rerun Streamlit : une commande est enregistrée dans st.session_state 
# et est automatiquement relancée en début de rerun tant qu'elle n'est pas terminée.
def activites_non_programmees_modifier_cellule(idx, col, val):
    
    st.session_state.setdefault("activites_non_programmees_modifier_cellule_cmd", 
        {
            "idx": idx,
            "col": col,
            "val": val,
            "step": 0,
        }
    )

    tracer.log(f"Début {idx} {col} {val}")

    erreur = affecter_valeur_df(idx, col, val, section_critique=st.session_state.activites_non_programmees_modifier_cellule_cmd)

    # Workaround pour forcer le réaffichage de la grille.
    # Sinon figeage grille après modification de cellule.
    forcer_reaffichage_activites_non_programmees() 
    
    if not erreur:
        forcer_reaffichage_df("activites_programmables")
    else:
        st.session_state.aggrid_activites_non_programmees_erreur = erreur

    tracer.log(f"Fin {idx} {col} {val}")
    del st.session_state["activites_non_programmees_modifier_cellule_cmd"]

# Menu activité à afficher dans la sidebar si click dans aggrid d'activités non programmées         }
def menu_activites_non_programmees(index_df):

    df = st.session_state.df
    df_display = st.session_state.activites_non_programmees_df_display
    nom_activite = df.at[index_df, "Activite"] if  isinstance(df, pd.DataFrame) and index_df is not None and index_df in df.index else ""
    nom_activite = nom_activite.strip() if pd.notna(nom_activite) else ""

    boutons_disabled = nom_activite == "" or pd.isna(index_df) or not isinstance(df, pd.DataFrame) or (isinstance(df, pd.DataFrame) and len(df) == 0)
    jours_possibles = get_jours_possibles(df, st.session_state.activites_programmees, index_df)

    # Affichage du label d'activité
    afficher_nom_activite(df, index_df, nom_activite)

    # Affichage du contrôle recherche sur le Web
    afficher_bouton_web(nom_activite, disabled=boutons_disabled or est_nom_pause(nom_activite))

    # Affichage du contrôle recherche itinéraire
    afficher_bouton_itineraire(df.loc[index_df, "Lieu"] if pd.notna(index_df) and len(df) > 0 else "")

    # Affichage contrôle Supprimer
    if st.button(LABEL_BOUTON_SUPPRIMER, use_container_width=CENTRER_BOUTONS, disabled=boutons_disabled, key="menu_activite_supprimer"):
        undo.save()
        demander_selection("activites_non_programmees", ligne_voisine_index(df_display, index_df), deselect="activites_programmees")
        demander_selection("creneaux_disponibles", get_creneau_proche(st.session_state.get("creneaux_disponibles"), df.loc[index_df])[0])
        st.session_state.forcer_maj_menu_activites_non_programmees = True
        supprimer_activite(index_df)
        forcer_reaffichage_df("activites_programmable_dans_creneau_selectionne")
        sql.sauvegarder_row(index_df)
        st.rerun()

    # Affichage contrôle Deprogrammer
    st.button(LABEL_BOUTON_DEPROGRAMMER, use_container_width=CENTRER_BOUTONS, disabled=True, key="menu_activite_deprogrammer")

    # Affichage contrôle Programmer
    if st.button(LABEL_BOUTON_PROGRAMMER, use_container_width=CENTRER_BOUTONS, disabled=boutons_disabled or not jours_possibles, key="menu_activite_programmer"):
        if "activites_non_programmees_jour_choisi" in st.session_state:
            jour_choisi = st.session_state.activites_non_programmees_jour_choisi
            undo.save()
            st.session_state.forcer_menu_activites_programmees = True
            demander_selection("activites_programmees", index_df, deselect="activites_non_programmees")
            bd_modifier_cellule(index_df, "Date", int(jour_choisi))
            demander_selection("creneaux_disponibles", get_creneau_proche(st.session_state.get("creneaux_disponibles"), df.loc[index_df])[0])
            forcer_reaffichage_df("creneaux_disponibles")
            sql.sauvegarder_row(index_df)
            st.rerun()

    # Affichage Liste des jours possibles
    jours_label = [f"{int(jour):02d}" for jour in jours_possibles]
    if jours_label and (not st.session_state.get("menu_activite_choix_jour_programmation") or st.session_state.menu_activite_choix_jour_programmation not in jours_label):
            st.session_state.menu_activite_choix_jour_programmation = jours_label[0]
    st.session_state.activites_non_programmees_jour_choisi = st.selectbox("Jours possibles", jours_label, label_visibility="visible", disabled=boutons_disabled or not jours_possibles, key = "menu_activite_choix_jour_programmation") # , width=90
        
    # Affichage de l'éditeur d'activité
    if st.button(LABEL_BOUTON_EDITER, use_container_width=CENTRER_BOUTONS, disabled=boutons_disabled,  key="menu_activite_bouton_editer"):
        if "editeur_activite_etat" in st.session_state:
            del st.session_state["editeur_activite_etat"]
        show_dialog_editeur_activite(df, index_df)

    # Affichage contrôle Ajouter
    afficher_bouton_nouvelle_activite(key="menu_activite_bouton_nouvelle_activite")

# Affichage de l'éditeur d'activité en mode modal
@st.dialog("Editeur d'activité")
def show_dialog_editeur_activite(df, index_df):
    afficher_nom_activite(df, index_df, afficher_label=False)
    afficher_editeur_activite(df, index_df)

# Affichage de l'éditeur d'activité
def afficher_editeur_activite(df, index_df=None, key="editeur_activite"):

    def enregistrer_modification_dans_row(df, row, colonne_df, valeur_courante, nouvelle_valeur):

        erreur = None
        if (pd.isna(valeur_courante) and (pd.notna(nouvelle_valeur) and str(nouvelle_valeur) != "")) or (pd.notna(valeur_courante) and str(valeur_courante) != str(nouvelle_valeur)):
        
            erreur = affecter_valeur_row(row, colonne_df, nouvelle_valeur)

            if erreur is not None:
                st.error(erreur)
            else:
                try:
                    if  colonne_df != "Lien Web":
                        if ptypes.is_numeric_dtype(df[colonne_df]) and not ptypes.is_numeric_dtype(row[colonne_df]):
                            if "." not in row[colonne_df] and "," not in row[colonne_df] and "e" not in row[colonne_df].lower():
                                row[colonne_df] = int(row[colonne_df])
                            else:
                                row[colonne_df] = float(row[colonne_df])
                except Exception as e:
                    erreur = f"⛔ Format numérique attendu pour cette colonne"
                    st.error(erreur)
                if (pd.isna(nouvelle_valeur) and not pd.isna(valeur_courante)) or (not pd.isna(nouvelle_valeur) and pd.isna(valeur_courante)) or nouvelle_valeur != valeur_courante:
                    if colonne_df == "Lien Web":
                        st.session_state.editeur_activite_etat["lien_modif"] = True
                    else:
                        st.session_state.editeur_activite_etat["col_modif"].append(colonne_df)
                        if est_activite_programmee(row):
                            # st.session_state.editeur_activite_etat["forcer_reaffichage_activites_non_programmees"] = True -> remplacé par __df_push_ver
                            if colonne_df in ["Debut", "Duree", "Activité"]:
                                st.session_state.editeur_activite_etat["forcer_reaffichage_creneaux_disponibles"] = True
                        elif est_activite_non_programmee(row):
                            # st.session_state.editeur_activite_etat["forcer_reaffichage_activites_non_programmees"] = True -> remplacé par __df_push_ver
                            st.session_state.editeur_activite_etat["forcer_reaffichage_activites_programmables"] = True
        return erreur
                
    # Rien à faire sur df vide
    if len(df) <= 0:
        return
    
    if index_df is None:
        if "editeur_activite_idx" in st.session_state:
            index_df = st.session_state.editeur_activite_idx 
    
    if index_df is not None:

        input_text_key_counter_key = key + "input_text_key_counter"
        st.session_state.setdefault(input_text_key_counter_key, 0)
        input_text_key_counter = st.session_state.get(input_text_key_counter_key)
        if "editeur_activite_etat" not in st.session_state:
            input_text_key_counter += 1
            st.session_state[input_text_key_counter_key] = input_text_key_counter

        st.session_state.setdefault("editeur_activite_etat", {
            "row": df.loc[index_df].copy(),
            "colonne_courante": None,
            "nouvelle_valeur": None,
            "col_modif": [],
            "forcer_reaffichage_activites_programmees": False,
            "forcer_reaffichage_activites_non_programmees": False,
            "forcer_reaffichage_creneaux_disponibles": False,
            "forcer_reaffichage_activites_programmables": False,
            "erreur": None,
        })

        row = st.session_state.editeur_activite_etat["row"]

        if est_activite_reserve(row):
            colonnes_editables = [col for col in df.columns if col not in ["Date", "Fin", "Debut_dt", "Duree_dt", "Debut", "Duree", "__uuid"]]
        else:
            colonnes_editables = [col for col in df.columns if col not in ["Date", "Fin", "Debut_dt", "Duree_dt", "__uuid"]]

        # Traitement de l'accentuation
        colonnes_editables_avec_accents = [RENOMMAGE_COLONNES.get(col, col) for col in colonnes_editables]
        
        colonne = st.selectbox("⚙️ Colonne", colonnes_editables_avec_accents, key=key+"_selectbox_choix_colonne")
        colonne_df = RENOMMAGE_COLONNES_INVERSE[colonne] if colonne in RENOMMAGE_COLONNES_INVERSE else colonne

        colonne_rerun_pred = st.session_state.editeur_activite_etat.get("colonne_courante")
        if colonne_rerun_pred is None or  colonne_rerun_pred != colonne_df:
            st.session_state.editeur_activite_etat["colonne_courante"] = colonne_df


        valeur_courante = row[colonne_df]

        st.session_state.editeur_activite_etat["nouvelle_valeur"] = st.text_input(f"✏️ Valeur", "" if pd.isna(valeur_courante) else str(valeur_courante), key=key+str(input_text_key_counter)) 
        erreur = enregistrer_modification_dans_row(df, row, colonne_df, row[colonne_df], st.session_state.editeur_activite_etat.get("nouvelle_valeur"))

        if st.button(LABEL_BOUTON_VALIDER, use_container_width=CENTRER_BOUTONS):
            if not erreur and st.session_state.editeur_activite_etat["col_modif"]:
                undo.save()
                try:
                    if st.session_state.editeur_activite_etat["col_modif"]:
                        cols = {}
                        for col in st.session_state.editeur_activite_etat["col_modif"]:
                            cols[col] = row[col]
                            bd_modifier_cellule(index_df, col, row[col])

                        if st.session_state.editeur_activite_etat["forcer_reaffichage_activites_programmees"]:
                            forcer_reaffichage_activites_programmees()
                        if st.session_state.editeur_activite_etat["forcer_reaffichage_activites_non_programmees"]:
                            forcer_reaffichage_activites_non_programmees()
                            
                        if st.session_state.editeur_activite_etat["forcer_reaffichage_creneaux_disponibles"]:
                            forcer_reaffichage_df("creneaux_disponibles")
                        if st.session_state.editeur_activite_etat["forcer_reaffichage_activites_programmables"]:
                            forcer_reaffichage_df("activites_programmables")

                        # Mise en attente du code de traitement des cellValueChanged utilisateur tant que le DOM n'a pas pris en compte les modifs
                        row_modification_request_set(index_df, cols)
                        if est_activite_programmee(row): signaler_df_push("activites_programmees")
                        if est_activite_non_programmee(row): signaler_df_push("activites_non_programmees")
                        
                        sql.sauvegarder_row(index_df)

                    st.rerun()
                except Exception as e:
                    st.error(f"⛔ {e}")
                    undo.undo()

            else:
                st.rerun()
                    
def valider_valeur(df, colonne, nouvelle_valeur):           
    erreur = None
    if colonne == "Debut" and not est_heure_valide(nouvelle_valeur):
        erreur = "⛔ Format attendu : HHhMM (ex : 10h00)"
    elif colonne == "Duree" and not est_duree_valide(nouvelle_valeur):
        erreur = "⛔ Format attendu : HhMM (ex : 1h00 ou 0h30)"
    elif colonne == "Relache" and not est_relache_valide(nouvelle_valeur):
        erreur = "⛔ Format attendu : 1, 10, pair, impair"
    elif colonne == "Reserve" and not est_reserve_valide(nouvelle_valeur):
        erreur = "⛔ Format attendu : Oui, Non"
    elif ptypes.is_numeric_dtype(df[colonne]) and not ptypes.is_numeric_dtype(nouvelle_valeur):
        try:
            if "." not in nouvelle_valeur and "," not in nouvelle_valeur and "e" not in nouvelle_valeur.lower():
                nouvelle_valeur = int(nouvelle_valeur)
            else:
                nouvelle_valeur = float(nouvelle_valeur)
        except:
            erreur = "⛔ Format numérique attendu"
    return erreur

# Affecte une nouvelle valeur à une cellule du df de base donnée par son index et sa colonne
def affecter_valeur_df(index, colonne, nouvelle_valeur, section_critique=None):
    
    def set_section_critique_step(section_critique, step):
        if section_critique is not None:
            section_critique["step"] = step

    df = st.session_state.df
    valeur_courante = df.at[index, colonne]
    step = section_critique["step"] if section_critique is not None else 0
    tracer.log(f"step {step}")
    erreur = None

    if step == 0:
        erreur = valider_valeur(df, colonne, nouvelle_valeur)
        if not erreur:
            set_section_critique_step(section_critique, 1)
            if colonne == "Debut" :
                heures, minutes = nouvelle_valeur.split("h")
                nouvelle_valeur = f"{int(heures):02d}h{int(minutes):02d}"
            if (pd.isna(valeur_courante) and pd.notna(nouvelle_valeur)) or nouvelle_valeur != valeur_courante:
                try:
                    df.at[index, colonne] = nouvelle_valeur
                except Exception as e:
                    erreur = f"⛔ {e}"
                else:
                    set_section_critique_step(section_critique, 2)
                    df.at[index, colonne] = valeur_courante
                    undo.save()
                    bd_modifier_cellule(index, colonne, nouvelle_valeur)
                    sql.sauvegarder_row(index)
    elif step == 1:
        if colonne == "Debut" :
            heures, minutes = nouvelle_valeur.split("h")
            nouvelle_valeur = f"{int(heures):02d}h{int(minutes):02d}"
        if (pd.isna(valeur_courante) and pd.notna(nouvelle_valeur)) or nouvelle_valeur != valeur_courante:
            try:
                df.at[index, colonne] = nouvelle_valeur
            except Exception as e:
                erreur = f"⛔ {e}"
            else:
                set_section_critique_step(section_critique, 2)
                df.at[index, colonne] = valeur_courante
                undo.save()
                bd_modifier_cellule(index, colonne, nouvelle_valeur, section_critique=section_critique)
                sql.sauvegarder_row(index)
    elif step == 2:
        df.at[index, colonne] = valeur_courante
        undo.save()
        bd_modifier_cellule(index, colonne, nouvelle_valeur, section_critique=section_critique)
        sql.sauvegarder_row(index)
        
    return erreur

# Affecte une nouvelle valeur à une cellule d'une row d'un df 
def affecter_valeur_row(row, colonne, nouvelle_valeur):
    valeur_courante = row[colonne]
    erreur = valider_valeur(row, colonne, nouvelle_valeur)
    if not erreur:
        if colonne == "Debut" :
            heures, minutes = nouvelle_valeur.split("h")
            nouvelle_valeur = f"{int(heures):02d}h{int(minutes):02d}"
        if (pd.isna(valeur_courante) and pd.notna(nouvelle_valeur)) or nouvelle_valeur != valeur_courante:
            try:
                row[colonne] = nouvelle_valeur
            except Exception as e:
                erreur = f"⛔ {e}"

    return erreur

# Vérifie qu'une valeur est bien Oui Non
def est_reserve_valide(val):
    return str(val).strip().lower() in ["oui", "non", ""]

# Vérifie qu'une valeur contient bien NaN ou "" ou quelque chose du type "1", "1,10", "1, 10", "1, pair", "12, impair"
def est_relache_valide(val):

    # Cas val vide ou NaN
    if pd.isna(val) or str(val).strip() == "":
        return True

    val_str = str(val).strip().lower()

    # Autorise : chiffres ou mots-clés (pair, impair) séparés par virgules
    # Exemples valides : "1", "1, 10", "1, impair", "2, pair"
    # Regex : liste d'éléments séparés par des virgules, chaque élément est un entier ou 'pair'/'impair'
    motif = r"^\s*(\d+|pair|impair)(\s*,\s*(\d+|pair|impair))*\s*$"

    return re.fullmatch(motif, val_str) is not None

# Vérifie si une date de référence est compatible avec la valeur de la colonne Relache qui donne les jours de relache pour une activité donnée
def est_hors_relache(relache_val, date_val):
    if pd.isna(relache_val) or pd.isna(date_val):
        return True  # Aucune relâche spécifiée ou date absente

    if not est_relache_valide(relache_val):
        return True
    
    try:
        date_int = int(float(date_val))
    except (ValueError, TypeError):
        return True  # Si la date n'est pas exploitable, on la considère programmable

    # Normaliser le champ Relache en chaîne
    if isinstance(relache_val, (int, float)):
        relache_str = str(int(relache_val))
    else:
        relache_str = str(relache_val).strip().lower()

    # Cas particulier : pair / impair
    if "pair" in relache_str and date_int % 2 == 0:
        return False
    if "impair" in relache_str and date_int % 2 != 0:
        return False

    # Cas général : liste explicite de jours (ex : "20,21")
    try:
        jours = [int(float(x.strip())) for x in relache_str.split(",")]
        if date_int in jours:
            return False
    except ValueError:
        pass  # ignorer s'il ne s'agit pas d'une liste de jours

    return True

# Supprime une row dans un df à partir de son index
def supprimer_row_df(df, idx):
    return df.drop(idx) if idx in df.index else df

# Supprime une row dans un df_display d'AgGrid à partir de son index dans le df principal (suppose que cet index est stocké dans la colonne __index du df_display)
def supprimer_row_df_display(df, idx):
    matches = df[df["__index"].astype(str) == str(idx)]
    return df.drop(matches.index) if not matches.empty else df

# Suppression d'une activité d'un df
def supprimer_activite(idx):
    if idx not in st.session_state.df.index:
        return
    jour = st.session_state.df.loc[idx]["Date"]
    uuid = st.session_state.df.loc[idx]["__uuid"]
    st.session_state.df.loc[idx] = pd.NA
    st.session_state.df.at[idx, "__uuid"] = uuid
    st.session_state.activites_programmees = supprimer_row_df(st.session_state.activites_programmees, idx)
    st.session_state.activites_non_programmees = supprimer_row_df(st.session_state.activites_non_programmees, idx)
    st.session_state.activites_programmees_df_display = supprimer_row_df_display(st.session_state.activites_programmees_df_display, idx)
    st.session_state.activites_programmees_df_display_copy = supprimer_row_df_display(st.session_state.activites_programmees_df_display_copy, idx)
    st.session_state.activites_non_programmees_df_display = supprimer_row_df_display(st.session_state.activites_non_programmees_df_display, idx)
    st.session_state.activites_non_programmees_df_display_copy = supprimer_row_df_display(st.session_state.activites_non_programmees_df_display_copy, idx)
    maj_options_date(st.session_state.df, st.session_state.activites_programmees, st.session_state.activites_programmees_df_display, jour)
    maj_options_date(st.session_state.df, st.session_state.activites_programmees, st.session_state.activites_non_programmees_df_display, jour)
    bd_maj_creneaux_disponibles()

# Modifie la valeur d'une cellule d'un df
def modifier_df_cell(df, idx, col, val):
    if idx in df.index:
        df.at[idx, col] = val

# Modifie la valeur d'une cellule d'un df_display
def modifier_df_display_cell(df, idx, col, val):
    matches = df[df["__index"].astype(str) == str(idx)]
    if not matches.empty:
        df.at[matches.index[0], col] = val

# Renvoie le nom d'une colonne en faisant la traduction colonne df -> colonne df_display
def df_display_col_nom(nom):
    return RENOMMAGE_COLONNES.get(nom, nom)

# Création de la liste des créneaux avant/après pour chaque activité programmée 
# le df des activités programmées est supposé etre trié par jour ("Date") et par heure de début ("Debut")
def get_creneaux(df, activites_programmees, traiter_pauses):

    def creer_creneau(row, borne_min, borne_max, avant, apres, type_creneau):
        titre = row["Activite"] if not pd.isna(row["Activite"]) else ""
        date_str = str(int(row["Date"])) if pd.notnull(row["Date"]) else ""
        return {
            "Date": date_str, # str pour ne pas avoir d'icone de filtre sur la colonne
            "Debut": borne_min.strftime('%Hh%M'),
            "Fin": borne_max.strftime('%Hh%M'),
            "Activité avant": avant,
            "Activité après": apres,
            "__type_creneau": type_creneau,
            "__index": row.name,
            "__uuid": str(uuid.uuid4())
        }
    
    params_to_hash = [
        traiter_pauses, 
        st.session_state.get("MARGE", MARGE).total_seconds(), 
        st.session_state.get("DUREE_REPAS", DUREE_REPAS).total_seconds(), 
        st.session_state.get("DUREE_CAFE", DUREE_CAFE).total_seconds(),
        st.session_state.get("periode_a_programmer_debut", BASE_DATE).isoformat(),
        st.session_state.get("periode_a_programmer_fin", BASE_DATE).isoformat(),
    ]

    hash_val  = hash_df(df, colonnes_a_garder=[col for col in df.columns if col not in ["Debut_dt", "Duree_dt", "__uuid"]], params=params_to_hash)
    hash_key = "creneaux__hash"
    key = "creneaux"
    
    if st.session_state.get(hash_key) != hash_val:
        
        creneaux = []
        bornes = []

        # Traitement des jours libres 
        jours_libres = []
        for jour in range(st.session_state.periode_a_programmer_debut.day, st.session_state.periode_a_programmer_fin.day + 1):
            if jour not in activites_programmees["Date"].values:
                jours_libres.append(jour)
        for jour in jours_libres:
            if exist_activites_programmables(jour, traiter_pauses):
                row = pd.Series({col: None for col in df.columns})
                row["Date"] = jour
                borne_min = datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))
                borne_max = datetime.datetime.combine(BASE_DATE, datetime.time(23, 59))
                creneaux.append(creer_creneau(row, borne_min, borne_max, "", "", "Journée"))

        if len(activites_programmees) > 0:
            # Initialisation de jour_courant au premier jour des activités programmées
            jour_courant = activites_programmees.iloc[0]["Date"]

            for _, row in activites_programmees.iterrows():

                # Heure de début d'activité
                heure_debut = row["Debut_dt"]
                # Heure de fin d'activité
                heure_fin = heure_debut + row["Duree_dt"] if pd.notnull(heure_debut) and pd.notnull(row["Duree_dt"]) else None
                # initialisation du tableau enregistrant pour chaque jour les bornes des creneaux rencontrés pour eviter les doublons
                if row ["Date"] != jour_courant:
                    bornes = []
                    jour_courant = row ["Date"]

                # Ajout des creneaux avant l'activité considérée s'ils existent
                if pd.notnull(heure_debut):
                    if get_activites_programmables_avant(df, activites_programmees, row, traiter_pauses):
                        borne_min, borne_max, pred = get_creneau_bounds_avant(activites_programmees, row)
                        if (borne_min, borne_max) not in bornes:
                            bornes.append((borne_min, borne_max))
                            creneaux.append(creer_creneau(row, borne_min, borne_max, pred["Activite"] if pred is not None else "", row["Activite"], "Avant"))

                # Ajout des creneaux après l'activité considérée s'ils existent
                if pd.notnull(heure_fin):
                    if get_activites_programmables_apres(df, activites_programmees, row, traiter_pauses):
                        borne_min, borne_max, next = get_creneau_bounds_apres(activites_programmees, row)
                        borne_max = borne_max if borne_max is not None else datetime.datetime.combine(BASE_DATE, datetime.time(23, 59))
                        if (borne_min, borne_max) not in bornes:
                            bornes.append((borne_min, borne_max))
                            creneaux.append(creer_creneau(row, borne_min, borne_max, row["Activite"], next["Activite"] if next is not None else "", "Après"))
        creneaux = sorted(creneaux, key=lambda x: int(x["Date"]))
        creneaux = pd.DataFrame(creneaux)
        st.session_state[key] = creneaux
        st.session_state[hash_key] = hash_val
    return st.session_state[key]

# Renvoie les bornes du créneau existant avant une activité donnée par son descripteur ligne_ref
def get_creneau_bounds_avant(activites_programmees, ligne_ref):
    date_ref = ligne_ref["Date"]
    debut_ref = ligne_ref["Debut_dt"] if pd.notnull(ligne_ref["Debut_dt"]) else datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))
    duree_ref = ligne_ref["Duree_dt"] if pd.notnull(ligne_ref["Duree_dt"]) else datetime.timedelta(0)
    fin_ref = debut_ref + duree_ref if pd.notnull(debut_ref) and pd.notnull(duree_ref) else None    

    # Chercher l'activité programmée précédente sur le même jour
    programmes_jour_ref = activites_programmees[activites_programmees["Date"] == date_ref]
    programmes_jour_ref = programmes_jour_ref.sort_values(by="Debut_dt")
    prev = programmes_jour_ref[programmes_jour_ref["Debut_dt"] < debut_ref].tail(1)

    # Calculer l'heure de début minimum du créneau
    if not prev.empty:
        prev_fin = datetime.datetime.combine(BASE_DATE, prev["Debut_dt"].iloc[0].time()) + prev["Duree_dt"].iloc[0]
        debut_min = prev_fin
    else:
        debut_min = datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))

    # Calculer l'heure de fin max du créneau
    fin_max = datetime.datetime.combine(BASE_DATE, debut_ref.time())

    return debut_min, fin_max, prev.iloc[0] if not prev.empty else None

# Renvoie les bornes du créneau existant après une activité donnée par son descripteur ligne_ref
# S'il n'y a pas d'activité suivante pour le même jour renvoie None pour fin_max
def get_creneau_bounds_apres(activites_programmees, ligne_ref):
    date_ref = ligne_ref["Date"]
    debut_ref = ligne_ref["Debut_dt"] if pd.notnull(ligne_ref["Debut_dt"]) else datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))
    duree_ref = ligne_ref["Duree_dt"] if pd.notnull(ligne_ref["Duree_dt"]) else datetime.timedelta(0)
    fin_ref = debut_ref + duree_ref if pd.notnull(debut_ref) and pd.notnull(duree_ref) else debut_ref    


    # Ajuster la date de référence si le jour a changé
    if fin_ref.day != debut_ref.day:
        date_ref = date_ref + fin_ref.day - debut_ref.day  

    # Chercher l'activité programmée suivante sur le même jour de référence
    programmes_jour_ref = activites_programmees[activites_programmees["Date"] == date_ref]
    programmes_jour_ref = programmes_jour_ref.sort_values(by="Debut_dt")
    next = programmes_jour_ref[programmes_jour_ref["Debut_dt"] + programmes_jour_ref["Duree_dt"] > fin_ref].head(1)

    # Calculer l'heure de fin max du créneau
    if not next.empty:
        fin_max = datetime.datetime.combine(BASE_DATE, next["Debut_dt"].iloc[0].time())
    else:
        fin_max = None # datetime.datetime.combine(BASE_DATE, datetime.time(23, 59))

    # Calculer l'heure de début minimum du créneau
    debut_min = datetime.datetime.combine(BASE_DATE, fin_ref.time())

    return debut_min, fin_max, next.iloc[0] if not next.empty else None

# Renvoie la liste des activités programmables avant une activité donnée par son descripteur ligne_ref
def get_activites_programmables_avant(df, activites_programmees, ligne_ref, traiter_pauses=True):
    date_ref = ligne_ref["Date"]
    debut_ref = ligne_ref["Debut_dt"] if pd.notnull(ligne_ref["Debut_dt"]) else datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))
    duree_ref = ligne_ref["Duree_dt"] if pd.notnull(ligne_ref["Duree_dt"]) else datetime.timedelta(0)
    fin_ref = debut_ref + duree_ref if pd.notnull(debut_ref) and pd.notnull(duree_ref) else None

    proposables = [] 

    debut_min, fin_max, _ = get_creneau_bounds_avant(activites_programmees, ligne_ref)
    if debut_min >= fin_max:
        return proposables  # Pas d'activités programmables avant si le créneau est invalide

    for _, row in df[df["Date"].isna()].iterrows():
        if pd.isna(row["Debut_dt"]) or pd.isna(row["Duree_dt"]):
            continue
        h_debut = datetime.datetime.combine(BASE_DATE, row["Debut_dt"].time())
        h_fin = h_debut + row["Duree_dt"]
        # L'activité doit commencer après debut_min et finir avant fin_max
        if h_debut >= debut_min + st.session_state.MARGE and h_fin <= fin_max - st.session_state.MARGE and est_hors_relache(row["Relache"], date_ref):
            nouvelle_ligne = row.drop(labels=["Debut_dt", "Duree_dt"]).to_dict()
            nouvelle_ligne["__type_activite"] = "ActiviteExistante"
            nouvelle_ligne["__index"] = row.name
            proposables.append(nouvelle_ligne)
    if traiter_pauses:
        ajouter_pauses(proposables, activites_programmees, ligne_ref, "Avant")
    return proposables

# Renvoie la liste des activités programmables après une activité donnée par son descripteur ligne_ref
def get_activites_programmables_apres(df, activites_programmees, ligne_ref, traiter_pauses=True):
    date_ref = ligne_ref["Date"]
    debut_ref = ligne_ref["Debut_dt"] if pd.notnull(ligne_ref["Debut_dt"]) else datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))
    duree_ref = ligne_ref["Duree_dt"] if pd.notnull(ligne_ref["Duree_dt"]) else datetime.timedelta(0)
    fin_ref = debut_ref + duree_ref if pd.notnull(debut_ref) and pd.notnull(duree_ref) else None   

    proposables = []

    debut_min, fin_max, _ = get_creneau_bounds_apres(activites_programmees, ligne_ref) # Attention fin_max est None si créneau se termine apres 23h59
    if fin_max is not None and debut_min >= fin_max:
        return proposables  # Pas d'activités programmables avant si le créneau est invalide

    if fin_ref.day != debut_ref.day:
        return proposables  # Pas d'activités programmables après si le jour a changé

    for _, row in df[df["Date"].isna()].iterrows():
        if pd.isna(row["Debut_dt"]) or pd.isna(row["Duree_dt"]):
            continue
        h_debut = datetime.datetime.combine(BASE_DATE, row["Debut_dt"].time())
        h_fin = h_debut + row["Duree_dt"]
        # L'activité doit commencer après debut_min et finir avant fin_max en tenant compte des marges et des relaches
        if h_debut >= debut_min + st.session_state.MARGE and (fin_max is None or h_fin <= fin_max - st.session_state.MARGE) and est_hors_relache(row["Relache"], date_ref):
            nouvelle_ligne = row.drop(labels=["Debut_dt", "Duree_dt"]).to_dict()
            nouvelle_ligne["__type_activite"] = "ActiviteExistante"
            nouvelle_ligne["__index"] = row.name
            proposables.append(nouvelle_ligne)
    if traiter_pauses:
        ajouter_pauses(proposables, activites_programmees, ligne_ref, "Après")
    return proposables

# Renvoie les activités programmables sur une journée entière donc les activités qui ne sont pas relache ce jour
def get_activites_programmables_sur_journee_entiere(date_ref, traiter_pauses=True):
    proposables = []

    for _, row in st.session_state.activites_non_programmees.iterrows():
        if est_hors_relache(row["Relache"], date_ref):
            nouvelle_ligne = row.drop(labels=["Debut_dt", "Duree_dt"]).to_dict()
            nouvelle_ligne["__type_activite"] = "ActiviteExistante"
            nouvelle_ligne["__index"] = row.name
            proposables.append(nouvelle_ligne)
    
    if traiter_pauses:
        h_dej = datetime.datetime.combine(BASE_DATE, datetime.time(12, 0))
        type_repas = "déjeuner"
        proposables.append(
            completer_ligne({
                "Debut": datetime.datetime.combine(BASE_DATE, datetime.time(12, 0)).strftime('%Hh%M'),
                "Fin": (h_dej + st.session_state.DUREE_REPAS).strftime('%Hh%M'),
                "Duree": duree_str(st.session_state.DUREE_REPAS),
                "Activite": f"Pause {type_repas}",
                "__type_activite": type_repas,
                "__uuid": str(uuid.uuid4()),
            }))
        
        h_dej = datetime.datetime.combine(BASE_DATE, datetime.time(12, 0))
        type_repas = "dîner"
        proposables.append(
            completer_ligne({
                "Debut": datetime.datetime.combine(BASE_DATE, datetime.time(20, 0)).strftime('%Hh%M'),
                "Fin": (h_dej + st.session_state.DUREE_REPAS).strftime('%Hh%M'),
                "Duree": duree_str(st.session_state.DUREE_REPAS),
                "Activite": f"Pause {type_repas}",
                "__type_activite": type_repas,
                "__uuid": str(uuid.uuid4()),
            }))
    
    return proposables

# Renvoie True s'il existe des activités programmables sur une journée entière donc des activités qui ne sont pas relache ce jour
def exist_activites_programmables(date_ref, traiter_pauses=False):
    if traiter_pauses:
        return True
    for _, row in st.session_state.activites_non_programmees.iterrows():
        if est_hors_relache(row["Relache"], date_ref):
            return True
    return False

# Vérifie si une pause d'un type donné est déjà présente pour un jour donné dans le dataframe des activités planiées
def pause_deja_existante(activites_programmees, jour, type_pause):
    activites_programmes_du_jour = activites_programmees[activites_programmees["Date"] == jour]
    return activites_programmes_du_jour["Activite"].astype(str).str.contains(type_pause, case=False, na=False).any() 

# Ajoute les pauses possibles (déjeuner, dîner, café) à une liste d'activités programmables pour une activité donnée par son descripteur ligne_ref
def ajouter_pauses(proposables, activites_programmees, ligne_ref, type_creneau):

    # Pause repas
    def ajouter_pause_repas(proposables, date_ref, debut_min, fin_max, pause_debut_min, pause_debut_max, type_repas):
        if not pause_deja_existante(activites_programmees, date_ref, type_repas):
            if type_creneau == "Avant":
                h_dej = min(max(fin_max - st.session_state.DUREE_REPAS - st.session_state.MARGE, 
                    datetime.datetime.combine(BASE_DATE, pause_debut_min)), 
                    datetime.datetime.combine(BASE_DATE, pause_debut_max))
                if h_dej - st.session_state.MARGE >= debut_min and h_dej + st.session_state.MARGE <= fin_max:
                    nouvelle_ligne = completer_ligne({
                        "Debut": h_dej.strftime('%Hh%M'),
                        "Fin": (h_dej + st.session_state.DUREE_REPAS).strftime('%Hh%M'),
                        "Duree": duree_str(st.session_state.DUREE_REPAS),
                        "Activite": f"Pause {type_repas}",
                        "__type_activite": type_repas,
                        "__uuid": str(uuid.uuid4()),
                    })
                    proposables.append(nouvelle_ligne)
            elif type_creneau == "Après": # Attention : dans ce cas fin_max est None si le créneau se termine apres 23h59
                h_dej = min(max(debut_min + st.session_state.MARGE, 
                    datetime.datetime.combine(BASE_DATE, pause_debut_min)), 
                    datetime.datetime.combine(BASE_DATE, pause_debut_max))
                if h_dej - st.session_state.MARGE >= debut_min and (fin_max is None or h_dej + st.session_state.MARGE <= fin_max):
                    nouvelle_ligne = completer_ligne({
                        "Debut": h_dej.strftime('%Hh%M'),
                        "Fin": (h_dej + st.session_state.DUREE_REPAS).strftime('%Hh%M'),
                        "Duree": duree_str(st.session_state.DUREE_REPAS),
                        "Activite": f"Pause {type_repas}",
                        "__type_activite": type_repas,
                        "__uuid": str(uuid.uuid4()),
                    })
                    proposables.append(nouvelle_ligne)
    
    def ajouter_pause_cafe(proposables, debut_min, fin_max):
        if not est_pause(ligne_ref):
            Lieu_ref = ligne_ref["Lieu"]
            if type_creneau == "Avant":
                i = activites_programmees.index.get_loc(ligne_ref.name)  
                Lieu_ref_prev = activites_programmees.iloc[i - 1]["Lieu"] if i > 0 else None
                h_cafe = fin_max - st.session_state.DUREE_CAFE
                if not pd.isna(Lieu_ref) and not pd.isna(Lieu_ref_prev) and Lieu_ref == Lieu_ref_prev: 
                    # Dans ce cas pas la peine de tenir compte de la marge avec l'activité précédente
                    if h_cafe >= debut_min: 
                        nouvelle_ligne = completer_ligne({
                            "Debut": h_cafe.strftime('%Hh%M'),
                            "Fin": (h_cafe + st.session_state.DUREE_CAFE).strftime('%Hh%M'),
                            "Duree": duree_str(st.session_state.DUREE_CAFE),
                            "Activite": f"Pause café",
                            "__type_activite": "café",
                            "__uuid": str(uuid.uuid4()),
                        })
                        proposables.append(nouvelle_ligne)
                else: 
                    # Dans ce cas on tient compte de la marge avec l'activité précédente sauf si debut_min = 0h00
                    marge_cafe = st.session_state.MARGE if debut_min != datetime.datetime.combine(BASE_DATE, datetime.time(0, 0)) else datetime.timedelta(minutes=0) 
                    if h_cafe >= debut_min + marge_cafe:
                        nouvelle_ligne = completer_ligne({
                            "Debut": h_cafe.strftime('%Hh%M'),
                            "Fin": (h_cafe + st.session_state.DUREE_CAFE).strftime('%Hh%M'),
                            "Duree": duree_str(st.session_state.DUREE_CAFE),
                            "Activite": "Pause café",
                            "__type_activite": "café",
                            "__uuid": str(uuid.uuid4()),
                        })
                        proposables.append(nouvelle_ligne)
            elif type_creneau == "Après": # Attention : dans ce cas fin_max est None si le créneau se termine apres 23h59
                i = activites_programmees.index.get_loc(ligne_ref.name)  
                Lieu_ref_suiv = activites_programmees.iloc[i + 1]["Lieu"] if i < len(activites_programmees) - 1 else None
                h_cafe = debut_min
                if not pd.isna(Lieu_ref) and not pd.isna(Lieu_ref_suiv) and Lieu_ref == Lieu_ref_suiv: 
                    # Dans ce cas pas la peine de tenir compte de la marge avec l'activité suivante 
                    if fin_max is None or h_cafe + st.session_state.DUREE_CAFE <= fin_max: 
                        nouvelle_ligne = completer_ligne({
                            "Debut": h_cafe.strftime('%Hh%M'),
                            "Fin": (h_cafe + st.session_state.DUREE_CAFE).strftime('%Hh%M'),
                            "Duree": duree_str(st.session_state.DUREE_CAFE),
                            "Activite": "Pause café",
                            "__type_activite": "café",
                            "__uuid": str(uuid.uuid4()),
                        })
                        proposables.append(nouvelle_ligne)
                else: 
                    # Dans ce cas on tient compte de la marge avec l'activité suivante sauf si fin_max is None (créneau se termine après 23h59)
                    marge_cafe = st.session_state.MARGE if fin_max is not None else datetime.timedelta(minutes=0)
                    if fin_max is None or h_cafe + st.session_state.DUREE_CAFE <= fin_max - marge_cafe:
                        nouvelle_ligne = completer_ligne({
                            "Debut": h_cafe.strftime('%Hh%M'),
                            "Fin": (h_cafe + st.session_state.DUREE_CAFE).strftime('%Hh%M'),
                            "Duree": duree_str(st.session_state.DUREE_CAFE),
                            "Activite": "Pause café",
                            "__type_activite": "café",
                            "__uuid": str(uuid.uuid4()),
                        })
                        proposables.append(nouvelle_ligne)

    date_ref = ligne_ref["Date"]
    debut_ref = ligne_ref["Debut_dt"] if pd.notnull(ligne_ref["Debut_dt"]) else datetime.datetime.combine(BASE_DATE, datetime.time(0, 0))
    duree_ref = ligne_ref["Duree_dt"] if pd.notnull(ligne_ref["Duree_dt"]) else datetime.timedelta(0)
    fin_ref = debut_ref + duree_ref if pd.notnull(debut_ref) and pd.notnull(duree_ref) else None    

    def desc(h, duree, nom):
        # return f"{int(date_ref)} de {h.strftime('%Hh%M')} à {(h + duree).time().strftime('%Hh%M')} ({formatter_timedelta(duree)}) - {nom}"
        return f"{int(date_ref)} - {h.strftime('%Hh%M')} - {nom}"
    
    # Récupération des bornes du créneau
    if type_creneau == "Avant":
        debut_min, fin_max, _ = get_creneau_bounds_avant(activites_programmees, ligne_ref)
    elif type_creneau == "Après":
        debut_min, fin_max, _ = get_creneau_bounds_apres(activites_programmees, ligne_ref)
    else:
        raise ValueError("type_creneau doit être 'Avant' ou 'Après'")

    # Pause déjeuner
    ajouter_pause_repas(proposables, date_ref, debut_min, fin_max, PAUSE_DEJ_DEBUT_MIN, PAUSE_DEJ_DEBUT_MAX, "déjeuner")

    # Pause dîner
    ajouter_pause_repas(proposables, date_ref, debut_min, fin_max, PAUSE_DIN_DEBUT_MIN, PAUSE_DIN_DEBUT_MAX, "dîner")

    # Pause café
    ajouter_pause_cafe(proposables, debut_min, fin_max)

def est_nom_pause(val):
    valeurs = val.split()
    if not valeurs:
        return False
    return val.split()[0].lower() == "pause"

def est_pause(ligne_ref):
    val = str(ligne_ref["Activite"]).strip()
    return est_nom_pause(val)

def est_pause_cafe(ligne_ref):
    if not est_pause(ligne_ref):
        return False
    val = str(ligne_ref["Activite"]).strip()
    valeurs = val.split()
    if not valeurs:
        return False
    if len(valeurs) < 2:
        return False
    return val.split()[0].lower() == "pause" and val.split()[1].lower() == "café"

def sauvegarder_contexte(df_hash=None):

    def serialiser_contexte(df):
        # Réindexer proprement pour éviter les trous
        df_sorted = df.copy()
        df_sorted = df_sorted.sort_values(by=["Date", "Debut_dt"])
        df_sorted = df_sorted.reset_index(drop=True)
        df_sorted = df_sorted.drop(columns=["Debut_dt", "Duree_dt", "__uuid"], errors='ignore')

        # Récupération de la worksheet à traiter
        wb = st.session_state.get("wb")

        if wb is not None:
            ws = wb.worksheets[0]

            # Effacer le contenu de la feuille Excel existante
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.value = None  # on garde le style, on efface juste la valeur
                    cell.hyperlink = None

            # Réinjecter les données du df dans la feuille Excel
            from copy import copy

            col_activite = None
            for cell in ws[1]:
                if cell.value and str(cell.value).strip().lower() in ["activité"]:
                    col_activite = cell.column
            source_font = ws.cell(row=1, column=1).font

            # Réécriture sans saut de ligne
            for i, (_, row) in enumerate(df_sorted.iterrows()):
                row_idx = i + 2  # ligne Excel (1-indexée + entête)
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)

                    if pd.isna(value):
                        cell.value = None
                    else:
                        try:
                            # Conserve les entiers réels, sinon cast en string
                            v = int(value)
                            if str(v) == str(value).strip():
                                cell.value = v
                            else:
                                cell.value = value
                        except (ValueError, TypeError):
                            cell.value = value

                        # Ajout d'hyperliens pour la colonne Activite
                        if col_activite is not None:
                            if col_idx == col_activite and "Hyperlien" in df.columns:
                                lien = row["Hyperlien"]
                                if lien:
                                    cell.hyperlink = lien
                                    cell.font = Font(color="0000EE", underline="single")
                                else:
                                    cell.hyperlink = None
                                    cell.font = copy(source_font)   

            # Sauvegarde dans un buffer mémoire
            buffer = io.BytesIO()
            wb.save(buffer)
        else:
            # Sauvegarde dans un buffer mémoire
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    df_sorted.to_excel(writer, index=False)

        # Revenir au début du buffer pour le téléchargement
        buffer.seek(0)
        return buffer

    # Version modale
    @st.dialog("Sauvegarder données")
    def show_dialog_sauvegarder_contexte(df, nom_fichier, df_hash=None):
        st.markdown("Voulez-vous sauvegarder les données ?")
        col1, col2 = st.columns([1, 1])
        with col1:
            if df_hash is None:
                df_hash = hash_df(st.session_state.df, colonnes_a_enlever=["Debut_dt", "Duree_dt", "__uuid"])
            prev_hash = st.session_state.get("__contexte_hash")
            buffer = st.session_state.get("__contexte_buffer")

            if df_hash != prev_hash or buffer is None:
                # Le df a changé, on régénère le buffer
                buffer = serialiser_contexte(st.session_state.df)
                st.session_state["__contexte_hash"] = df_hash
                st.session_state["__contexte_buffer"] = buffer

            # Bouton de téléchargement
            if st.download_button(
                label="Valider",
                data=st.session_state["__contexte_buffer"],
                file_name=nom_fichier,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=CENTRER_BOUTONS
            ):
                st.rerun()
        with col2:
            if st.button(LABEL_BOUTON_ANNULER, use_container_width=CENTRER_BOUTONS):
                st.rerun()

    # Version Non Modale
    nom_fichier = st.session_state.get("fn", "planning_avignon.xlsx")
    
    if df_hash is None:
        df_hash = hash_df(st.session_state.get("df"), colonnes_a_enlever=["Debut_dt", "Duree_dt", "__uuid"])
    prev_hash = st.session_state.get("__contexte_hash")
    buffer = st.session_state.get("__contexte_buffer")

    if (df_hash != prev_hash or buffer is None) and est_contexte_valide():
        # Le df a changé, on régénère le buffer
        buffer = serialiser_contexte(st.session_state.df)
        st.session_state["__contexte_hash"] = df_hash
        st.session_state["__contexte_buffer"] = buffer

    # Bouton de téléchargement
    st.download_button(
        label=LABEL_BOUTON_SAUVEGARDER,
        data=st.session_state.get("__contexte_buffer", ""),
        file_name=nom_fichier,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=CENTRER_BOUTONS,
        disabled=not est_contexte_valide()
    )

# Programme une activité non programmée à une date donnée
def programmer_activite_non_programmee(date_ref, activite):

    df = st.session_state.df
    type_activite = activite["__type_activite"]
    undo.save()
    if type_activite == "ActiviteExistante":
        # Pour les spectacles, on programme la date et l'heure
        index = activite["__index"]
        bd_modifier_cellule(index, "Date", date_ref)
    elif type_activite == "déjeuner":
        # Pour les pauses, on ne programme pas d'heure spécifique
        index = len(df)
        bd_ajouter_activite(
            idx=index, 
            nom="Pause déjeuner",
            jour=date_ref, 
            debut=activite["Debut"],
            duree=formatter_timedelta(st.session_state.DUREE_REPAS),
            )
    elif type_activite == "dîner":
        # Pour les pauses, on ne programme pas d'heure spécifique
        index = len(df)
        bd_ajouter_activite(
            idx=index, 
            nom="Pause dîner",
            jour=date_ref, 
            debut=activite["Debut"],
            duree=formatter_timedelta(st.session_state.DUREE_REPAS),
            )
    elif type_activite == "café":
        # Pour les pauses, on ne programme pas d'heure spécifique
        index = len(df)
        bd_ajouter_activite(
            idx=index, 
            nom="Pause café",
            jour=date_ref, 
            debut=activite["Debut"],
            duree=formatter_timedelta(st.session_state.DUREE_CAFE),
            )
    else:
        return

    demander_selection("activites_programmees", index, deselect="activites_non_programmees")
    demander_selection("creneaux_disponibles", get_creneau_proche(st.session_state.creneaux_disponibles, activite)[0])
    st.session_state["activites_programmables_selected_row"] = None
    forcer_reaffichage_df("creneaux_disponibles")
    sql.sauvegarder_row(index)
    st.rerun()

# Renvoie les jours possibles pour programmer une activité donnée par son idx
def get_jours_possibles(df, activites_programmees, idx_activite):
    try:
        jours_possibles = []

        # Retour si index non valide
        if idx_activite not in df.index:
            return jours_possibles

        # Récupérer la durée de l'activité à considérer
        ligne_a_considerer = df.loc[idx_activite]
        debut = ligne_a_considerer["Debut_dt"]
        fin = ligne_a_considerer["Debut_dt"] + ligne_a_considerer["Duree_dt"]

        if activites_programmees is not None:
            for jour in range(st.session_state.periode_a_programmer_debut.day, st.session_state.periode_a_programmer_fin.day + 1):
                
                if not est_hors_relache(ligne_a_considerer["Relache"], jour):
                    continue

                activites_programmes_du_jour = activites_programmees[activites_programmees["Date"] == jour].sort_values("Debut_dt")

                if not activites_programmes_du_jour.empty:
                    # Créneau entre minuit et première activité du jour
                    premiere_activite_du_jour = activites_programmes_du_jour.iloc[0]
                    borne_inf = datetime.datetime.combine(BASE_DATE, datetime.time.min)  # 00h00
                    borne_sup = premiere_activite_du_jour["Debut_dt"]
                    if debut > borne_inf and fin < borne_sup - st.session_state.MARGE:
                        jours_possibles.append(jour)
                        continue  # on prend le premier créneau dispo du jour

                    # Ensuite, créneaux entre chaque activité programmée
                    for _, ligne in activites_programmes_du_jour.iterrows():
                        borne_inf, borne_sup, _ = get_creneau_bounds_apres(activites_programmes_du_jour, ligne)
                        if debut > borne_inf + st.session_state.MARGE and (borne_sup is None or fin < borne_sup - st.session_state.MARGE):
                            jours_possibles.append(jour)
                            break  # jour validé, on passe au suivant
                else: # jour libre
                    jours_possibles.append(jour)
    except Exception as e:
        print(f"Erreur in get_jours_possibles : {e}")
    return jours_possibles

# Renvoie les jours possibles pour programmer une activité donnée par son idx
def est_jour_possible(df, activites_programmees, idx_activite, jour):
    try:
        # Retour si index non valide
        if idx_activite not in df.index:
            return False

        # Récupérer la durée de l'activité à considérer
        ligne_a_considerer = df.loc[idx_activite]
        debut = ligne_a_considerer["Debut_dt"]
        fin = ligne_a_considerer["Debut_dt"] + ligne_a_considerer["Duree_dt"]

        if activites_programmees is not None:
                
            if not est_hors_relache(ligne_a_considerer["Relache"], jour):
                return False

            activites_programmes_du_jour = activites_programmees[activites_programmees["Date"] == jour].sort_values("Debut_dt")

            if not activites_programmes_du_jour.empty:
                # Créneau entre minuit et première activité du jour
                premiere_activite_du_jour = activites_programmes_du_jour.iloc[0]
                borne_inf = datetime.datetime.combine(BASE_DATE, datetime.time.min)  # 00h00
                borne_sup = premiere_activite_du_jour["Debut_dt"]
                if debut > borne_inf + st.session_state.MARGE and fin < borne_sup - st.session_state.MARGE:
                    return True

                # Ensuite, créneaux entre chaque activité programmée
                for _, ligne in activites_programmes_du_jour.iterrows():
                    borne_inf, borne_sup, _ = get_creneau_bounds_apres(activites_programmes_du_jour, ligne)
                    if debut > borne_inf + st.session_state.MARGE and (borne_sup is None or fin < borne_sup - st.session_state.MARGE):
                        return True
            else: # jour libre
                return True
    except Exception as e:
        print(f"Erreur in get_jours_possibles : {e}")
    return False

# Transforme en set un __options_date au format json
def parse_options_date(s):
    """Retourne un set[int] à partir du JSON éventuellement hétérogène."""
    if not s or pd.isna(s):
        return set()
    try:
        lst = json.loads(s)
    except Exception:
        return set()
    # force en int, ignore ce qui n'est pas convertible
    out = set()
    for x in lst:
        try:
            out.add(str(x))
        except Exception:
            pass
    return out

# Met au format json un __options_date au format set
def dump_options_date(sset):
    """Serialize un set[int] en JSON trié."""
    return json.dumps(sorted(str(x) for x in sset))

# Met à jour la colonne __options_date d'un df_display donné pour un jour donné
def maj_options_date(df, activites_programmees, df_display, jour):
    """
    - jour: int (jour modifié)
    Met à jour uniquement les lignes dont __options_date contient `jour`.
    Retourne la liste des index modifiés.
    """
    if jour is None or pd.isna(jour):
        return

    jour = str(jour)

    changed_idx = []

    # Pré-filtrage simple : on parcourt uniquement les lignes où la chaîne n'est pas vide.
    # (on pourrait accélérer avec .dropna() / .astype(str), mais restons sûrs)
    for i, s in df_display["__options_date"].items():
        if not s:
            continue
    
        # parse -> set[str]
        opts = parse_options_date(s)
        
        # Activité courante
        row = df_display.loc[i]

        # S'il s'agit d'une activité programmée au jour dit...
        if row["Date"] == jour:
            # S'il s'agit d'une activité réservée on vérifie que le menu est vide. Si ce n'est pas le cas on le vide.
            if est_activite_reserve(df.loc[i]):
                if opts != set():
                    df_display.at[i, "__options_date"] = dump_options_date(set())
                    changed_idx.append(i)
            # Sinon on vérifie que le menu n'est pas vide (cas d'une activité qui serait passée de réservée à non réservée).
            # Dans ce cas on reconstruit le menu.
            else:
                if opts == set():
                    df_display.at[i, "__options_date"] = dump_options_date(get_jours_possibles_from_activite_programmee(row))
                    changed_idx.append(i)
            # Sinon rien d'autre à faire
            # car s'il s'agit d'une activité reprogrammée au jour dit ce jour était déjà dans le menu avant reprogrammation et doit y rester
            # et sinon ce jour est déjà dans le menu et doit y rester aussi pour que le déploiement dudit menu n'oblige pas à changer de jour.
        
        else:

            # si le jour n'était pas présent ET que la règle ne le concerne pas, on peut sauter
            # (mais on doit tout de même appeler la règle si tu veux ajouter quand c'est possible)
            allowed = est_jour_possible(df, activites_programmees, i, int(jour))

            # remove si plus possible
            if not allowed and jour in opts:
                opts.remove(jour)
                if len(opts) == 1 and '' in opts:
                    opts = set() # un menu ne doit pas avoir un seul élément vide
                df_display.at[i, "__options_date"] = dump_options_date(opts)
                changed_idx.append(i)

            # add si maintenant possible
            elif allowed and jour not in opts:
                opts.add(jour)
                if len(opts) == 1:
                    opts.add('') # il faut un item vide dans un menu avec des jours valides pour permettre la déprogrammation
                df_display.at[i, "__options_date"] = dump_options_date(opts)
                changed_idx.append(i)

    return changed_idx

# idem get_jours_possibles avec en paramètre une row d'activité programmée contenant en colonne __index l'index du df de base
# Les paramètres df et activites_programmees de get_jours_possibles sont supposés etre stockés dans st.session_state
def get_jours_possibles_from_activite_programmee(row: pd.Series):
    jours = get_jours_possibles(st.session_state.df, st.session_state.activites_programmees, row["__index"])
    jour_courant = int(row["Date"]) if pd.notna(row["Date"]) and row["Date"] is not None else row["Date"]
    if pd.notna(row["__index"]) and row["__index"] in st.session_state.df.index:
        if not est_activite_reserve(st.session_state.df.loc[row["__index"]]):
            jours = [jour_courant] + jours + [""] if jours != [] else [jour_courant] + [""]
        else: 
            jours = []
    return sorted([str(j) for j in jours]) if isinstance(jours, list) else []

# idem get_jours_possibles avec en paramètre une row d'activité non programmée contenant en colonne __index l'index du df de base
# Les paramètres df et activites_programmees de get_jours_possibles sont supposés etre stockés dans st.session_state
def get_jours_possibles_from_activite_non_programmee(row: pd.Series):
    jours = get_jours_possibles(st.session_state.df, st.session_state.activites_programmees, row["__index"])
    jours = [""] + jours if jours != [] else jours
    return [str(j) for j in jours] if isinstance(jours, list) else []

# Calcule les options des dates pour les activiés programmées
def calculer_options_date_activites_programmees(df_display):
    # Hash non pertinent en l'état car cette fonction n'est appelée par bd_maj_activites_non_programmees que si les données d'entrée on changé
    # hash_val  = hash_df(
    #     df_display, 
    #     colonnes_a_garder=["Date", "Debut", "Duree"], 
    #     params=[
    #         st.session_state.periode_a_programmer_debut.isoformat(), 
    #         st.session_state.periode_a_programmer_fin.isoformat(),
    #         str(st.session_state.MARGE.total_seconds()),
    #         str(st.session_state.DUREE_REPAS.total_seconds()),
    #         str(st.session_state.DUREE_CAFE.total_seconds())])
    # hash_key = "options_date_activites_programmees__hash"
    # key = "options_date_activites_programmees"
    # if st.session_state.get(hash_key) != hash_val:
    #     st.session_state[key] = df_display.apply(lambda row: get_jours_possibles_from_activite_programmee(row), axis=1)
    #     st.session_state[hash_key] = hash_val
    # return st.session_state[key]
    return df_display.apply(lambda row: get_jours_possibles_from_activite_programmee(row), axis=1)

# Calcule les options des dates pour les activiés non programmées
def calculer_options_date_activites_non_programmees(df_display):
    # Hash non pertinent en l'état car cette fonction n'est appelée par bd_maj_activites_non_programmees que si les données d'entrée on changé
    # hash_val  = hash_df(
    #     df_display, 
    #     colonnes_a_garder=["Date", "Debut", "Duree"], 
    #     params=[
    #         st.session_state.periode_a_programmer_debut.isoformat(), 
    #         st.session_state.periode_a_programmer_fin.isoformat(),
    #         str(st.session_state.MARGE.total_seconds()),
    #         str(st.session_state.DUREE_REPAS.total_seconds()),
    #         str(st.session_state.DUREE_CAFE.total_seconds())])
    # hash_key = "options_date_activites_non_programmees__hash"
    # key = "options_date_activites_non_programmees"
    # if st.session_state.get(hash_key) != hash_val:
    #     st.session_state[key] = df_display.apply(lambda row: get_jours_possibles_from_activite_non_programmee(row), axis=1)
    #     st.session_state[hash_key] = hash_val
    # return st.session_state[key]
    return df_display.apply(lambda row: get_jours_possibles_from_activite_non_programmee(row), axis=1)

# Programme une activité choisie en fonction des jours possibles
def programmer_activite_par_choix_activite():

    df = st.session_state.get("df")
    if df is None or len(df) <= 0:
        return

    st.markdown("##### Programmation d'une nouvelle activité")

    # Filtrer les activités non programmées
    activites_programmees = st.session_state.get("activites_programmees")
    activites_non_programmees = st.session_state.get("activites_non_programmees")

    # Liste d'options formatées
    options_activites = []
    for idx, row in activites_non_programmees.iterrows():
        if get_jours_possibles(df, activites_programmees, idx):
            label = f"[{row["Debut"]} - {row["Fin"]}] - {str(row["Activite"]).strip()}"
            options_activites.append((label, idx))

    # Afficher la selectbox des activités
    activite_selectionee = st.selectbox("Choix de l'activité à programmer :", options_activites, format_func=lambda x: x[0])
    if activite_selectionee:
        idx_choisi = activite_selectionee[1]

        # Déterminer les jours disponibles 
        jours_possibles = get_jours_possibles(df, activites_programmees, idx_choisi)
        jours_label = [f"{int(jour):02d}" for jour in jours_possibles]

        jour_selection = st.selectbox("Choix du jour :", jours_label)

        # Bouton pour confirmer
        if jour_selection:
            if st.button(LABEL_BOUTON_PROGRAMMER, key="AjouterAuPlanningParChoixActivite"):
                jour_choisi = int(jour_selection.split()[-1])

                # On peut maintenant modifier le df
                df.at[idx_choisi, "Date"] = jour_choisi
                st.rerun()

def get_creneau_proche(creneaux: pd.DataFrame, activite):
    """
    A partir d'une liste de créneaux, renvoie le créneau le plus proche d'une activité donnée selon les critères suivants:
      1 : Date activité manquante → premier créneau qui contient l’activité
      2 : même jour, créneau dont Début >= Fin activité
      3 : même jour, créneau qui contient totalement l’activité
      4 : même jour, créneau dont Fin <= Début activité
      5 : jour futur le plus proche
      6 : jour passé le plus proche
      fallback : aucune correspondance, première ligne
    
    S’il n’y a aucun candidat (i.e. vide ou Date non utilisables) -> (index de la 1ère ligne, ligne, 'fallback').
    
    Paramètres:
    - creneaux: liste de creneaux fournie sous la forme d'un DataFrame tel que renvoyé par get_creneaux
    - activite: activité
    
    retours:
    - index du créneau sélectionné
    - créneau sélectionné
    - critère de choix
    """
    def _interval_distance_to_window(d, f, win_start, win_end):
        """
        Distance entre l'intervalle [d,f] et la fenêtre [win_start, win_end].
        0 si recouvrement; sinon la distance minimale entre bords.
        Robuste si d ou f manquent (on réduit à un point).
        """
        if pd.isna(d) and pd.isna(f):
            return 10**9
        if pd.isna(d): d = f
        if pd.isna(f): f = d
        if d > f:
            d, f = f, d
        # Overlap ?
        if not (f < win_start or d > win_end):
            return 0
        # Sinon distance au plus proche bord
        if f < win_start:
            return win_start - f
        else:  # d > win_end
            return d - win_end

    if creneaux is None or creneaux.empty or activite is None:
        return None, None, None

    hdeb = activite.get("Debut")
    hfin = activite.get("Fin")
    win_start = hhmm_to_min(hdeb)
    win_end   = hhmm_to_min(hfin)

    jour = safe_int(activite.get("Date"))

    work = creneaux.copy()
    work["Date"]   = pd.to_numeric(work["Date"], errors="coerce")
    work["_debut"] = work["Debut"].map(hhmm_to_min)
    work["_fin"]   = work["Fin"].map(hhmm_to_min)

    # corrige inversions Debut/Fin
    mask_swap = work["_debut"].notna() & work["_fin"].notna() & (work["_fin"] < work["_debut"])
    if mask_swap.any():
        tmp = work.loc[mask_swap, "_debut"].copy()
        work.loc[mask_swap, "_debut"] = work.loc[mask_swap, "_fin"]
        work.loc[mask_swap, "_fin"]   = tmp

    # ---------- 1 : Date activité manquante → créneau qui contient l’activité
    if jour is None or pd.isna(jour):
        for idx, creneau in work.iterrows():
            p = get_proposables(creneau, traiter_pauses=False)
            if activite["__uuid"] in p["__uuid"].values:
                return idx, creneau, "1"
        idx0 = creneaux.index[0]
        return idx0, creneaux.loc[idx0], "fallback"

    # ---------- 2 : même jour, Debut >= Fin activité
    j = int(jour)
    if win_end is not None:
        r2 = work[(work["Date"] == j) & (work["_debut"].notna()) & (work["_debut"] >= win_end)]
        if not r2.empty:
            idx = (r2["_debut"] - win_end).idxmin()
            return idx, creneaux.loc[idx], "2"

    # ---------- 3 : même jour, créneau contenant l’activité
    if win_start is not None and win_end is not None:
        r3 = work[(work["Date"] == j)
                 & work["_debut"].notna() & work["_fin"].notna()
                 & (work["_debut"] <= win_start) & (work["_fin"] >= win_end)]
        if not r3.empty:
            slack_left  = win_start - r3["_debut"]
            slack_right = r3["_fin"] - win_end
            r3 = r3.assign(_slack_total = slack_left + slack_right,
                           _slack_left = slack_left,
                           _slack_right = slack_right)
            cand = r3.sort_values(
                by=["_slack_total", "_slack_left", "_slack_right", "_debut", "_fin"],
                ascending=[True, True, True, True, True]
            ).iloc[0]
            return cand.name, creneaux.loc[cand.name], "3"

    # ---------- 4 : même jour, Fin <= Début activité
    if win_start is not None:
        r4 = work[(work["Date"] == j) & (work["_fin"].notna()) & (work["_fin"] <= win_start)]
        if not r4.empty:
            idx = (win_start - r4["_fin"]).idxmin()
            return idx, creneaux.loc[idx], "4"

    # distance au créneau pour futur/passé
    work["_win_dist"] = work.apply(
        lambda r: _interval_distance_to_window(r["_debut"], r["_fin"], win_start, win_end), axis=1
    )

    # ---------- 5 : jour futur le plus proche
    r5 = work[(work["Date"].notna()) & (work["Date"] >= j)]
    if not r5.empty:
        r5 = r5.assign(_day_dist=(r5["Date"] - j).astype("int64"))
        cand = r5.sort_values(by=["_day_dist", "_win_dist", "_debut", "_fin"],
                              ascending=[True, True, True, True]).iloc[0]
        return cand.name, creneaux.loc[cand.name], "5"

    # ---------- 6 : jour passé le plus proche
    r6 = work[(work["Date"].notna()) & (work["Date"] <= j)]
    if not r6.empty:
        r6 = r6.assign(_day_dist=(j - r6["Date"]).astype("int64"))
        cand = r6.sort_values(by=["_day_dist", "_win_dist", "_debut", "_fin"],
                              ascending=[True, True, True, True]).iloc[0]
        return cand.name, creneaux.loc[cand.name], "6"

    # ---------- fallback
    idx0 = creneaux.index[0]
    return idx0, creneaux.loc[idx0], "fallback"

def get_proposables(creneau, traiter_pauses=False):

    proposables = []

    df = st.session_state.get("df")
    if df is None or len(df) <= 0:
        return proposables

    type_creneau = creneau["__type_creneau"]
    idx = creneau["__index"]
    date_ref = int(creneau["Date"]) # date_ref doit être en int !

    if type_creneau == "Avant":
        activites_programmees = st.session_state.get("activites_programmees")
        if activites_programmees is None:
            return proposables
        try:
            ligne_ref = activites_programmees.loc[idx]
        except Exception as e:
            print(f"Erreur afficher_creneaux_disponibles : {e}")
            return proposables
        proposables = get_activites_programmables_avant(df, activites_programmees, ligne_ref, traiter_pauses)

    elif type_creneau == "Après":
        activites_programmees = st.session_state.get("activites_programmees")
        if activites_programmees is None:
            return proposables
        try:
            ligne_ref = activites_programmees.loc[idx]
        except Exception as e:
            print(f"Erreur afficher_creneaux_disponibles : {e}")
            return proposables
        proposables = get_activites_programmables_apres(df, activites_programmees, ligne_ref, traiter_pauses)

    elif type_creneau == "Journée":
        proposables = get_activites_programmables_sur_journee_entiere(date_ref, traiter_pauses)

    proposables = pd.DataFrame(proposables).sort_values(by=["Debut"], ascending=[True]) if proposables else pd.DataFrame(proposables)
    proposables["Date"] = creneau["Date"] # ou str(date_ref) car col Date au format string dans les df_display !

    return proposables

# Programme une activité en fonction des créneaux possibles
def afficher_creneaux_disponibles():

    def on_toggle_pauses():
        st.session_state.traiter_pauses_change = True
        st.session_state.traiter_pauses = st.session_state.traiter_pauses_cb
        bd_maj_creneaux_disponibles()
        sql.sauvegarder_param("traiter_pauses")
        st.session_state.creneaux_disponibles_choix_activite = None

    df = st.session_state.get("df")
    if df is None or len(df) <= 0:
        return
    
    with st.expander("**Créneaux disponibles**", expanded=True):

        # Gestion du flag de traitement des pauses
        traiter_pauses = st.checkbox("Tenir compte des pauses", value=st.session_state.get("traiter_pauses", False), key="traiter_pauses_cb", on_change=on_toggle_pauses)  
        traiter_pauses_change = st.session_state.get("traiter_pauses_change", False)
        st.session_state["traiter_pauses_change"] = False

        creneaux_disponibles = st.session_state.get("creneaux_disponibles")
        if creneaux_disponibles is None or creneaux_disponibles.empty:
            return 

        proposables = []

        st.session_state.creneaux_disponibles_choix_activite = None

        # Récupération du creneau enregistré au run précédent
        choix_creneau_pred = st.session_state["creneaux_disponibles_selected_row"] if "creneaux_disponibles_selected_row" in st.session_state else None

        # Affichage de la grille des créneaux disponibles
        choix_creneau = afficher_df(
            "Créneaux disponibles", 
            creneaux_disponibles, 
            header_names={"Debut": "Début"},
            fixed_columns={"Date": 55, "Debut": 55, "Fin": 55}, 
            hide=["__type_creneau", "__index", "__uuid"], 
            key="creneaux_disponibles", 
            hide_label=True, 
            colorisation=True)

        if choix_creneau is not None:

            # Choix d'une activité à programmer dans le creneau choisi
            if (choix_creneau_pred is not None and choix_creneau_pred["__uuid"] != choix_creneau["__uuid"]) or \
                traiter_pauses_change or \
                "activites_programmables" not in st.session_state:

                if choix_creneau_pred is not None and choix_creneau_pred["__uuid"] != choix_creneau["__uuid"]:
                    forcer_reaffichage_df("activites_programmables")
            
                proposables = get_proposables(choix_creneau, traiter_pauses)

                st.session_state.activites_programmables = proposables

                # Resélection automatique de l'activité précédemment séléectionnée si elle existe dans la nouvelle liste de proposables
                st.session_state.setdefault("activites_programmables_select_auto", True)
                if st.session_state["activites_programmables_select_auto"]:
                    current_selected_row = st.session_state.get("activites_programmables_selected_row")
                    current_selected_row_idx = get_index_from_uuid(proposables, current_selected_row["__uuid"]) if current_selected_row is not None else None
                    current_selected_row_idx = current_selected_row_idx if current_selected_row_idx is not None else proposables.index[0] if isinstance(proposables, pd.DataFrame) else None
                    demander_selection("activites_programmables", current_selected_row_idx)
                else:
                    st.session_state["activites_programmables_select_auto"] = True
            else: 
                proposables = st.session_state.get("activites_programmables", [])
        else:
            st.session_state.activites_programmables = None

    if isinstance(proposables, pd.DataFrame) and not proposables.empty:
        with st.expander("**Activités programmables**", expanded=True):
                date_ref = int(choix_creneau["Date"]) # date_ref doit être en int !
                st.markdown(f"Sur le créneau du {int(date_ref)} de {choix_creneau["Debut"]} à {choix_creneau["Fin"]}")

                activite = afficher_df(
                    "Activités programmables", 
                    proposables, 
                    header_names={"Debut": "Début", "Duree": "Durée", "Activite": "Activité", "Relache": "Relâche", "Priorite": "Prio", "Reserve": "Réservé"},
                    fixed_columns={"Date": 55, "Debut": 55, "Fin": 55, "Duree": 55}, 
                    hide=["__type_activite", "__index", "__uuid"], 
                    key="activites_programmables", 
                    hide_label=True, 
                    colorisation=True,
                )

                st.markdown(f"{activite["Activite"]} le {activite["Date"]} à {activite["Debut"]}" if activite is not None else "Aucune activité sélectionnée")

                # Gestion du bouton Programmer
                if st.button(LABEL_BOUTON_PROGRAMMER, disabled=activite is None, key="PagePrincipaleProgrammerParCréneau"):
                    if activite is not None:
                        st.session_state.forcer_menu_activites_programmees = True
                        programmer_activite_non_programmee(date_ref, activite)

# Force le reaffichage de l'agrid des activités programmées
def forcer_reaffichage_activites_programmees():
    st.session_state.activites_programmees_key_counter += 1 

# Force le reaffichage de l'agrid des activités non programmées
def forcer_reaffichage_activites_non_programmees():
    st.session_state.activites_non_programmees_key_counter += 1 

# Signale au DOM une modification de df_display en incrémentant la première ligne de la colonne __df_push_ver.
# Cette incrémentation est captée par le JsCode JS_SELECT_DESELECT_ONCE, lequel declenche un selectionChanged de type "api"
# qui permet à Streamlit de renvoyer la prise en compte des modifications du df_display via response["data"] sans attendre de clic utilisateur.
def signaler_df_push(grid_name):
    df_display = st.session_state.get(grid_name + "_df_display")
    if df_display is not None:
        df_display.loc[df_display.index[0], "__df_push_ver"] = int(df_display.iloc[0]["__df_push_ver"] or 0) + 1

# Initialisation des variables d'état du contexte après chargement des données du contexte
def initialiser_etat_contexte(df, wb, fn, fp, ca):
    st.session_state.df = df
    st.session_state.wb = wb
    st.session_state.fn = fn
    st.session_state.fp = fp
    st.session_state.ca = ca
    st.session_state.nouveau_fichier = True
    st.session_state.compteur_activite = 0
    st.session_state.menu_activites = {"menu": "menu_activites_non_programmees", "index_df": None}
    st.session_state.menu_activites_programmees = None
    st.session_state.menu_activites_non_programmees = None
    st.session_state.forcer_menu_activites_programmees = False
    st.session_state.forcer_menu_activites_non_programmees = False
    st.session_state.forcer_maj_menu_activites_programmees = False
    st.session_state.forcer_maj_menu_activites_non_programmees = copy.deepcopy(SEL_REQUEST_DEFAUT)
    st.session_state.activites_programmees_sel_request = copy.deepcopy(SEL_REQUEST_DEFAUT)
    st.session_state.activites_non_programmees_sel_request = copy.deepcopy(SEL_REQUEST_DEFAUT)
    st.session_state.creneaux_disponibles_sel_request = copy.deepcopy(SEL_REQUEST_DEFAUT)
    st.session_state.activites_programmables_sel_request =copy.deepcopy(SEL_REQUEST_DEFAUT)

    forcer_reaffichage_df("creneaux_disponibles")

# Ajout d'une nouvelle activité 
def afficher_bouton_nouvelle_activite(disabled=False, key="ajouter_activite"):
    import numpy as np

    df = st.session_state.df

    # Initialiser le DataFrame dans session_state si absent
    if "compteur_activite" not in st.session_state:
        st.session_state.compteur_activite = 0

    # Bouton Ajouter
    if st.button(LABEL_BOUTON_AJOUTER, use_container_width=CENTRER_BOUTONS, disabled=disabled, key=key):

        undo.save()
        
        new_idx = bd_ajouter_activite()

        demander_selection("activites_non_programmees", new_idx, deselect="activites_programmees")
        st.session_state.editeur_activite_idx = new_idx
        
        # Bascule du menu activité sur le menu_activites_non_programmees
        st.session_state.menu_activites = {
            "menu": "menu_activites_non_programmees",
            "index_df": new_idx
        }

        forcer_reaffichage_df("activites_programmables")
        sql.sauvegarder_row(new_idx)
        st.rerun()

# Charge le fichier Excel contenant les activités à programmer
def charger_contexte_depuis_fichier():

    @st.dialog("Charger fichier")
    def dialog_charger_fichier():
        # Chargement du fichier Excel contenant les activités à programmer
        fd = st.file_uploader(
            "Choix du fichier Excel contenant les activités à programmer", 
            type=["xlsx"], 
            label_visibility="collapsed",
            key="file_uploader",
        )

        if st.button("Charger", use_container_width=CENTRER_BOUTONS, disabled=fd is None):

            try:
                st.session_state.contexte_invalide = True
                curseur_attente()
                df = pd.read_excel(fd)
                wb = load_workbook(fd)
                lnk = get_liens_activites(wb)
                sheetnames = wb.sheetnames
                ca = pd.read_excel(fd, sheet_name=sheetnames[1]) if len(sheetnames) > 1 else None
                df = nettoyer_donnees(df, fd.name) # si ok RAZ du contexte_invalide

                if "contexte_invalide" not in st.session_state:
                    df = add_persistent_uuid(df)
                    df = add_hyperliens(df, lnk)
                    fn = fd.name if fd is not None else ""
                    fp = upload_excel_to_dropbox(fd.getvalue(), fd.name) if fd is not None else ""
                    undo.save()
                    initialiser_etat_contexte(df, wb, fn, fp, ca)
                    initialiser_periode_programmation(df)
                    st.session_state["push_periode_programmation_modele_values"] = True 
                    # undo.init(verify=False)
                    bd_maj_contexte(maj_donnees_calculees=True)
                    sql.sauvegarder_contexte()
                    selection = st.session_state.activites_non_programmees.index[0] if len(st.session_state.activites_non_programmees) > 0 else None
                    demander_selection("activites_non_programmees", selection, deselect="activites_programmees")
                    st.session_state.menu_activites = {
                        "menu": "menu_activites_non_programmees",
                        "index_df": selection
                    }
                    st.session_state.forcer_maj_menu_activites_non_programmees = True
                    st.rerun()

            except Exception as e:
                st.sidebar.error(f"Erreur de chargement du fichier : {e}")

    if st.button("Charger", use_container_width=CENTRER_BOUTONS):
        dialog_charger_fichier()

# Initialisation des types d'un df vide
def initialiser_dtypes(df):
    for col in df.columns:
        if col in COLONNES_TYPE_INT:
            df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
        elif col in COLONNES_TYPE_STRING:
            df[col] = df[col].astype("string")
        elif col in COLONNES_TYPE_OBJECT:
            df[col] = df[col].astype("object")
    if "Debut_dt" not in df.columns:
        df["Debut_dt"] = pd.Series(dtype="datetime64[ns]")
    else:
        df["Debut_dt"] = df["Debut_dt"].astype("datetime64[ns]")
    if "Duree_dt" not in df.columns:
        df["Duree_dt"] = pd.Series(dtype="timedelta64[ns]")
    else:
        df["Duree_dt"] = df["Duree_dt"].astype("timedelta64[ns]")

# Initialisation d'un nouveau contexte
def initialiser_nouveau_contexte():

    if "contexte_invalide" in st.session_state:
        del st.session_state["contexte_invalide"]

    df = pd.DataFrame(columns=COLONNES_ATTENDUES)
    df = add_persistent_uuid(df)
    df = add_hyperliens(df)
    initialiser_dtypes(df)
    wb = None
    fn = "planning_avignon.xlsx"
    fp = ""
    ca = pd.DataFrame(columns=COLONNES_ATTENDUES_CARNET_ADRESSES)
    
    initialiser_etat_contexte(df, wb, fn, fp, ca)
    initialiser_periode_programmation(df)
    sql.sauvegarder_contexte()

# Création d'un nouveau contexte
def creer_nouveau_contexte():
    if st.button(LABEL_BOUTON_NOUVEAU, use_container_width=CENTRER_BOUTONS, key="creer_nouveau_contexte"):
        curseur_attente()
        undo.save()
        initialiser_nouveau_contexte()
        bd_maj_contexte(maj_donnees_calculees=True)
        st.rerun()

# Indique si le contexte est vlide pour traitement
def est_contexte_valide():
    return "df" in st.session_state and isinstance(st.session_state.df, pd.DataFrame) and "contexte_invalide" not in st.session_state

# Ajout d'une nouvelle activité à la bd contexte
# @chrono
def bd_ajouter_activite(idx=None, nom=None, jour=None, debut=None, duree=None):
    def get_nom_nouvelle_activite(df):
        noms_existants = df["Activite"].dropna().astype(str).str.strip().tolist()
        while True:
            st.session_state.compteur_activite += 1
            nom_candidat = f"Activité {st.session_state.compteur_activite}"
            if nom_candidat not in noms_existants:
                return nom_candidat
            
    def get_next_free_index(df):
        existing = set(df.index)
        i = 0
        while i in existing:
            i += 1
        return i
    
    df = st.session_state.get("df", None)
    if df is None:
        return
    
    idx = get_next_free_index(df) if idx is None else idx
    nom = get_nom_nouvelle_activite(df) if nom is None else nom
    jour = pd.NA if jour is None else jour
    debut = "09h00" if debut is None else debut
    duree = "1h00" if duree is None else duree

    df.at[idx, "Date"] = jour
    df.at[idx, "Debut"] = debut
    df.at[idx, "Duree"] = duree
    df.at[idx, "Activite"] = nom
    df.at[idx, "Hyperlien"] = f"https://www.festivaloffavignon.com/resultats-recherche?recherche={nom.replace(' ', '+')}"
    add_persistent_uuid(df, idx)
    bd_maj_donnees_calculees_row(idx, full=False)

    row = df.loc[[idx]]

    if est_activite_programmee(row):
        st.session_state.activites_programmees = pd.concat([st.session_state.activites_programmees, row]).sort_values(by=["Date", "Debut"], ascending=[True, True])

        row = bd_creer_df_display_activites_programmees(row)
        st.session_state.activites_programmees_df_display = pd.concat([st.session_state.activites_programmees_df_display, row]).sort_values(by=["Date", "Début"], ascending=[True, True])
        st.session_state.activites_programmees_df_display_copy = st.session_state.activites_programmees_df_display.copy()

        bd_maj_creneaux_disponibles()

    elif est_activite_non_programmee(row):
        st.session_state.activites_non_programmees = pd.concat([st.session_state.activites_non_programmees, row]).sort_values(by=["Date", "Debut"], ascending=[True, True])

        row = bd_creer_df_display_activites_non_programmees(row)
        st.session_state.activites_non_programmees_df_display = pd.concat([st.session_state.activites_non_programmees_df_display, row]).sort_values(by=["Date", "Début"], ascending=[True, True])
        st.session_state.activites_non_programmees_df_display_copy = st.session_state.activites_non_programmees_df_display.copy()

        bd_maj_creneaux_disponibles()
    
    return idx

def bd_creer_df_display_activites_non_programmees(activites_non_programmees, maj_options_date=True):
    df_display = activites_non_programmees.copy()
    df_display["__index"] = df_display.index
    if "__options_date" not in df_display or maj_options_date:
        df_display["__options_date"] = calculer_options_date_activites_non_programmees(df_display) 
        df_display["__options_date"] = df_display["__options_date"].map(safe_json_dump)
    df_display["Date"] = df_display["Date"].apply(lambda x: str(int(x)) if pd.notna(x) and float(x).is_integer() else "")
    df_display["__desel_ver"] = st.session_state.activites_programmees_sel_request["desel"]["ver"] if "activites_programmees_sel_request" in st.session_state else 0
    df_display["__desel_id"] =  get_uuid(df_display, st.session_state.activites_programmees_sel_request["desel"]["id"]) if "activites_programmees_sel_request" in st.session_state else None
    df_display["__sel_ver"] = st.session_state.activites_programmees_sel_request["sel"]["ver"] if "activites_programmees_sel_request" in st.session_state else 0
    df_display["__sel_id"] =  get_uuid(df_display, st.session_state.activites_programmees_sel_request["sel"]["id"]) if "activites_programmees_sel_request" in st.session_state else None
    df_display["__sel_source"] = "api"
    df_display["__df_push_ver"] = 0
    df_display = ensure_addr_cols(df_display)
    df_display.drop(columns=["Debut_dt", "Duree_dt"], inplace=True)
    df_display.rename(columns=RENOMMAGE_COLONNES, inplace=True)
    df_display = df_display.where(df_display.notna(), None)
    return df_display

def bd_creer_df_display_activites_programmees(activites_programmees, maj_options_date=True):
    df_display = activites_programmees.copy()
    df_display["__jour"] = df_display["Date"].apply(lambda x: int(str(int(float(x)))[-2:]) if pd.notna(x) else None)
    df_display["__index"] = df_display.index
    if "__options_date" not in df_display or maj_options_date:
        df_display["__options_date"] = calculer_options_date_activites_programmees(df_display) 
        df_display["__options_date"] = df_display["__options_date"].map(safe_json_dump)
    df_display["__non_reserve"] = df_display["Reserve"].astype(str).str.strip().str.lower() != "oui"
    df_display["Date"] = df_display["Date"].apply(lambda x: str(int(x)) if pd.notna(x) and float(x).is_integer() else "")
    df_display["__desel_ver"] = st.session_state.activites_programmees_sel_request["desel"]["ver"] if "activites_programmees_sel_request" in st.session_state else 0
    df_display["__desel_id"] =  get_uuid(df_display, st.session_state.activites_programmees_sel_request["desel"]["id"]) if "activites_programmees_sel_request" in st.session_state else None
    df_display["__sel_ver"] = st.session_state.activites_programmees_sel_request["sel"]["ver"] if "activites_programmees_sel_request" in st.session_state else 0
    df_display["__sel_id"] =  get_uuid(df_display, st.session_state.activites_programmees_sel_request["sel"]["id"]) if "activites_programmees_sel_request" in st.session_state else None
    df_display["__sel_source"] = "api"
    df_display["__df_push_ver"] = 0
    df_display = ensure_addr_cols(df_display)
    df_display.drop(columns=["Debut_dt", "Duree_dt"], inplace=True)
    df_display.rename(columns=RENOMMAGE_COLONNES, inplace=True)
    df_display = df_display.where(df_display.notna(), None)
    return df_display

# Met à jour le contexte pour une activité dont la date de programmation passe de jour à None
# Si le param jour est à None on prend comme date de programmation antérieure la valeur présente dans l'activité 
# (ce qui suppose que cette valeur n'a pas été modifiée préalablement par un bd_modifier_cell). 
# @chrono
def bd_deprogrammer(idx, jour=None):
    
    if "df" not in st.session_state:
        return
    
    if idx not in st.session_state.df.index:
        return
    
    if "activites_programmees" not in st.session_state:
        return
    
    if "activites_non_programmees" not in st.session_state:
        return

    if idx in st.session_state.activites_programmees.index:
        row = st.session_state.activites_programmees.loc[[idx]]
        if jour is None:
            jour = row.loc[idx]["Date"]
        row.at[idx, "Date"] = None
        st.session_state.activites_programmees.drop(index=idx, inplace=True)
        st.session_state.activites_non_programmees = pd.concat([st.session_state.activites_non_programmees, row]).sort_values(by=["Debut_dt"], ascending=[True])

        row = st.session_state.activites_programmees_df_display.loc[[idx]]
        row.at[idx, "Date"] = ""
        row.drop(columns=["__jour", "__non_reserve"], inplace=True)
        st.session_state.activites_programmees_df_display.drop(index=idx, inplace=True)
        st.session_state.activites_non_programmees_df_display = pd.concat([st.session_state.activites_non_programmees_df_display, row]).sort_values(by=["Début"], ascending=[True])

        maj_options_date(st.session_state.df, st.session_state.activites_programmees, st.session_state.activites_programmees_df_display, jour)
        maj_options_date(st.session_state.df, st.session_state.activites_programmees, st.session_state.activites_non_programmees_df_display, jour)

        st.session_state.activites_programmees_df_display_copy = st.session_state.activites_programmees_df_display.copy()
        st.session_state.activites_non_programmees_df_display_copy = st.session_state.activites_non_programmees_df_display.copy()

        bd_maj_creneaux_disponibles()

# Déprogrammation d'une activité programmée (si pause suppression, si activité ordinaire date à None)
def bd_deprogrammer_activite_programmee(idx):
    df = st.session_state.df
    if est_pause(df.loc[idx]):
        supprimer_activite(idx)
    else:
        if idx not in st.session_state.df.index:
            return
        jour = st.session_state.df.loc[idx]["Date"]
        modifier_df_cell(st.session_state.df, idx, "Date", None)
        bd_deprogrammer(idx, jour)

# Met à jour les variables d'état relatives aux activités programmées
# @chrono
def bd_maj_activites_programmees(maj_options_date=True):
    if st.session_state.get("df", None) is None:
        return  
    activites_programmees = get_activites_programmees(st.session_state.df)
    st.session_state.activites_programmees = activites_programmees
    df_display = bd_creer_df_display_activites_programmees(activites_programmees, maj_options_date)
    st.session_state.activites_programmees_df_display = df_display
    st.session_state.activites_programmees_df_display_copy = df_display.copy()

# Met à jour le contexte complet (activités programmées, non programmées et créneaux disponibles)
def bd_maj_contexte(maj_donnees_calculees=True, maj_options_date=True):
    st.session_state.setdefault("bd_maj_contexte_cmd", {"maj_donnees_calculees": maj_donnees_calculees, "maj_options_date": maj_options_date})
    tracer.log(f"Debut", types=["gen"])
    if maj_donnees_calculees:
        bd_maj_donnees_calculees()
    bd_maj_activites_programmees(maj_options_date) # pour mise à jour menus options date
    bd_maj_activites_non_programmees(maj_options_date) # pour mise à jour menus options date
    bd_maj_creneaux_disponibles()
    tracer.log(f"Fin", types=["gen"])
    del st.session_state["bd_maj_contexte_cmd"]

# Met à jour la variable d'état qui donne la liste des créneaux disponibles
# @chrono
def bd_maj_creneaux_disponibles():
    df = st.session_state.get("df")
    if df is None:
        return
    
    activites_programmees = st.session_state.get("activites_programmees")
    if activites_programmees is None:
        return
    
    traiter_pauses = st.session_state.get("traiter_pauses", False)
    st.session_state.creneaux_disponibles = get_creneaux(df, activites_programmees, traiter_pauses) 
    # if st.session_state.creneaux_disponibles is not None and len(st.session_state.creneaux_disponibles) > 0:
    #     demander_selection("creneaux_disponibles", st.session_state.creneaux_disponibles.index[0])

# Met à jour les données calculées d'une ligne
def bd_maj_donnees_calculees_row(idx, full=True):
    df = st.session_state.get("df", None)
    if df is None:
        return
    if idx not in df.index:
        return
    try:
        if len(df) > 0:
            debut = heure_parse(df.loc[idx, "Debut"])
            duree = duree_parse(df.loc[idx, "Duree"])
            
            df.at[idx, "Debut_dt"] = debut
            df.at[idx, "Duree_dt"] = duree

            fin = calculer_fin_row(df.loc[idx])
            df.at[idx, "Fin"] = fin

            if full:
                df = st.session_state.get("activites_programmees", None)
                if df is not None and idx in df.index:
                    df.at[idx, "Debut_dt"] = debut
                    df.at[idx, "Duree_dt"] = duree
                    df.at[idx, "Fin"] = fin

                df = st.session_state.get("activites_programmees_df_display", None)
                if df is not None and idx in df.index:
                    df.at[idx, "Fin"] = fin
                    st.session_state.activites_programmees_df_display_copy = st.session_state.activites_programmees_df_display.copy()
                
                df = st.session_state.get("activites_non_programmees", None)
                if df is not None and idx in df.index:
                    df.at[idx, "Debut_dt"] = debut
                    df.at[idx, "Duree_dt"] = duree
                    df.at[idx, "Fin"] = fin
                df = st.session_state.get("activites_non_programmees_df_display", None)
                if df is not None and idx in df.index:
                    df.at[idx, "Fin"] = fin
                    st.session_state.activites_programmees_df_display_copy = st.session_state.activites_programmees_df_display.copy()

    except:
        pass        

# Met à jour les données calculées sur st.session_state
# A utiliser conjointement avec bd_maj_activites_programmees, bd_maj_activites_non_programmees et , bd_maj_creneaux_disponibles
# pour reconstituer le contexte apres chargement de nouvelles données via fichier ou google sheet
def bd_maj_donnees_calculees():
    df = st.session_state.get("df", None)
    if df is None:
        return
    try:
        if len(df) > 0:
            df["Debut_dt"] = df["Debut"].apply(heure_parse)
            df["Duree_dt"] = df["Duree"].apply(duree_parse)
            df["Fin"] = df.apply(calculer_fin_row, axis=1)    
            df["Hyperlien"] = get_liens_activites()   
    except:
        pass        

# Met à jour les variables d'état relatives aux activités non programmées
# @chrono
def bd_maj_activites_non_programmees(maj_options_date=True):
    if st.session_state.get("df", None) is None:
        return
    activites_non_programmees = get_activites_non_programmees(st.session_state.df)
    st.session_state.activites_non_programmees = activites_non_programmees
    df_display = bd_creer_df_display_activites_non_programmees(activites_non_programmees, maj_options_date)
    st.session_state.activites_non_programmees_df_display = df_display
    st.session_state.activites_non_programmees_df_display_copy = df_display.copy()

def bd_modifier_cellule(idx, col, val, section_critique=False):

    if section_critique:
        st.session_state.setdefault("bd_modifier_cellule_cmd", 
            {
                "idx": idx,
                "col": col,
                "val": val
            }
        )

    tracer.log(f"Debut {idx} {col} {val}", types=["gen"])

    df = st.session_state.df
    oldval = df.loc[idx, col]
    modifier_df_cell(df, idx, col, val)
    if col == "Date":
        jour = safe_int(val)

        # Programmation d'une activité non programmée
        if (pd.isna(oldval) or oldval == "") and not (pd.isna(val) or val == ""):
            if jour is not None:
                bd_programmer(idx, jour)
        
        # Déprogrammation d'une activité programmée
        elif not (pd.isna(oldval) or oldval == "") and (pd.isna(val) or val == ""):
            jour = safe_int(oldval)
            if jour is not None:
                bd_deprogrammer(idx, jour)            

        # Reprogrammation d'une activité programmée    
        elif est_activite_programmee(df.loc[idx]):
            modifier_df_cell(st.session_state.activites_programmees, idx, col, val)
            modifier_df_display_cell(st.session_state.activites_programmees_df_display, idx, df_display_col_nom(col), str(val))
            modifier_df_display_cell(st.session_state.activites_programmees_df_display, idx, "__jour", int(val) if safe_int(val) is not None else None)
            st.session_state.activites_programmees = st.session_state.activites_programmees.sort_values(by=["Date", "Debut"], ascending=[True, True])
            st.session_state.activites_programmees_df_display = st.session_state.activites_programmees_df_display.sort_values(by=["Date", "Début"], ascending=[True, True])

            maj_options_date(df, st.session_state.activites_programmees, st.session_state.activites_programmees_df_display, safe_int(oldval))
            maj_options_date(df, st.session_state.activites_programmees, st.session_state.activites_non_programmees_df_display, safe_int(oldval))
            maj_options_date(df, st.session_state.activites_programmees, st.session_state.activites_programmees_df_display, jour)
            maj_options_date(df, st.session_state.activites_programmees, st.session_state.activites_non_programmees_df_display, jour)

            st.session_state.activites_programmees_df_display_copy = st.session_state.activites_programmees_df_display.copy()
            bd_maj_creneaux_disponibles()
    
    else:
        
        if col == "Activite":
            lnk = df.loc[idx, "Hyperlien"]
            if isinstance(lnk, str) and isinstance(oldval, str) and isinstance(val, str):
                ancien_nom_dans_lnk = oldval.replace(' ', '+')
                if ancien_nom_dans_lnk in lnk:
                    lnk = lnk.replace(ancien_nom_dans_lnk, val.replace(' ', '+'))
                    modifier_df_cell(df, idx, "Hyperlien", lnk)
                    if est_activite_programmee(df.loc[idx]):
                        modifier_df_cell(st.session_state.activites_programmees, idx, "Hyperlien", lnk)
                        modifier_df_display_cell(st.session_state.activites_programmees_df_display, idx, "Hyperlien", lnk)
                    elif est_activite_non_programmee(df.loc[idx]):
                        modifier_df_cell(st.session_state.activites_non_programmees, idx, "Hyperlien", lnk)
                        modifier_df_display_cell(st.session_state.activites_non_programmees_df_display, idx, "Hyperlien", lnk)
        elif col == "Lieu":
            if est_activite_programmee(df.loc[idx]):
                set_addr_cols(st.session_state.activites_programmees_df_display, idx, val)
            elif est_activite_non_programmee(df.loc[idx]):
                set_addr_cols(st.session_state.activites_non_programmees_df_display, idx, val)
    
        if est_activite_programmee(df.loc[idx]):
            modifier_df_cell(st.session_state.activites_programmees, idx, col, val)
            modifier_df_display_cell(st.session_state.activites_programmees_df_display, idx, df_display_col_nom(col), val)
            if col == "Debut":
                st.session_state.activites_programmees = st.session_state.activites_programmees.sort_values(by=["Date", "Debut"], ascending=[True, True])
                st.session_state.activites_programmees_df_display = st.session_state.activites_programmees_df_display.sort_values(by=["Date", "Début"], ascending=[True, True])
            elif col == "Reserve":
                if st.session_state.activites_programmees_df_display.loc[idx]["__index"] == idx:
                    non_reserve = str(st.session_state.activites_programmees_df_display.loc[idx][df_display_col_nom("Reserve")].strip().lower()) != "oui"
                    modifier_df_cell(st.session_state.activites_programmees_df_display, idx, "__non_reserve", non_reserve)
                else:
                    st.session_state.activites_programmees_df_display["__non_reserve"] = st.session_state.activites_programmees_df_display["Reserve"].astype(str).str.strip().str.lower() != "oui"
            st.session_state.activites_programmees_df_display_copy = st.session_state.activites_programmees_df_display.copy()
        
        elif est_activite_non_programmee(df.loc[idx]):
            modifier_df_cell(st.session_state.activites_non_programmees, idx, col, val)
            modifier_df_display_cell(st.session_state.activites_non_programmees_df_display, idx, df_display_col_nom(col), val)
            if col == "Debut":
                st.session_state.activites_non_programmees = st.session_state.activites_non_programmees.sort_values(by=["Date", "Debut"], ascending=[True, True])
                st.session_state.activites_non_programmees_df_display = st.session_state.activites_non_programmees_df_display.sort_values(by=["Date", "Début"], ascending=[True, True])
            st.session_state.activites_non_programmees_df_display_copy = st.session_state.activites_non_programmees_df_display.copy()
        
        if col in ["Debut", "Duree"]:
            bd_maj_donnees_calculees_row(idx)
        
        if col in ["Debut", "Duree", "Relache"] or est_activite_programmee(df.loc[idx]) and col == "Reserve":
            if pd.notna(df.loc[idx]["Date"]):
                jour = df.loc[idx]["Date"]
                maj_options_date(df, st.session_state.activites_programmees, st.session_state.activites_programmees_df_display, jour)
                maj_options_date(df, st.session_state.activites_programmees, st.session_state.activites_non_programmees_df_display, jour)

        bd_maj_creneaux_disponibles()

    tracer.log(f"Fin {idx} {col} {val}", types=["gen"])
    
    if section_critique:
        del st.session_state["bd_modifier_cellule_cmd"]

# Met à jour le contexte pour activité dont la Date de programmation passe de None à jour
# @chrono
def bd_programmer(idx, jour=None):
    
    if "df" not in st.session_state:
        return
    
    if idx not in st.session_state.df.index:
        return
    
    if "activites_programmees" not in st.session_state:
        return
    
    if "activites_non_programmees" not in st.session_state:
        return

    if jour is None:
        return

    if idx in st.session_state.activites_non_programmees.index:

        row = st.session_state.activites_non_programmees.loc[[idx]]
        row.at[idx, "Date"] = jour
        st.session_state.activites_non_programmees.drop(index=idx, inplace=True)
        st.session_state.activites_programmees = pd.concat([st.session_state.activites_programmees, row]).sort_values(by=["Date", "Debut_dt"], ascending=[True, True])

        row = st.session_state.activites_non_programmees_df_display.loc[[idx]]
        row.at[idx, "Date"] = str(jour)
        row["__jour"] = row["Date"].apply(lambda x: int(str(int(float(x)))[-2:]) if pd.notna(x) else None)
        row["__non_reserve"] = row["Réservé"].astype(str).str.strip().str.lower() != "oui"
        st.session_state.activites_non_programmees_df_display.drop(index=idx, inplace=True)
        st.session_state.activites_programmees_df_display = pd.concat([st.session_state.activites_programmees_df_display, row]).sort_values(by=["Date", "Début"], ascending=[True, True])

        maj_options_date(st.session_state.df, st.session_state.activites_programmees, st.session_state.activites_programmees_df_display, jour)
        maj_options_date(st.session_state.df, st.session_state.activites_programmees, st.session_state.activites_non_programmees_df_display, jour)

        st.session_state.activites_programmees_df_display_copy = st.session_state.activites_programmees_df_display.copy()
        st.session_state.activites_non_programmees_df_display_copy = st.session_state.activites_non_programmees_df_display.copy()
        
        bd_maj_creneaux_disponibles()

# Affichage des contrôles d'édition
def afficher_controles_edition():
    if st.button(LABEL_BOUTON_DEFAIRE, 
        disabled=not st.session_state.get("historique_undo"), 
        use_container_width=CENTRER_BOUTONS, 
        key="undo_btn") and st.session_state.historique_undo:
        undo.undo()
    if st.button(LABEL_BOUTON_REFAIRE, 
        disabled=not st.session_state.get("historique_redo"), 
        use_container_width=CENTRER_BOUTONS, 
        key="redo_btn") and st.session_state.historique_redo:
        undo.redo()

# Affichage des choix généraux
def afficher_infos_generales():

    df = st.session_state.get("df")
    if df is None:
        return
    
    with st.expander("ℹ️ Infos"):
        # Vérification de l'
        afficher_aide()        
        
        # Vérification de cohérence des informations du df
        verifier_coherence(df) 

        # Vérification de cohérence des informations du df
        afficher_periode_programmation()

        # Affichage des paramètres
        afficher_parametres()

# Affiche le nom d'activité
def afficher_nom_activite(df, index_df, nom_activite=None, afficher_label=True):

    # afficher_label = False if not st.session_state.sidebar_menus else afficher_label
    
    if index_df is not None:
        row = df.loc[index_df]
        if nom_activite == None:
            nom_activite = row["Activite"].strip()
        if est_activite_programmee(row):
            label_activite = f"Le {int(row["Date"])} de {row["Debut"]} à {row["Fin"]}"
            if est_activite_reserve(row):
                st_info_avec_label(label_activite, nom_activite, afficher_label=afficher_label, color="red")
            else:
                st_info_avec_label(label_activite, nom_activite, afficher_label=afficher_label)
        else:
            label_activite = f"De {row["Debut"]} à {row["Fin"]}"
            st_info_avec_label(label_activite, nom_activite, afficher_label=afficher_label)
    else:
        if nom_activite == None:
            nom_activite = ""
        label_activite = "De ..h.. à ..h.."
        st_info_avec_label(label_activite, nom_activite, afficher_label=afficher_label)
    
# Affiche un nom d'activité clickable qui switche le menu d'activités alternatif (sert en mode MODE_ACTIVITE_UNIQUE)
def afficher_nom_activite_clickable(df, index_df, nom_activite=None, afficher_label=True):

    hit = False
    key = "nom_activite_clickable" # if st.session_state.sidebar_menus else None
    # afficher_label = False if not st.session_state.sidebar_menus else afficher_label
    activite_programmee = False

    if index_df is not None:
        row = df.loc[index_df]
        activite_reservee = est_activite_reserve(row)
        activite_programmee = est_activite_programmee(row)

        # Injecte le CSS permettent de styler le primary button affiché par st_info_avec_label avec param key 
        injecter_css_pour_primary_buttons("error" if activite_reservee else "info")

        if nom_activite == None:
            nom_activite = row["Activite"].strip()
        if est_activite_programmee(row):
            label_activite = f"Le {int(row["Date"])} de {row["Debut"]} à {row["Fin"]}"
            if activite_reservee:
                hit = st_info_avec_label(label_activite, nom_activite, key, afficher_label=afficher_label, color="red")
            else:
                hit = st_info_avec_label(label_activite, nom_activite, key, afficher_label=afficher_label)
        else:
            label_activite = f"De {row["Debut"]} à {row["Fin"]}"
            hit = st_info_avec_label(label_activite, nom_activite, key, afficher_label=afficher_label)
    else:
        if nom_activite == None:
            nom_activite = ""
        label_activite = "De ..h.. à ..h.."

        # Injecte le CSS permettent de styler le primary button affiché par st_info_avec_label avec param key 
        injecter_css_pour_primary_buttons("info")
        hit = st_info_avec_label(label_activite, nom_activite, key, afficher_label=afficher_label)
    
    if hit:
        if activite_programmee:
            new_index_df = st.session_state.activites_non_programmees_sel_request["sel"]["id"] #_selected_row
            if new_index_df is not None:
                st.session_state.menu_activites = {
                    "menu": "menu_activites_non_programmees",
                    "index_df": new_index_df
                }
                demander_selection("activites_non_programmees", new_index_df, deselect="activites_programmees")
        else:
            new_index_df = st.session_state.activites_programmees_sel_request["sel"]["id"] #_selected_row
            if new_index_df is not None:
                st.session_state.menu_activites = {
                    "menu": "menu_activites_programmees",
                    "index_df": new_index_df
                }
                demander_selection("activites_programmees", new_index_df, deselect="activites_non_programmees")
        st.rerun()

# Affichage du status du GS Worker
def afficher_worker_status():
    st.sidebar.subheader("Google ")
    s = wk.get_sync_status()
    col1, col2 = st.sidebar.columns(2)
    col1.metric("Alive", "✅" if s.get("alive") else "❌")
    col2.metric("Pending", s.get("pending", 0))
    st.sidebar.caption(f"Last OK: {s.get('last_ok')}")
    if s.get("last_err"):
        st.sidebar.error(s["last_err"])
    if st.sidebar.button("Ping worker"):
        ok = wk.enqueue_noop()
        st.sidebar.write("Enqueue:", "ok" if ok else "queue None")

    # ###################################################################################################################
    # A BANNIR ABSOLUMENT CAR streamlit_autorefresh INTERROMPT TOUT TRAITEMENT QUI N'EST PAS MIS EN SECTION CRITIQUE ET
    # POUR CEUX QUI LE SONT EMPECHE QU'ILS SE TERMINENT SI LA PLUS LONGUE DE LEURS ETAPES EST PLUS LONGUE QUE LE TIMEOUT 
    # D'AUTOREFRESH, D'OU FIGEAGE D'UI ET EVENTUELLE PERTE DE COHERENCE DU CONTEXTE;
    # ####################################################################################################################
    # # Auto-refresh tant qu’il y a du travail
    # from streamlit_autorefresh import st_autorefresh
    # if s.get("pending", 0) > 0:
    #     st_autorefresh(interval=1000, key="gsync_poll")

# Affichage du status du GS Worker (version discrète)
def afficher_worker_status_discret(with_pending=True):
    s = wk.get_sync_status() if "wk" in globals() else {}
    alive   = bool(s.get("alive"))
    pending = int(s.get("pending", 0))
    last_ok = s.get("last_ok")
    last_err = s.get("last_err")

    color = "#16a34a" if alive else "#ef4444"   # vert / rouge
    title = "OK" if alive else "Hors ligne"

    if with_pending:
        if pending > 0:
            title = f"Sync en cours… ({pending})"
        if last_err:
            title = f"Erreur: {last_err}"

        html = f"""
        <div style="
            display:flex;align-items:center;gap:.5rem;
            font-size:0.90rem; line-height:1.2; margin:.25rem 0 .25rem .1rem;">
        <span title="{title}" style="color:{color};font-size:1rem;">●</span>
        <span style="opacity:.9;">Google&nbsp;Sheet</span>
        {"<span style='margin-left:auto;opacity:.6;font-variant-numeric:tabular-nums;'>"+str(pending)+"</span>" if pending>0 else ""}
        </div>
        """
    else:
        html = f"""
        <div style="
            display:flex;align-items:center;gap:.5rem;
            font-size:0.90rem; line-height:1.2; margin:.25rem 0 .25rem .1rem;">
        <span title="{title}" style="color:{color};font-size:1rem;">●</span>
        <span style="opacity:.9;">Google&nbsp;Sheet</span>
        </div>
        """
    st.sidebar.markdown(html, unsafe_allow_html=True)

    # ###################################################################################################################
    # A BANNIR ABSOLUMENT CAR streamlit_autorefresh INTERROMPT TOUT TRAITEMENT QUI N'EST PAS MIS EN SECTION CRITIQUE ET
    # POUR CEUX QUI LE SONT EMPECHE QU'ILS SE TERMINENT SI LA PLUS LONGUE DE LEURS ETAPES EST PLUS LONGUE QUE LE TIMEOUT 
    # D'AUTOREFRESH, D'OU FIGEAGE D'UI ET EVENTUELLE PERTE DE COHERENCE DU CONTEXTE;
    # ####################################################################################################################
    # # Auto-refresh UNIQUEMENT si des tâches sont en attente
    # if pending > 0:
    #     try:
    #         from streamlit_autorefresh import st_autorefresh
    #         st_autorefresh(interval=1000, key="gsync_poll")
    #     except Exception:
    #         pass    

# Affichage de la la sidebar min avec menus fichier et edition 
# (le reste est affiché dans d'affichage de données en fonction du contexte)
def afficher_sidebar():

    st.sidebar.title("Menu principal")

    with st.sidebar.expander("Fichier"):
        creer_nouveau_contexte()
        charger_contexte_depuis_fichier()
        sauvegarder_contexte()

    with st.sidebar.expander("Edition"):
        afficher_controles_edition()

# Affichage du menu activité de la sidebar
def afficher_menu_activite():

    df = st.session_state.get("df")
    if df is None:
        return
    
    if est_contexte_valide():
        with st.sidebar.expander("Activités", expanded=True):
            if "menu_activites" in st.session_state and isinstance(st.session_state.menu_activites, dict):
                if st.session_state.menu_activites["menu"] == "menu_activites_programmees":
                    menu_activites_programmees(
                        st.session_state.menu_activites["index_df"]
                    )

                elif st.session_state.menu_activites["menu"] == "menu_activites_non_programmees":
                    menu_activites_non_programmees(
                        st.session_state.menu_activites["index_df"]
                    )
        
        # Désactivation des flags de forçage de menu activités
        if st.session_state.forcer_menu_activites_programmees and st.session_state.menu_activites["menu"] == "menu_activites_programmees":
            st.session_state.forcer_menu_activites_programmees = False
        if st.session_state.forcer_menu_activites_non_programmees and st.session_state.menu_activites["menu"] == "menu_activites_non_programmees":
            st.session_state.forcer_menu_activites_non_programmees = False
    
    afficher_worker_status_discret()

# Gestion des sections critiques de traitement.
# Ces sections critiques sont utilisées notamment pour gérer la modification de cellules depuis les grilles.
# Dans ce cas en effet la modification de cellule depuis la grille est validée par un click row 
# qui peut entraîner une interruption du script python et donc une incohérence de contexte.
# Le mécanisme de section critique permet une relance automatique du traitement jusqu'à complétion 
# en cas d'interruption par un rerun Streamlit : une commande est enregistrée dans st.session_state 
# et est automatiquement relancée en début de rerun par la fonction ci-dessous tant qu'elle n'est pas terminée.
def traiter_sections_critiques():

    cmd = st.session_state.get("bd_maj_contexte_cmd")
    if cmd:
        bd_maj_contexte(cmd["maj_donnees_calculees"], cmd["maj_options_date"])
    
    cmd = st.session_state.get("bd_modifier_cellule_cmd")
    if cmd:
        bd_modifier_cellule(cmd["idx"], cmd["col"], cmd["val"])
    
    cmd = st.session_state.get("activites_programmees_modifier_cellule_cmd")
    if cmd:
        activites_programmees_modifier_cellule(cmd["idx"], cmd["col"], cmd["val"])
    
    cmd = st.session_state.get("activites_programmees_deprogrammer_cmd")
    if cmd:
        activites_programmees_deprogrammer(cmd["idx"])
    
    cmd = st.session_state.get("activites_programmees_reprogrammer_cmd")
    if cmd:
        activites_programmees_reprogrammer(cmd["idx"], cmd["jour"])
    
    cmd = st.session_state.get("activites_non_programmees_modifier_cellule_cmd")
    if cmd:
        activites_non_programmees_modifier_cellule(cmd["idx"], cmd["col"], cmd["val"])
    
    cmd = st.session_state.get("activites_non_programmees_programmer_cmd")
    if cmd:
        activites_non_programmees_programmer(cmd["idx"], cmd["jour"])

# Permet d'éviter le blocage de l'UI au retour d'appel d'une page web dans le meme onglet (same tab)
@st.cache_resource
def inject_ios_soft_revive_global():
    st.markdown("""
        <script>
        (function(){
        var isIOS = /iPad|iPhone|iPod/.test(navigator.userAgent);

        function cameFromBackForward(){
            try {
            var nav = performance.getEntriesByType && performance.getEntriesByType('navigation');
            return !!(nav && nav[0] && nav[0].type === 'back_forward');
            } catch(e){ return false; }
        }

        function softRevive(){
            try { document.activeElement && document.activeElement.blur && document.activeElement.blur(); } catch(e){}
            try { window.dispatchEvent(new Event('focus')); } catch(e){}
            try { window.dispatchEvent(new Event('resize')); } catch(e){}
            // petit “reflow” pour réveiller WebKit
            try {
            var html = document.documentElement;
            var prev = html.style.webkitTransform;
            html.style.webkitTransform = 'translateZ(0)';
            void html.offsetHeight;
            html.style.webkitTransform = prev || '';
            } catch(e){}
        }

        window.addEventListener('pageshow', function(e){
            if (!isIOS) return;
            if (e.persisted || cameFromBackForward()){
            // réveille la page parent
            softRevive();
            // Laisse les iframes (grilles) gérer leur propre refresh (voir 2B)
            }
        }, false);
        })();
        </script>
    """, unsafe_allow_html=True)
    return True

# Initialisation de la page HTML
def initialiser_page():

    # Injecte le JS qui permet d'éviter un figeage au retour d'appel d'une page web dans le meme onglet (same tab)
    inject_ios_soft_revive_global()

# Trace le début d'un rerun
def tracer_rerun():
    st.session_state.setdefault("main_counter", 0)
    st.session_state.main_counter += 1
    tracer.log(f"____________MAIN {st.session_state.main_counter}______________", types=["gen","main"])

# Opérations à ne faire qu'une seule fois au boot de l'appli
@st.cache_resource
def app_boot():

    cold_start = not sql.db_exists()
    tracer.log(f"Cold Start {cold_start}", types=["main"])

    # DEBUG ONLY - Reset DB
    # with sqlite3.connect(DB_PATH) as con:
    #     cur = con.cursor()
    #     # supprime les tables si elles existent
    #     cur.executescript("""
    #         DROP TABLE IF EXISTS df_principal;
    #         DROP TABLE IF EXISTS meta;
    #         DROP TABLE IF EXISTS carnet;
    #     """)
    #     con.commit()
    # DEBUG ONLY

    sql.init_db()                           # Crée les tables si besoin
    if cold_start and WITH_GOOGLE_SHEET:    # Hydratation des tables avec les données Google Sheet en cas de cold start et si Google Sheet est utilisé
        charger_contexte_depuis_gsheet()
        tracer.log(f"Type de ca {type(st.session_state.ca)}", types=["cold start"])
        sql.sauvegarder_contexte(enqueue=False)

def main():

    # Affichage de la version de streamlit-aggrid
    # import pkg_resources
    # version = pkg_resources.get_distribution("streamlit-aggrid").version
    # st.write("Version streamlit-aggrid :", version)

    # Trace le début d'un rerun
    tracer_rerun()
  
    # Connexion à la Google Sheet et lancement du GS Worker chargé de la sauvegarde Google Sheet en temps masqué (seulement si WITH_GOOGLE_SHEET est True)
    if WITH_GOOGLE_SHEET:
        gs.connect()
        wk.ensure_worker_alive()

    # Opérations à ne faire qu'une seule fois au démarrage appli 
    app_boot()

    # Chargement de contexte depuis SqLite en charge de la persistence à chaud 
    # (à faire à chaque rerun pour tenir compte des reinit de st.session_state en cours de session)
    charger_contexte_depuis_sql()

    # Gestion des sections critiques
    traiter_sections_critiques()

    # Configuration de la page HTML
    initialiser_page()

    # Affichage du titre
    afficher_titre("Planificateur Avignon Off")

    # Affichage de la sidebar
    afficher_sidebar()

   # Si le contexte est valide, on le traite
    if est_contexte_valide():

        # Affichage des infos générales
        afficher_infos_generales()
        
        # Affichage des activités programmées
        afficher_activites_programmees()

        # Affichage des activités non programmées
        afficher_activites_non_programmees()

        # Affichage des créneaux disponibles et des activités programmables
        afficher_creneaux_disponibles()      

        # # Affichage du menu activité de la sidebar
        afficher_menu_activite()
    else:
        message = st.session_state.get("contexte_invalide_message")
        if message is not None:
            st.error(st.session_state.get("contexte_invalide_message"))

if __name__ == "__main__":
    main()
