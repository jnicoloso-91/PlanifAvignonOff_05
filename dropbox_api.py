###############
# API DropBox #
###############

import streamlit as st
import dropbox
from io import BytesIO
from openpyxl import load_workbook, Workbook

def get_dropbox_client() -> dropbox.Dropbox:
    """
    Retourne un client Dropbox qui renouvelle automatiquement
    l'access token à partir du refresh_token.
    """
    cfg = st.secrets["dropbox"]
    return dropbox.Dropbox(
        app_key=cfg["app_key"],
        app_secret=cfg["app_secret"],
        oauth2_refresh_token=cfg["refresh_token"]
    )

# Sauvegarde sur Dropbox le fichier Excel de l'utilisateur 
# Cette sauvegarde permet de garder une trace de la mise en page du fichier utilisateur
def upload_excel_to_dropbox(file_bytes, filename, dropbox_path="/uploads/"):
    dbx = get_dropbox_client()
    full_path = f"{dropbox_path}{filename}"

    try:
        dbx.files_upload(file_bytes, full_path, mode=dropbox.files.WriteMode("overwrite"))
        # st.success(f"✅ Fichier '{filename}' uploadé dans Dropbox à {full_path}")
        return full_path
    except Exception as e:
        print(f"upload_excel_to_dropbox : {e}")
        return ""

# Renvoie le fichier Excel de l'utilisateur sauvegardé sur DropBox
# Cette sauvegarde permet de garder une trace de la mise en page du fichier utilisateur
def download_excel_from_dropbox(file_path):
    dbx = get_dropbox_client()
    try:
        metadata, res = dbx.files_download(file_path)
        file_bytes = BytesIO(res.content)
        return load_workbook(file_bytes)
    except Exception as e:
        # st.error(f"❌ Erreur lors du téléchargement depuis Dropbox : {e}")
        return Workbook()

